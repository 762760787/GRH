#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Application de Gestion des Ressources Humaines - Mairie de Ngoundiane
Développée en Python avec Tkinter pour une interface desktop moderne
Auteur: Assistant IA - Développeur Senior Desktop (25 ans d'expérience)
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import hashlib
import os
import shutil
from datetime import datetime, timedelta
import calendar
from PIL import Image, ImageTk
import subprocess
import platform
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

class HRManagementApp:
    def __init__(self):
        """Initialisation de l'application"""
        self.root = tk.Tk()
        self.root.title("Gestion RH - Mairie de Ngoundiane")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 800)
        
        # Variables globales
        self.current_user = None
        self.current_employee_id = None
        self.db_path = "hr_database.db"
        self.documents_folder = "documents"
        self.photos_folder = "photos"
        self.courriers_folder = "courriers_files"  # Nouveau dossier pour les fichiers de courriers
        
        # Créer les dossiers nécessaires
        os.makedirs(self.documents_folder, exist_ok=True)
        os.makedirs(self.photos_folder, exist_ok=True)
        os.makedirs(self.courriers_folder, exist_ok=True)  # Créer le dossier courriers
        
        # Palette de couleurs verte professionnelle
        self.colors = {
            'primary_green': '#2E7D32',      # Vert principal
            'light_green': '#4CAF50',        # Vert clair
            'dark_green': '#1B5E20',         # Vert foncé
            'accent_green': '#81C784',       # Vert accent
            'background': '#F8F9FA',         # Arrière-plan clair
            'white': '#FFFFFF',              # Blanc
            'light_gray': '#E8F5E8',         # Gris vert clair
            'text_dark': '#2C3E50',          # Texte foncé
            'text_light': '#7F8C8D',         # Texte clair
            'error': '#E74C3C',              # Rouge erreur
            'warning': '#F39C12',            # Orange avertissement
            'success': '#27AE60'             # Vert succès
        }
        
        # Configuration du style
        self.setup_styles()
        
        # Initialisation de la base de données
        self.init_database()
        
        # Démarrage avec l'écran de connexion
        self.show_login_screen()
        
    def setup_styles(self):
        """Configuration des styles visuels modernes"""
        style = ttk.Style()
        
        # Configuration du thème
        style.theme_use('clam')
        
        # Style pour les boutons principaux
        style.configure('Primary.TButton',
                       background=self.colors['primary_green'],
                       foreground='white',
                       borderwidth=0,
                       focuscolor='none',
                       padding=(20, 10))
        
        style.map('Primary.TButton',
                 background=[('active', self.colors['light_green']),
                           ('pressed', self.colors['dark_green'])])
        
        # Style pour les boutons secondaires
        style.configure('Secondary.TButton',
                       background=self.colors['accent_green'],
                       foreground=self.colors['text_dark'],
                       borderwidth=1,
                       focuscolor='none',
                       padding=(15, 8))
        
        # Style pour les labels de titre
        style.configure('Title.TLabel',
                       background=self.colors['background'],
                       foreground=self.colors['primary_green'],
                       font=('Segoe UI', 16, 'bold'))
        
        # Style pour les labels de section
        style.configure('Section.TLabel',
                       background=self.colors['background'],
                       foreground=self.colors['text_dark'],
                       font=('Segoe UI', 12, 'bold'))
        
        # Style pour le notebook (onglets)
        style.configure('Custom.TNotebook',
                       background=self.colors['background'],
                       borderwidth=0)
        
        style.configure('Custom.TNotebook.Tab',
                       background=self.colors['light_gray'],
                       foreground=self.colors['text_dark'],
                       padding=(20, 10),
                       font=('Segoe UI', 10, 'bold'))
        
        style.map('Custom.TNotebook.Tab',
                 background=[('selected', self.colors['primary_green']),
                           ('active', self.colors['accent_green'])],
                 foreground=[('selected', 'white')])
        
        # Style pour les treeviews
        style.configure('Custom.Treeview',
                       background='white',
                       foreground=self.colors['text_dark'],
                       rowheight=30,
                       fieldbackground='white',
                       font=('Segoe UI', 10))
        
        style.configure('Custom.Treeview.Heading',
                       background=self.colors['primary_green'],
                       foreground='white',
                       font=('Segoe UI', 10, 'bold'))
        
    def init_database(self):
        """Initialisation de la base de données SQLite"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Table des utilisateurs
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table des employés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                matricule TEXT UNIQUE NOT NULL,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                birth_place TEXT,
                address TEXT,
                phone TEXT,
                email TEXT,
                marital_status TEXT,
                dependents INTEGER DEFAULT 0,
                social_security TEXT,
                bank_details TEXT,
                hire_date TEXT,
                contract_type TEXT,
                contract_start TEXT,
                contract_end TEXT,
                department TEXT,
                job_title TEXT,
                status TEXT DEFAULT 'Active',
                photo_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table de l'historique de carrière
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS career_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                act_number TEXT,
                nature TEXT,
                subject TEXT,
                act_date TEXT,
                effective_date TEXT,
                document_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id)
            )
        ''')
        
        # Table des documents
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                category TEXT,
                name TEXT,
                file_path TEXT,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id)
            )
        ''')
        
        # Table des types de congés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leave_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                days_per_year INTEGER DEFAULT 0,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table des congés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leaves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                leave_type_id INTEGER,
                start_date TEXT,
                end_date TEXT,
                days_count INTEGER,
                status TEXT DEFAULT 'Approved',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id),
                FOREIGN KEY (leave_type_id) REFERENCES leave_types (id)
            )
        ''')
        
        # Table des courriers - MISE À JOUR avec colonne pour fichiers
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS courriers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_ordre TEXT UNIQUE NOT NULL,
                type_courrier TEXT NOT NULL,
                nombre_pieces INTEGER DEFAULT 1,
                date_arrivee_expedition TEXT NOT NULL,
                expediteur_destinataire TEXT NOT NULL,
                objet TEXT NOT NULL,
                numero_archive TEXT,
                observation TEXT,
                file_path TEXT,
                date_creation TEXT DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT
            )
        ''')
        # AJOUTEZ CE BLOC DE CODE ICI
        # Ajouter les colonnes CNI et nationalité si elles n'existent pas
        try:
            cursor.execute('ALTER TABLE employees ADD COLUMN cni TEXT')
            cursor.execute('ALTER TABLE employees ADD COLUMN nationalite TEXT')
        except sqlite3.OperationalError:
            # Les colonnes existent déjà
            pass
        # Ajouter la colonne file_path si elle n'existe pas déjà
        try:
            cursor.execute('ALTER TABLE courriers ADD COLUMN file_path TEXT')
        except sqlite3.OperationalError:
            # La colonne existe déjà
            pass
        
        # Insérer des utilisateurs par défaut
        try:
            admin_hash = hashlib.sha256('admin'.encode()).hexdigest()
            user_hash = hashlib.sha256('user'.encode()).hexdigest()
            
            cursor.execute('INSERT OR IGNORE INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                          ('admin', admin_hash, 'admin'))
            cursor.execute('INSERT OR IGNORE INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                          ('user', user_hash, 'user'))
        except sqlite3.IntegrityError:
            pass
        
        # Insérer des types de congés par défaut
        default_leave_types = [
            ('Congé Annuel', 30, 'Congé annuel réglementaire'),
            ('Congé Maladie', 0, 'Congé pour maladie'),
            ('Congé Maternité', 0, 'Congé de maternité'),
            ('Congé Paternité', 0, 'Congé de paternité'),
            ('Permission Exceptionnelle', 0, 'Permission pour événements familiaux')
        ]
        
        for leave_type in default_leave_types:
            cursor.execute('INSERT OR IGNORE INTO leave_types (name, days_per_year, description) VALUES (?, ?, ?)',
                          leave_type)
        
        conn.commit()
        conn.close()
        
    def show_login_screen(self):
        """Affichage de l'écran de connexion (design moderne)"""
        # Nettoyer la fenêtre
        for widget in self.root.winfo_children():
            widget.destroy()

        self.root.configure(bg=self.colors['background'])
        self.root.bind('<Return>', lambda e: self.login())

        # --- Conteneur Principal ---
        main_frame = tk.Frame(self.root, bg=self.colors['background'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Panneau Gauche (Branding) ---
        left_panel = tk.Frame(main_frame, bg=self.colors['primary_green'])
        left_panel.place(relx=0, rely=0, relwidth=0.4, relheight=1)

        # Titre et description dans le panneau gauche
        tk.Label(left_panel, text="Mairie de\nNgoundiane", font=('Segoe UI', 36, 'bold'), fg='white', bg=self.colors['primary_green']).pack(pady=(150, 20))
        tk.Label(left_panel, text="Système de Gestion\ndes Ressources Humaines", font=('Segoe UI', 18), fg=self.colors['light_gray'], bg=self.colors['primary_green']).pack(pady=10)
        tk.Label(left_panel, text="© 2024", font=('Segoe UI', 10), fg=self.colors['accent_green'], bg=self.colors['primary_green']).pack(side=tk.BOTTOM, pady=20)

        # --- Panneau Droit (Formulaire de Connexion) ---
        right_panel = tk.Frame(main_frame, bg=self.colors['background'])
        right_panel.place(relx=0.4, rely=0, relwidth=0.6, relheight=1)

        # Frame pour centrer le contenu
        login_container = tk.Frame(right_panel, bg=self.colors['background'])
        login_container.place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(login_container, text="Authentification", font=('Segoe UI', 28, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background']).pack(pady=(0, 40))

        # --- Champ Nom d'utilisateur ---
        user_frame = tk.Frame(login_container, bg=self.colors['background'])
        user_frame.pack(pady=10)
        tk.Label(user_frame, text="👤", font=('Segoe UI', 14), fg=self.colors['text_light'], bg=self.colors['background']).pack(side=tk.LEFT, padx=(0,10))
        self.username_entry = tk.Entry(user_frame, font=('Segoe UI', 14), width=25, relief='flat', bg=self.colors['light_gray'])
        self.username_entry.pack(side=tk.LEFT)
        self.username_entry.insert(0, "Nom d'utilisateur")
        self.username_entry.config(fg=self.colors['text_light'])

        # --- Champ Mot de passe ---
        pass_frame = tk.Frame(login_container, bg=self.colors['background'])
        pass_frame.pack(pady=10)
        tk.Label(pass_frame, text="🔑", font=('Segoe UI', 14), fg=self.colors['text_light'], bg=self.colors['background']).pack(side=tk.LEFT, padx=(0,10))
        self.password_entry = tk.Entry(pass_frame, font=('Segoe UI', 14), width=25, relief='flat', bg=self.colors['light_gray'])
        self.password_entry.pack(side=tk.LEFT)
        self.password_entry.insert(0, "Mot de passe")
        self.password_entry.config(fg=self.colors['text_light'])

        # --- Logique pour les placeholders ---
        def on_user_click(event):
            if self.username_entry.get() == "Nom d'utilisateur":
                self.username_entry.delete(0, tk.END)
                self.username_entry.config(fg=self.colors['text_dark'])
        def on_user_leave(event):
            if self.username_entry.get() == '':
                self.username_entry.insert(0, "Nom d'utilisateur")
                self.username_entry.config(fg=self.colors['text_light'])

        def on_pass_click(event):
            if self.password_entry.get() == "Mot de passe":
                self.password_entry.delete(0, tk.END)
                self.password_entry.config(fg=self.colors['text_dark'], show='*')
        def on_pass_leave(event):
            if self.password_entry.get() == '':
                self.password_entry.insert(0, "Mot de passe")
                self.password_entry.config(fg=self.colors['text_light'], show='')

        self.username_entry.bind('<FocusIn>', on_user_click)
        self.username_entry.bind('<FocusOut>', on_user_leave)
        self.password_entry.bind('<FocusIn>', on_pass_click)
        self.password_entry.bind('<FocusOut>', on_pass_leave)

        # --- Bouton de connexion ---
        login_btn = tk.Button(login_container,
                             text="Se Connecter",
                             font=('Segoe UI', 14, 'bold'),
                             bg=self.colors['primary_green'],
                             fg='white',
                             relief='flat',
                             bd=0,
                             padx=40,
                             pady=12,
                             cursor='hand2',
                             command=self.login)
        login_btn.pack(pady=40)

        # --- Animation du bouton au survol ---
        def on_enter(e):
            login_btn.config(bg=self.colors['light_green'])
        def on_leave(e):
            login_btn.config(bg=self.colors['primary_green'])

        login_btn.bind("<Enter>", on_enter)
        login_btn.bind("<Leave>", on_leave)

        # --- Info utilisateurs ---
        # tk.Label(login_container,
        #          text="Utilisateurs par défaut: admin/admin ou user/user",
        #          font=('Segoe UI', 9),
        #          fg=self.colors['text_light'],
        #          bg=self.colors['background']).pack()

        # Focus initial
        # self.root.after(100, lambda: self.root.focus_force())

    def login(self):
        """Gestion de la connexion utilisateur"""
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not username or not password:
            messagebox.showerror("Erreur", "Veuillez saisir le nom d'utilisateur et le mot de passe")
            return
            
        # Vérification dans la base de données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute('SELECT id, username, role FROM users WHERE username = ? AND password_hash = ?',
                      (username, password_hash))
        
        user = cursor.fetchone()
        conn.close()
        
        if user:
            self.current_user = {
                'id': user[0],
                'username': user[1],
                'role': user[2]
            }
            self.show_main_dashboard()
        else:
            messagebox.showerror("Erreur", "Nom d'utilisateur ou mot de passe incorrect")
            self.password_entry.delete(0, tk.END)
            
    def show_main_dashboard(self):
        """Affichage du tableau de bord principal"""
        # Nettoyer la fenêtre
        for widget in self.root.winfo_children():
            widget.destroy()
            
        self.root.configure(bg=self.colors['background'])
        
        # Frame principal
        main_container = tk.Frame(self.root, bg=self.colors['background'])
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Header avec titre et info utilisateur
        header_frame = tk.Frame(main_container, bg=self.colors['primary_green'], height=80)
        header_frame.pack(fill='x', pady=(0, 10))
        header_frame.pack_propagate(False)
        
        # Titre principal
        title_label = tk.Label(header_frame,
                              text="🏛️ Système de Gestion RH - Mairie de Ngoundiane",
                              font=('Segoe UI', 20, 'bold'),
                              fg='white',
                              bg=self.colors['primary_green'])
        title_label.pack(side='left', padx=20, pady=20)
        
        # Info utilisateur et déconnexion
        user_frame = tk.Frame(header_frame, bg=self.colors['primary_green'])
        user_frame.pack(side='right', padx=20, pady=20)
        
        user_label = tk.Label(user_frame,
                             text=f"👤 {self.current_user['username']} ({self.current_user['role']})",
                             font=('Segoe UI', 12),
                             fg='white',
                             bg=self.colors['primary_green'])
        user_label.pack(side='left', padx=(0, 15))
        
        logout_btn = tk.Button(user_frame,
                              text="Déconnexion",
                              font=('Segoe UI', 10),
                              bg=self.colors['error'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2',
                              command=self.logout)
        logout_btn.pack(side='right')
        
        # Container pour le contenu principal
        content_container = tk.Frame(main_container, bg=self.colors['background'])
        content_container.pack(fill='both', expand=True)
        
        # Sidebar pour la navigation
        sidebar = tk.Frame(content_container, bg=self.colors['white'], width=250)
        sidebar.pack(side='left', fill='y', padx=(0, 10))
        sidebar.pack_propagate(False)
        
        # Titre sidebar
        sidebar_title = tk.Label(sidebar,
                                text="📋 MODULES",
                                font=('Segoe UI', 14, 'bold'),
                                fg=self.colors['primary_green'],
                                bg=self.colors['white'])
        sidebar_title.pack(pady=(20, 10))
        
        # Boutons de navigation
        nav_buttons = [
            ("👥 Tableau de Bord", self.show_dashboard_content),
            ("📁 Gestion Employés", self.show_employees_module),
            ("🏖️ Gestion Congés", self.show_leaves_module),
            ("📮 Gestion Courriers", self.show_mail_module),
            ("📊 Rapports", self.show_reports_module),
            ("⚙️ Configuration", self.show_settings_module)
        ]
        
        self.nav_buttons = {}
        for text, command in nav_buttons:
            btn = tk.Button(sidebar,
                           text=text,
                           font=('Segoe UI', 11),
                           bg=self.colors['light_gray'],
                           fg=self.colors['text_dark'],
                           relief='flat',
                           bd=0,
                           padx=20,
                           pady=12,
                           width=25,
                           anchor='w',
                           cursor='hand2',
                           command=command)
            btn.pack(fill='x', padx=10, pady=2)
            self.nav_buttons[text] = btn
            
            # Effet hover
            def on_enter(e, button=btn):
                if button['bg'] != self.colors['primary_green']:
                    button.configure(bg=self.colors['accent_green'])
            def on_leave(e, button=btn):
                if button['bg'] != self.colors['primary_green']:
                    button.configure(bg=self.colors['light_gray'])
                    
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
        
        # Zone de contenu principal
        self.main_content = tk.Frame(content_container, bg=self.colors['background'])
        self.main_content.pack(side='right', fill='both', expand=True)
        
        # Afficher le contenu du tableau de bord par défaut
        self.show_dashboard_content()
        self.set_active_nav_button("👥 Tableau de Bord")
        
    def set_active_nav_button(self, active_text):
        """Met en évidence le bouton de navigation actif"""
        for text, btn in self.nav_buttons.items():
            if text == active_text:
                btn.configure(bg=self.colors['primary_green'], fg='white')
            else:
                btn.configure(bg=self.colors['light_gray'], fg=self.colors['text_dark'])
                
    def show_dashboard_content(self):
        """Affichage du contenu du tableau de bord - MISE À JOUR avec totaux corrects"""
        self.clear_main_content()
        self.set_active_nav_button("👥 Tableau de Bord")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📊 Tableau de Bord",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les statistiques
        stats_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        stats_frame.pack(fill='x', padx=20)
        
        # Récupérer les statistiques CORRIGÉES
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Total employés actifs
        cursor.execute('SELECT COUNT(*) FROM employees WHERE status = "Active"')
        total_employees = cursor.fetchone()[0]
        
        # Employés en congé aujourd'hui
        today = datetime.now().strftime('%d/%m/%Y')
        cursor.execute('''
            SELECT COUNT(DISTINCT e.id) FROM employees e
            JOIN leaves l ON e.id = l.employee_id
            WHERE l.start_date <= ? AND l.end_date >= ? AND l.status = "Approved"
        ''', (today, today))
        employees_on_leave = cursor.fetchone()[0]
        
        # Anniversaires ce mois
        current_month = datetime.now().strftime('%m')
        cursor.execute('''
            SELECT COUNT(*) FROM employees 
            WHERE substr(birth_date, 4, 2) = ? AND status = "Active"
        ''', (current_month,))
        birthdays_this_month = cursor.fetchone()[0]
        
        # Total courriers (tous types confondus)
        cursor.execute('SELECT COUNT(*) FROM courriers')
        total_courriers = cursor.fetchone()[0]
        
        # Total congés cette année
        current_year = str(datetime.now().year)
        cursor.execute('''
            SELECT COUNT(*) FROM leaves 
            WHERE strftime('%Y', date(substr(start_date, 7, 4) || '-' || substr(start_date, 4, 2) || '-' || substr(start_date, 1, 2))) = ?
        ''', (current_year,))
        total_leaves_year = cursor.fetchone()[0]
        
        conn.close()
        
        # Cartes de statistiques - SUPPRESSION de "Contrats à Renouveler"
        stats_data = [
            ("👥 Total Employés", total_employees, self.colors['primary_green']),
            ("🏖️ En Congé Aujourd'hui", employees_on_leave, self.colors['warning']),
            ("📮 Total Courriers", total_courriers, '#6f42c1'),
            ("🏖️ Congés Cette Année", total_leaves_year, self.colors['success']),
            ("🎂 Anniversaires ce Mois", birthdays_this_month, '#ff6b6b')
        ]
        
        for i, (title, value, color) in enumerate(stats_data):
            card = tk.Frame(stats_frame, bg='white', relief='solid', bd=1)
            card.grid(row=i//3, column=i%3, padx=10, pady=10, sticky='ew')
            
            # Titre de la carte
            card_title = tk.Label(card,
                                 text=title,
                                 font=('Segoe UI', 12, 'bold'),
                                 fg=self.colors['text_dark'],
                                 bg='white')
            card_title.pack(pady=(15, 5))
            
            # Valeur
            card_value = tk.Label(card,
                                 text=str(value),
                                 font=('Segoe UI', 24, 'bold'),
                                 fg=color,
                                 bg='white')
            card_value.pack(pady=(0, 15))
            
        # Configurer les colonnes pour qu'elles s'étendent uniformément
        for i in range(3):
            stats_frame.grid_columnconfigure(i, weight=1)
            
        # Section des alertes
        alerts_frame = tk.LabelFrame(self.main_content,
                                   text="🚨 Alertes et Notifications",
                                   font=('Segoe UI', 14, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'],
                                   padx=10,
                                   pady=10)
        alerts_frame.pack(fill='both', expand=True, padx=20, pady=(30, 20))
        
        # Liste des alertes
        alerts_text = tk.Text(alerts_frame,
                             font=('Segoe UI', 11),
                             bg='white',
                             fg=self.colors['text_dark'],
                             relief='flat',
                             wrap='word',
                             state='disabled')
        alerts_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Ajouter les alertes
        alerts_content = []
        
        if birthdays_this_month > 0:
            alerts_content.append(f"🎂 {birthdays_this_month} anniversaire(s) ce mois")
            
        if employees_on_leave > 0:
            alerts_content.append(f"🏖️ {employees_on_leave} employé(s) actuellement en congé")
            
        if total_courriers > 0:
            alerts_content.append(f"📮 {total_courriers} courrier(s) enregistré(s) au total")
            
        if not alerts_content:
            alerts_content.append("✅ Aucune alerte pour le moment")
            
        alerts_text.config(state='normal')
        for alert in alerts_content:
            alerts_text.insert(tk.END, alert + "\n\n")
        alerts_text.config(state='disabled')
        
    def show_employees_module(self):
        """Module de gestion des employés"""
        self.clear_main_content()
        self.set_active_nav_button("📁 Gestion Employés")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="👥 Gestion des Employés",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Barre d'outils
        toolbar = tk.Frame(self.main_content, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=(0, 10))
        
        # Boutons d'action
        add_btn = tk.Button(toolbar,
                           text="➕ Nouvel Employé",
                           font=('Segoe UI', 11, 'bold'),
                           bg=self.colors['primary_green'],
                           fg='white',
                           relief='flat',
                           bd=0,
                           padx=15,
                           pady=8,
                           cursor='hand2',
                           command=self.add_new_employee)
        add_btn.pack(side='left', padx=(0, 10))
        
        # Champ de recherche
        search_frame = tk.Frame(toolbar, bg=self.colors['background'])
        search_frame.pack(side='right')
        
        tk.Label(search_frame,
                text="🔍 Rechercher:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(side='left', padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_employees)
        search_entry = tk.Entry(search_frame,
                               textvariable=self.search_var,
                               font=('Segoe UI', 11),
                               width=25,
                               relief='solid',
                               bd=1)
        search_entry.pack(side='left')
        
        # Liste des employés
        list_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Treeview pour la liste
        columns = ('Photo', 'Matricule', 'Nom Complet', 'Poste', 'Département', 'Statut')
        self.employees_tree = ttk.Treeview(list_frame,
                                          columns=columns,
                                          show='headings',
                                          style='Custom.Treeview',
                                          height=15)
        
        # Configuration des colonnes
        self.employees_tree.heading('Photo', text='Photo')
        self.employees_tree.heading('Matricule', text='Matricule')
        self.employees_tree.heading('Nom Complet', text='Nom Complet')
        self.employees_tree.heading('Poste', text='Poste')
        self.employees_tree.heading('Département', text='Département')
        self.employees_tree.heading('Statut', text='Statut')
        
        self.employees_tree.column('Photo', width=80, anchor='center')
        self.employees_tree.column('Matricule', width=100, anchor='center')
        self.employees_tree.column('Nom Complet', width=200, anchor='w')
        self.employees_tree.column('Poste', width=150, anchor='w')
        self.employees_tree.column('Département', width=150, anchor='w')
        self.employees_tree.column('Statut', width=100, anchor='center')
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.employees_tree.yview)
        h_scrollbar = ttk.Scrollbar(list_frame, orient='horizontal', command=self.employees_tree.xview)
        self.employees_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Placement
        self.employees_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # Double-clic pour ouvrir le dossier employé
        self.employees_tree.bind('<Double-1>', self.open_employee_file)
        
        # Menu contextuel
        self.create_employee_context_menu()
        
        # Charger les employés
        self.load_employees()
        
    def create_employee_context_menu(self):
        """Créer le menu contextuel pour la liste des employés"""
        self.employee_context_menu = tk.Menu(self.root, tearoff=0)
        self.employee_context_menu.add_command(label="📂 Ouvrir le dossier", command=self.open_employee_file)
        self.employee_context_menu.add_command(label="✏️ Modifier", command=self.edit_employee)
        self.employee_context_menu.add_separator()
        self.employee_context_menu.add_command(label="🗑️ Supprimer", command=self.delete_employee)
        
        def show_context_menu(event):
            try:
                self.employee_context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.employee_context_menu.grab_release()
                
        self.employees_tree.bind("<Button-3>", show_context_menu)  # Clic droit
        
    def load_employees(self):
        """Charger la liste des employés"""
        # Vider la liste actuelle
        for item in self.employees_tree.get_children():
            self.employees_tree.delete(item)
            
        # Récupérer les employés de la base de données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        search_term = self.search_var.get() if hasattr(self, 'search_var') else ""
        
        if search_term:
            cursor.execute('''
                SELECT id, matricule, first_name, last_name, job_title, department, status, photo_path
                FROM employees 
                WHERE first_name LIKE ? OR last_name LIKE ? OR matricule LIKE ?
                ORDER BY last_name, first_name
            ''', (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
        else:
            cursor.execute('''
                SELECT id, matricule, first_name, last_name, job_title, department, status, photo_path
                FROM employees 
                ORDER BY last_name, first_name
            ''')
            
        employees = cursor.fetchall()
        conn.close()
        
        # Ajouter les employés à la liste
        for emp in employees:
            emp_id, matricule, first_name, last_name, job_title, department, status, photo_path = emp
            full_name = f"{first_name} {last_name}"
            
            # Indicateur photo
            photo_indicator = "📷" if photo_path and os.path.exists(photo_path) else "👤"
            
            # Couleur selon le statut
            tags = []
            if status == "Active":
                tags = ['active']
            elif status == "En Congé":
                tags = ['on_leave']
            else:
                tags = ['inactive']
                
            self.employees_tree.insert('', 'end',
                                     values=(photo_indicator, matricule, full_name, 
                                           job_title or '', department or '', status),
                                     tags=tags)
        
        # Configuration des couleurs par tag
        self.employees_tree.tag_configure('active', background='#E8F5E8')
        self.employees_tree.tag_configure('on_leave', background='#FFF3E0')
        self.employees_tree.tag_configure('inactive', background='#FFEBEE')
        
    def filter_employees(self, *args):
        """Filtrer les employés selon la recherche"""
        self.load_employees()
        
    def add_new_employee(self):
        """Ajouter un nouvel employé"""
        self.current_employee_id = None
        self.show_employee_form()
        
    def edit_employee(self):
        """Modifier un employé sélectionné"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé à modifier")
            return
            
        # Récupérer l'ID de l'employé
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            self.current_employee_id = result[0]
            self.show_employee_form()
            
    def delete_employee(self):
        """Supprimer un employé"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé à supprimer")
            return
            
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        full_name = item['values'][2]
        
        if messagebox.askyesno("Confirmation", 
                              f"Êtes-vous sûr de vouloir supprimer l'employé {full_name} (Matricule: {matricule}) ?\n\nCette action est irréversible."):
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer l'ID
            cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
            result = cursor.fetchone()
            
            if result:
                emp_id = result[0]
                
                # Supprimer les données liées
                cursor.execute('DELETE FROM career_history WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM documents WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM leaves WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
                
                conn.commit()
                messagebox.showinfo("Succès", "Employé supprimé avec succès")
                self.load_employees()
            
            conn.close()
            
    def open_employee_file(self, event=None):
        """Ouvrir le dossier complet d'un employé"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé")
            return
            
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            self.current_employee_id = result[0]
            self.show_employee_details()
            
    def show_employee_form(self):
        """Afficher le formulaire d'ajout/modification d'employé"""
        # Créer une nouvelle fenêtre
        form_window = tk.Toplevel(self.root)
        form_window.title("Nouvel Employé" if not self.current_employee_id else "Modifier Employé")
        form_window.geometry("800x700")
        form_window.configure(bg=self.colors['background'])
        form_window.transient(self.root)
        form_window.grab_set()
        
        # Variables pour les champs
        self.form_vars = {}
        self.form_text_widgets = {} # <-- LIGNE A AJOUTER
        fields = [
            'matricule', 'first_name', 'last_name', 'gender', 'birth_date', 'birth_place',
            'address', 'phone', 'email', 'marital_status', 'dependents', 'social_security',
            'bank_details', 'hire_date', 'contract_type', 'contract_start', 'contract_end',
            'department', 'job_title', 'status', 'cni', 'nationalite' # <-- MODIFICATION ICI
        ]
        
        for field in fields:
            self.form_vars[field] = tk.StringVar()
            
        # Si modification, charger les données existantes
        if self.current_employee_id:
            self.load_employee_data()
            
        # Titre
        title = tk.Label(form_window,
                        text="📝 " + ("Nouvel Employé" if not self.current_employee_id else "Modifier Employé"),
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(form_window, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet Informations Personnelles
        personal_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(personal_frame, text="👤 Informations Personnelles")
        
        # Onglet Informations Contractuelles
        contract_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(contract_frame, text="📋 Informations Contractuelles")
        
        # Remplir l'onglet personnel
        self.create_personal_info_tab(personal_frame)
        
        # Remplir l'onglet contractuel
        self.create_contract_info_tab(contract_frame)
        
        # Boutons d'action
        buttons_frame = tk.Frame(form_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        save_btn = tk.Button(buttons_frame,
                            text="💾 Enregistrer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=lambda: self.save_employee(form_window))
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=form_window.destroy)
        cancel_btn.pack(side='right')
        
    def create_personal_info_tab(self, parent):
        """Créer l'onglet des informations personnelles"""
        # Frame avec scrollbar
        canvas = tk.Canvas(parent, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Champs du formulaire
        fields_config = [
            ("Matricule *", 'matricule', 'entry'),
            ("Prénom *", 'first_name', 'entry'),
            ("Nom *", 'last_name', 'entry'),
            ("CNI (Carte Nationale d'Identité)", 'cni', 'entry'),      # <-- LIGNE A AJOUTER
            ("Nationalité", 'nationalite', 'entry'), 
            ("Genre", 'gender', 'combo', ['Masculin', 'Féminin']),
            ("Date de Naissance (jj/mm/aaaa)", 'birth_date', 'entry'),
            ("Lieu de Naissance", 'birth_place', 'entry'),
            ("Adresse", 'address', 'text'),
            ("Téléphone", 'phone', 'entry'),
            ("Email", 'email', 'entry'),
            ("Situation Matrimoniale", 'marital_status', 'combo', ['Célibataire', 'Marié(e)', 'Divorcé(e)', 'Veuf/Veuve']),
            ("Nombre de Personnes à Charge", 'dependents', 'entry'),
            ("Numéro de Sécurité Sociale", 'social_security', 'entry'),
            ("RIB/Détails Bancaires", 'bank_details', 'text')
        ]
        
        row = 0
        for field_config in fields_config:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(scrollable_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', padx=20, pady=(10, 5))
            
            # Champ de saisie
            if field_type == 'entry':
                widget = tk.Entry(scrollable_frame,
                                textvariable=self.form_vars[var_name],
                                font=('Segoe UI', 11),
                                width=40,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(scrollable_frame,
                                    textvariable=self.form_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=37,
                                    state='readonly')
            elif field_type == 'text':
                widget = tk.Text(scrollable_frame,
                               font=('Segoe UI', 11),
                               width=40,
                               height=3,
                               relief='solid',
                               bd=1)
                # On stocke une référence au widget pour le manipuler plus tard
                self.form_text_widgets[var_name] = widget
                    
            widget.grid(row=row+1, column=0, sticky='w', padx=20, pady=(0, 5))
            row += 2
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def create_contract_info_tab(self, parent):
        """Créer l'onglet des informations contractuelles"""
        # Frame avec scrollbar
        canvas = tk.Canvas(parent, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Champs contractuels
        fields_config = [
            ("Date d'Embauche (jj/mm/aaaa) *", 'hire_date', 'entry'),
            ("Type de Contrat", 'contract_type', 'combo', ['CDI', 'CDD', 'Stage', 'Consultant']),
            ("Début de Contrat (jj/mm/aaaa)", 'contract_start', 'entry'),
            ("Fin de Contrat (jj/mm/aaaa)", 'contract_end', 'entry'),
            ("Département", 'department', 'entry'),
            ("Poste/Fonction *", 'job_title', 'entry'),
            ("Statut", 'status', 'combo', ['Active', 'En Congé', 'Suspendu', 'Retraité', 'Démissionné'])
        ]
        
        row = 0
        for field_config in fields_config:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(scrollable_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', padx=20, pady=(10, 5))
            
            # Champ de saisie
            if field_type == 'entry':
                widget = tk.Entry(scrollable_frame,
                                textvariable=self.form_vars[var_name],
                                font=('Segoe UI', 11),
                                width=40,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(scrollable_frame,
                                    textvariable=self.form_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=37,
                                    state='readonly')
                    
            widget.grid(row=row+1, column=0, sticky='w', padx=20, pady=(0, 5))
            row += 2
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def load_employee_data(self):
        """Charger les données d'un employé existant"""
        if not self.current_employee_id:
            return
            
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row # Permet d'accéder aux colonnes par leur nom
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM employees WHERE id = ?', (self.current_employee_id,))
        employee_data = cursor.fetchone()
        conn.close()
        
        if employee_data:
            # 1. Remplir les champs standards (Entry, Combobox)
            for field_name, var in self.form_vars.items():
                if field_name in employee_data.keys() and field_name not in self.form_text_widgets:
                    value = employee_data[field_name] or ''
                    var.set(value)

            # 2. Remplir manuellement les champs Text (Adresse et RIB)
            for field_name, text_widget in self.form_text_widgets.items():
                if field_name in employee_data.keys():
                    value = employee_data[field_name] or ''
                    text_widget.delete('1.0', tk.END)  # Vider le champ
                    text_widget.insert('1.0', value)   # Insérer la nouvelle valeur
                        
    def save_employee(self, form_window):
        """Enregistrer les données de l'employé"""
        # Validation des champs obligatoires
        required_fields = {
            'matricule': 'Matricule',
            'first_name': 'Prénom',
            'last_name': 'Nom',
            'hire_date': 'Date d\'embauche',
            'job_title': 'Poste/Fonction'
        }
        
        for field, label in required_fields.items():
            if not self.form_vars[field].get().strip():
                messagebox.showerror("Erreur", f"Le champ '{label}' est obligatoire")
                return
                
        # Validation du format des dates
        date_fields = ['birth_date', 'hire_date', 'contract_start', 'contract_end']
        for field in date_fields:
            date_value = self.form_vars[field].get().strip()
            if date_value and not self.validate_date_format(date_value):
                messagebox.showerror("Erreur", f"Format de date invalide pour '{field}'. Utilisez jj/mm/aaaa")
                return
                
        # Vérifier l'unicité du matricule
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        if self.current_employee_id:
            cursor.execute('SELECT id FROM employees WHERE matricule = ? AND id != ?',
                          (self.form_vars['matricule'].get(), self.current_employee_id))
        else:
            cursor.execute('SELECT id FROM employees WHERE matricule = ?',
                          (self.form_vars['matricule'].get(),))
                          
        if cursor.fetchone():
            messagebox.showerror("Erreur", "Ce matricule existe déjà")
            conn.close()
            return
            
        # Préparer les données
        data = []
        fields = [
            'matricule', 'first_name', 'last_name', 'gender', 'birth_date', 'birth_place',
            'address', 'phone', 'email', 'marital_status', 'dependents', 'social_security',
            'bank_details', 'hire_date', 'contract_type', 'contract_start', 'contract_end',
            'department', 'job_title', 'status', 'cni', 'nationalite' # <-- MODIFICATION ICI
        ]
        
        for field in fields:
            # Si le champ est un widget Text
            if field in self.form_text_widgets:
                value = self.form_text_widgets[field].get('1.0', tk.END).strip()
                data.append(value if value else None)
            # Sinon (c'est un Entry ou Combobox)
            else:
                value = self.form_vars[field].get().strip()
                data.append(value if value else None)
            
        try:
            if self.current_employee_id:
                # Mise à jour
                sql = '''UPDATE employees SET 
                        matricule=?, first_name=?, last_name=?, gender=?, birth_date=?, birth_place=?,
                        address=?, phone=?, email=?, marital_status=?, dependents=?, social_security=?,
                        bank_details=?, hire_date=?, contract_type=?, contract_start=?, contract_end=?,
                        department=?, job_title=?, status=?, cni=?, nationalite=?, updated_at=CURRENT_TIMESTAMP
                        WHERE id=?'''
                data.append(self.current_employee_id)
                cursor.execute(sql, data)
                message = "Employé modifié avec succès"
            else:
                # Insertion
                sql = '''INSERT INTO employees 
                        (matricule, first_name, last_name, gender, birth_date, birth_place,
                         address, phone, email, marital_status, dependents, social_security,
                         bank_details, hire_date, contract_type, contract_start, contract_end,
                         department, job_title, status, cni, nationalite) 
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                cursor.execute(sql, data)
                message = "Employé ajouté avec succès"
                
            conn.commit()
            messagebox.showinfo("Succès", message)
            form_window.destroy()
            self.load_employees()  # Recharger la liste
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def validate_date_format(self, date_string):
        """Valider le format de date jj/mm/aaaa"""
        try:
            datetime.strptime(date_string, '%d/%m/%Y')
            return True
        except ValueError:
            return False
            
    def show_employee_details(self):
        """Afficher les détails complets d'un employé"""
        if not self.current_employee_id:
            return
            
        # Créer une nouvelle fenêtre
        details_window = tk.Toplevel(self.root)
        details_window.title("Dossier Employé")
        details_window.geometry("1000x800")
        details_window.configure(bg=self.colors['background'])
        details_window.transient(self.root)
        
        # Récupérer les données de l'employé
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM employees WHERE id = ?', (self.current_employee_id,))
        employee = cursor.fetchone()
        conn.close()
        
        if not employee:
            messagebox.showerror("Erreur", "Employé non trouvé")
            details_window.destroy()
            return
            
        # Titre avec nom de l'employé
        title = tk.Label(details_window,
                        text=f"📁 Dossier de {employee[2]} {employee[3]}",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(details_window, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglets
        self.create_employee_info_tab(notebook, employee)
        self.create_career_history_tab(notebook)
        self.create_documents_tab(notebook)
        self.create_leaves_history_tab(notebook)
        
    def create_employee_info_tab(self, notebook, employee):
        """Créer l'onglet des informations de l'employé"""
        info_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(info_frame, text="👤 Informations")
        
        # Frame principal avec scrollbar
        canvas = tk.Canvas(info_frame, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Section Photo
        photo_frame = tk.LabelFrame(scrollable_frame,
                                   text="📷 Photo",
                                   font=('Segoe UI', 12, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'])
        photo_frame.pack(fill='x', padx=20, pady=10)
        
        # Placeholder pour la photo SANS width et height pour qu'il s'adapte à l'image
        photo_label = tk.Label(photo_frame,
                              bg=self.colors['light_gray'],
                              relief='solid',
                              bd=1)
        photo_label.pack(side='left', padx=10, pady=10)
        
        # On charge la photo existante de l'employé dès l'ouverture
        # L'index 21 correspond à la colonne 'photo_path' dans la BDD
        photo_path = employee[21]
        self.display_photo(photo_label, photo_path)
        
        # Boutons photo
        photo_buttons = tk.Frame(photo_frame, bg=self.colors['background'])
        photo_buttons.pack(side='left', padx=20, pady=10)
        
        upload_photo_btn = tk.Button(photo_buttons,
                                    text="📁 Charger Photo",
                                    font=('Segoe UI', 10),
                                    bg=self.colors['accent_green'],
                                    fg=self.colors['text_dark'],
                                    relief='flat',
                                    bd=0,
                                    padx=15,
                                    pady=5,
                                    cursor='hand2',
                                    command=lambda: self.upload_employee_photo(photo_label))
        upload_photo_btn.pack(pady=2)
        
        # Informations personnelles
        personal_frame = tk.LabelFrame(scrollable_frame,
                                      text="👤 Informations Personnelles",
                                      font=('Segoe UI', 12, 'bold'),
                                      fg=self.colors['primary_green'],
                                      bg=self.colors['background'])
        personal_frame.pack(fill='x', padx=20, pady=10)
        
        # Mapping des champs
        personal_fields = [
            ("Matricule:", employee[1]),
            ("Nom Complet:", f"{employee[2]} {employee[3]}"),
            ("CNI:", employee[24] or 'Non renseigné'),             # <-- LIGNE A AJOUTER (index 24)
            ("Nationalité:", employee[25] or 'Non renseigné'),      # <-- LIGNE A AJOUTER (index 25)
            ("Genre:", employee[4]),
            ("Date de Naissance:", employee[5]),
            ("Lieu de Naissance:", employee[6]),
            ("Adresse:", employee[7]),
            ("Téléphone:", employee[8]),
            ("Email:", employee[9]),
            ("Situation Matrimoniale:", employee[10]),
            ("Personnes à Charge:", employee[11]),
            ("Sécurité Sociale:", employee[12])
        ]
        
        for i, (label, value) in enumerate(personal_fields):
            row_frame = tk.Frame(personal_frame, bg=self.colors['background'])
            row_frame.pack(fill='x', padx=10, pady=2)
            
            tk.Label(row_frame,
                    text=label,
                    font=('Segoe UI', 10, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background'],
                    width=20,
                    anchor='w').pack(side='left')
                    
            tk.Label(row_frame,
                    text=value or 'Non renseigné',
                    font=('Segoe UI', 10),
                    fg=self.colors['text_dark'] if value else self.colors['text_light'],
                    bg=self.colors['background'],
                    anchor='w').pack(side='left', padx=(10, 0))
        
        # Informations contractuelles
        contract_frame = tk.LabelFrame(scrollable_frame,
                                      text="📋 Informations Contractuelles",
                                      font=('Segoe UI', 12, 'bold'),
                                      fg=self.colors['primary_green'],
                                      bg=self.colors['background'])
        contract_frame.pack(fill='x', padx=20, pady=10)
        
        contract_fields = [
            ("Date d'Embauche:", employee[14]),
            ("Type de Contrat:", employee[15]),
            ("Début de Contrat:", employee[16]),
            ("Fin de Contrat:", employee[17]),
            ("Département:", employee[18]),
            ("Poste/Fonction:", employee[19]),
            ("Statut:", employee[20])
        ]
        
        for label, value in contract_fields:
            row_frame = tk.Frame(contract_frame, bg=self.colors['background'])
            row_frame.pack(fill='x', padx=10, pady=2)
            
            tk.Label(row_frame,
                    text=label,
                    font=('Segoe UI', 10, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background'],
                    width=20,
                    anchor='w').pack(side='left')
                    
            tk.Label(row_frame,
                    text=value or 'Non renseigné',
                    font=('Segoe UI', 10),
                    fg=self.colors['text_dark'] if value else self.colors['text_light'],
                    bg=self.colors['background'],
                    anchor='w').pack(side='left', padx=(10, 0))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def upload_employee_photo(self, photo_label):
        """Charger une photo pour l'employé"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner une photo",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.gif *.bmp")]
        )
        
        if file_path:
            try:
                # Copier le fichier dans le dossier photos
                filename = f"emp_{self.current_employee_id}_{os.path.basename(file_path)}"
                dest_path = os.path.join(self.photos_folder, filename)
                shutil.copy2(file_path, dest_path)
                
                # Mettre à jour la base de données
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('UPDATE employees SET photo_path = ? WHERE id = ?',
                              (dest_path, self.current_employee_id))
                conn.commit()
                conn.close()
                
                # Afficher la photo
                self.display_photo(photo_label, dest_path)
                
                messagebox.showinfo("Succès", "Photo chargée avec succès")
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors du chargement de la photo: {str(e)}")
                
    def display_photo(self, label, photo_path):
        """Afficher une photo dans un label"""
        try:
            if photo_path and os.path.exists(photo_path):
                # Charger et redimensionner l'image
                image = Image.open(photo_path)
                # MODIFIEZ LA TAILLE ICI (largeur, hauteur)
                image = image.resize((200, 210), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(image)
                
                label.configure(image=photo, text="")
                label.image = photo  # Garder une référence pour éviter qu'elle disparaisse
            else:
                # Si aucune photo n'est trouvée, on vide le label
                label.configure(image="", text="Aucune photo")
        except Exception as e:
            label.configure(image="", text="Erreur photo")
            print(f"Erreur lors de l'affichage de la photo : {e}")
            
    def create_career_history_tab(self, notebook):
        """Créer l'onglet historique de carrière"""
        history_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(history_frame, text="📈 Historique Carrière")
        
        # Barre d'outils
        toolbar = tk.Frame(history_frame, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=10)
        
        add_act_btn = tk.Button(toolbar,
                               text="➕ Nouvel Acte",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.add_career_act)
        add_act_btn.pack(side='left')
        
        # Liste des actes
        columns = ('N° Acte', 'Nature', 'Objet', 'Date Acte', 'Date Effet', 'Document')
        self.career_tree = ttk.Treeview(history_frame,
                                       columns=columns,
                                       show='headings',
                                       style='Custom.Treeview',
                                       height=12)
        
        # Configuration des colonnes
        for col in columns:
            self.career_tree.heading(col, text=col)
            self.career_tree.column(col, width=120, anchor='center')
            
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(history_frame, orient='vertical', command=self.career_tree.yview)
        self.career_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Placement
        self.career_tree.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=(0, 20))
        v_scrollbar.pack(side='right', fill='y', padx=(0, 20), pady=(0, 20))
        
        # Charger l'historique
        self.load_career_history()
        
    def add_career_act(self):
        """Ajouter un acte administratif"""
        # Fenêtre de saisie
        act_window = tk.Toplevel(self.root)
        act_window.title("Nouvel Acte Administratif")
        act_window.geometry("600x500")
        act_window.configure(bg=self.colors['background'])
        act_window.transient(self.root)
        act_window.grab_set()
        
        # Variables
        act_vars = {
            'act_number': tk.StringVar(),
            'nature': tk.StringVar(),
            'subject': tk.StringVar(),
            'act_date': tk.StringVar(),
            'effective_date': tk.StringVar()
        }
        
        # Titre
        title = tk.Label(act_window,
                        text="📝 Nouvel Acte Administratif",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(act_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20)
        
        fields = [
            ("Numéro d'Acte:", 'act_number', 'entry'),
            ("Nature:", 'nature', 'combo', ['Nomination', 'Promotion', 'Mutation', 'Sanction', 'Formation', 'Autre']),
            ("Objet:", 'subject', 'text'),
            ("Date de l'Acte (jj/mm/aaaa):", 'act_date', 'entry'),
            ("Date d'Effet (jj/mm/aaaa):", 'effective_date', 'entry')
        ]
        
        widgets = {}
        row = 0
        for field_config in fields:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(form_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', pady=(10, 5))
            
            # Widget
            if field_type == 'entry':
                widget = tk.Entry(form_frame,
                                textvariable=act_vars[var_name],
                                font=('Segoe UI', 11),
                                width=50,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(form_frame,
                                    textvariable=act_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=47,
                                    state='readonly')
            elif field_type == 'text':
                widget = tk.Text(form_frame,
                               font=('Segoe UI', 11),
                               width=50,
                               height=4,
                               relief='solid',
                               bd=1)
                widgets[var_name] = widget  # Stocker pour récupérer le contenu plus tard
                
            widget.grid(row=row+1, column=0, sticky='w', pady=(0, 5))
            row += 2
            
        # Bouton pour attacher un document
        doc_frame = tk.Frame(form_frame, bg=self.colors['background'])
        doc_frame.grid(row=row, column=0, sticky='w', pady=10)
        
        self.selected_doc_path = None
        
        attach_btn = tk.Button(doc_frame,
                              text="📎 Attacher Document",
                              font=('Segoe UI', 10),
                              bg=self.colors['accent_green'],
                              fg=self.colors['text_dark'],
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2',
                              command=self.select_document)
        attach_btn.pack(side='left')
        
        self.doc_label = tk.Label(doc_frame,
                                 text="Aucun document sélectionné",
                                 font=('Segoe UI', 10),
                                 fg=self.colors['text_light'],
                                 bg=self.colors['background'])
        self.doc_label.pack(side='left', padx=(10, 0))
        
        # Boutons d'action
        buttons_frame = tk.Frame(act_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=20)
        
        save_btn = tk.Button(buttons_frame,
                            text="💾 Enregistrer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=lambda: self.save_career_act(act_vars, widgets, act_window))
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=act_window.destroy)
        cancel_btn.pack(side='right')
        
    def select_document(self):
        """Sélectionner un document à attacher"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner un document",
            filetypes=[("Tous les fichiers", "*.*"), ("PDF", "*.pdf"), ("Images", "*.jpg *.jpeg *.png"), ("Word", "*.docx *.doc")]
        )
        
        if file_path:
            self.selected_doc_path = file_path
            filename = os.path.basename(file_path)
            self.doc_label.configure(text=f"📄 {filename}", fg=self.colors['primary_green'])
            
    def save_career_act(self, act_vars, widgets, window):
        """Enregistrer un acte administratif"""
        # Validation
        required_fields = ['act_number', 'nature', 'act_date']
        for field in required_fields:
            if not act_vars[field].get().strip():
                messagebox.showerror("Erreur", f"Le champ '{field}' est obligatoire")
                return
                
        # Validation des dates
        for date_field in ['act_date', 'effective_date']:
            date_value = act_vars[date_field].get().strip()
            if date_value and not self.validate_date_format(date_value):
                messagebox.showerror("Erreur", f"Format de date invalide pour '{date_field}'. Utilisez jj/mm/aaaa")
                return
                
        # Récupérer le contenu du champ texte
        subject = widgets['subject'].get('1.0', tk.END).strip() if 'subject' in widgets else act_vars['subject'].get()
        
        # Copier le document si sélectionné
        doc_path = None
        if self.selected_doc_path:
            try:
                filename = f"act_{self.current_employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(self.selected_doc_path)}"
                doc_path = os.path.join(self.documents_folder, filename)
                shutil.copy2(self.selected_doc_path, doc_path)
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la copie du document: {str(e)}")
                return
                
        # Enregistrer en base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO career_history 
                (employee_id, act_number, nature, subject, act_date, effective_date, document_path)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                self.current_employee_id,
                act_vars['act_number'].get(),
                act_vars['nature'].get(),
                subject,
                act_vars['act_date'].get(),
                act_vars['effective_date'].get() or None,
                doc_path
            ))
            
            conn.commit()
            messagebox.showinfo("Succès", "Acte administratif enregistré avec succès")
            window.destroy()
            self.load_career_history()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def load_career_history(self):
        """Charger l'historique de carrière"""
        if not hasattr(self, 'career_tree'):
            return
            
        # Vider la liste
        for item in self.career_tree.get_children():
            self.career_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT act_number, nature, subject, act_date, effective_date, document_path
            FROM career_history 
            WHERE employee_id = ?
            ORDER BY act_date DESC
        ''', (self.current_employee_id,))
        
        acts = cursor.fetchall()
        conn.close()
        
        for act in acts:
            act_number, nature, subject, act_date, effective_date, document_path = act
            doc_indicator = "📄" if document_path and os.path.exists(document_path) else ""
            
            self.career_tree.insert('', 'end', values=(
                act_number,
                nature,
                subject[:30] + "..." if len(subject) > 30 else subject,
                act_date,
                effective_date or "",
                doc_indicator
            ))
            
    def create_documents_tab(self, notebook):
        """Créer l'onglet gestion des documents"""
        docs_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(docs_frame, text="📄 Documents")
        
        # Barre d'outils
        toolbar = tk.Frame(docs_frame, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=10)
        
        add_doc_btn = tk.Button(toolbar,
                               text="📁 Ajouter Document",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.add_document)
        add_doc_btn.pack(side='left')
        
        # Liste des documents
        columns = ('Catégorie', 'Nom du Document', 'Date d\'Ajout', 'Actions')
        self.docs_tree = ttk.Treeview(docs_frame,
                                     columns=columns,
                                     show='headings',
                                     style='Custom.Treeview',
                                     height=12)
        
        # Configuration des colonnes
        for col in columns:
            self.docs_tree.heading(col, text=col)
            if col == 'Nom du Document':
                self.docs_tree.column(col, width=300, anchor='w')
            else:
                self.docs_tree.column(col, width=150, anchor='center')
                
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(docs_frame, orient='vertical', command=self.docs_tree.yview)
        self.docs_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Placement
        self.docs_tree.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=(0, 20))
        v_scrollbar.pack(side='right', fill='y', padx=(0, 20), pady=(0, 20))
        
        # Double-clic pour ouvrir le document
        self.docs_tree.bind('<Double-1>', self.open_document)
        
        # Charger les documents
        self.load_documents()
        
    def add_document(self):
        """Ajouter un document"""
        # Sélectionner le fichier
        file_path = filedialog.askopenfilename(
            title="Sélectionner un document",
            filetypes=[
                ("Tous les fichiers", "*.*"),
                ("PDF", "*.pdf"),
                ("Images", "*.jpg *.jpeg *.png *.gif *.bmp"),
                ("Word", "*.docx *.doc"),
                ("Excel", "*.xlsx *.xls")
            ]
        )
        
        if not file_path:
            return
            
        # Demander la catégorie
        categories = [
            "Documents Officiels",
            "État Civil",
            "Diplômes et Certifications",
            "Correspondance",
            "Contrats",
            "Évaluations",
            "Autres"
        ]
        
        category = simpledialog.askstring(
            "Catégorie",
            "Choisissez une catégorie:\n" + "\n".join(f"{i+1}. {cat}" for i, cat in enumerate(categories)) + "\n\nEntrez le numéro (1-7):"
        )
        
        if not category or not category.isdigit() or int(category) < 1 or int(category) > len(categories):
            messagebox.showerror("Erreur", "Catégorie invalide")
            return
            
        selected_category = categories[int(category) - 1]
        
        # Demander le nom du document
        doc_name = simpledialog.askstring("Nom du Document", "Entrez le nom du document:")
        if not doc_name:
            doc_name = os.path.basename(file_path)
            
        try:
            # Copier le fichier
            filename = f"doc_{self.current_employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(file_path)}"
            dest_path = os.path.join(self.documents_folder, filename)
            shutil.copy2(file_path, dest_path)
            
            # Enregistrer en base
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO documents (employee_id, category, name, file_path)
                VALUES (?, ?, ?, ?)
            ''', (self.current_employee_id, selected_category, doc_name, dest_path))
            
            conn.commit()
            conn.close()
            
            messagebox.showinfo("Succès", "Document ajouté avec succès")
            self.load_documents()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'ajout du document: {str(e)}")
            
    def load_documents(self):
        """Charger la liste des documents"""
        if not hasattr(self, 'docs_tree'):
            return
            
        # Vider la liste
        for item in self.docs_tree.get_children():
            self.docs_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT category, name, uploaded_at, file_path
            FROM documents 
            WHERE employee_id = ?
            ORDER BY uploaded_at DESC
        ''', (self.current_employee_id,))
        
        documents = cursor.fetchall()
        conn.close()
        
        for doc in documents:
            category, name, uploaded_at, file_path = doc
            # Formater la date
            try:
                date_obj = datetime.strptime(uploaded_at, '%Y-%m-%d %H:%M:%S')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = uploaded_at
                
            self.docs_tree.insert('', 'end', values=(
                category,
                name,
                formatted_date,
                "👁️ Ouvrir"
            ))
            
    def open_document(self, event=None):
        """Ouvrir un document sélectionné"""
        selection = self.docs_tree.selection()
        if not selection:
            return
            
        item = self.docs_tree.item(selection[0])
        doc_name = item['values'][1]
        
        # Récupérer le chemin du fichier
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT file_path FROM documents 
            WHERE employee_id = ? AND name = ?
        ''', (self.current_employee_id, doc_name))
        
        result = cursor.fetchone()
        conn.close()
        
        if result and os.path.exists(result[0]):
            try:
                # Ouvrir le fichier avec l'application par défaut
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', result[0]))
                elif platform.system() == 'Windows':  # Windows
                    os.startfile(result[0])
                else:  # Linux
                    subprocess.call(('xdg-open', result[0]))
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'ouvrir le document: {str(e)}")
        else:
            messagebox.showerror("Erreur", "Document non trouvé")
            
    def create_leaves_history_tab(self, notebook):
        """Créer l'onglet historique des congés"""
        leaves_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(leaves_frame, text="🏖️ Historique Congés")
        
        # Informations sur les soldes de congés
        balance_frame = tk.LabelFrame(leaves_frame,
                                     text="📊 Soldes de Congés",
                                     font=('Segoe UI', 12, 'bold'),
                                     fg=self.colors['primary_green'],
                                     bg=self.colors['background'])
        balance_frame.pack(fill='x', padx=20, pady=10)
        
        # Calculer les soldes (simplifié)
        current_year = datetime.now().year
        annual_leave_balance = 30  # Exemple: 30 jours par an
        
        balance_label = tk.Label(balance_frame,
                                text=f"Congés Annuels {current_year}: {annual_leave_balance} jours disponibles",
                                font=('Segoe UI', 11),
                                fg=self.colors['text_dark'],
                                bg=self.colors['background'])
        balance_label.pack(padx=10, pady=10)
        
        # Liste des congés
        columns = ('Type', 'Début', 'Fin', 'Durée', 'Statut', 'Notes')
        self.leaves_tree = ttk.Treeview(leaves_frame,
                                       columns=columns,
                                       show='headings',
                                       style='Custom.Treeview',
                                       height=10)
        
        # Configuration des colonnes
        for col in columns:
            self.leaves_tree.heading(col, text=col)
            self.leaves_tree.column(col, width=120, anchor='center')
            
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(leaves_frame, orient='vertical', command=self.leaves_tree.yview)
        self.leaves_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Placement
        self.leaves_tree.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=(0, 20))
        v_scrollbar.pack(side='right', fill='y', padx=(0, 20), pady=(0, 20))
        
        # Charger l'historique des congés
        self.load_leaves_history()
        
    def load_leaves_history(self):
        """Charger l'historique des congés"""
        if not hasattr(self, 'leaves_tree'):
            return
            
        # Vider la liste
        for item in self.leaves_tree.get_children():
            self.leaves_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT lt.name, l.start_date, l.end_date, l.days_count, l.status, l.notes
            FROM leaves l
            JOIN leave_types lt ON l.leave_type_id = lt.id
            WHERE l.employee_id = ?
            ORDER BY l.start_date DESC
        ''', (self.current_employee_id,))
        
        leaves = cursor.fetchall()
        conn.close()
        
        for leave in leaves:
            leave_type, start_date, end_date, days_count, status, notes = leave
            
            self.leaves_tree.insert('', 'end', values=(
                leave_type,
                start_date,
                end_date,
                f"{days_count} jour(s)",
                status,
                notes or ""
            ))
            
    def show_leaves_module(self):
        """Module de gestion des congés"""
        self.clear_main_content()
        self.set_active_nav_button("🏖️ Gestion Congés")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="🏖️ Gestion des Congés et Absences",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les sous-modules
        notebook = ttk.Notebook(self.main_content, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet planification des congés
        planning_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(planning_frame, text="📅 Planification")
        
        # Onglet calendrier
        calendar_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(calendar_frame, text="📆 Calendrier")
        
        # Onglet configuration
        config_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(config_frame, text="⚙️ Configuration")
        
        # Remplir les onglets
        self.create_leave_planning_tab(planning_frame)
        self.create_leave_calendar_tab(calendar_frame)
        self.create_leave_config_tab(config_frame)
        
    def create_leave_planning_tab(self, parent):
        """Créer l'onglet de planification des congés"""
        # Formulaire de planification
        form_frame = tk.LabelFrame(parent,
                                  text="📝 Planifier un Congé",
                                  font=('Segoe UI', 12, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg=self.colors['background'])
        form_frame.pack(fill='x', padx=20, pady=20)
        
        # Variables du formulaire
        self.leave_vars = {
            'employee': tk.StringVar(),
            'leave_type': tk.StringVar(),
            'start_date': tk.StringVar(),
            'end_date': tk.StringVar(),
            'notes': tk.StringVar()
        }
        
        # Champs du formulaire
        fields_frame = tk.Frame(form_frame, bg=self.colors['background'])
        fields_frame.pack(fill='x', padx=20, pady=20)
        
        # Sélection employé
        tk.Label(fields_frame,
                text="Employé:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=0, column=0, sticky='w', pady=5)
        
        # Récupérer la liste des employés
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id, first_name, last_name FROM employees WHERE status = "Active" ORDER BY last_name')
        employees = cursor.fetchall()
        
        employee_choices = [f"{emp[1]} {emp[2]} (ID: {emp[0]})" for emp in employees]
        employee_combo = ttk.Combobox(fields_frame,
                                     textvariable=self.leave_vars['employee'],
                                     values=employee_choices,
                                     font=('Segoe UI', 11),
                                     width=40,
                                     state='readonly')
        employee_combo.grid(row=0, column=1, sticky='w', padx=(10, 0), pady=5)
        
        # Type de congé
        tk.Label(fields_frame,
                text="Type de Congé:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=1, column=0, sticky='w', pady=5)
        
        cursor.execute('SELECT name FROM leave_types ORDER BY name')
        leave_types = [row[0] for row in cursor.fetchall()]
        conn.close()
        
        leave_type_combo = ttk.Combobox(fields_frame,
                                       textvariable=self.leave_vars['leave_type'],
                                       values=leave_types,
                                       font=('Segoe UI', 11),
                                       width=40,
                                       state='readonly')
        leave_type_combo.grid(row=1, column=1, sticky='w', padx=(10, 0), pady=5)
        
        # Date de début
        tk.Label(fields_frame,
                text="Date de Début (jj/mm/aaaa):",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=2, column=0, sticky='w', pady=5)
        
        start_date_entry = tk.Entry(fields_frame,
                                   textvariable=self.leave_vars['start_date'],
                                   font=('Segoe UI', 11),
                                   width=43,
                                   relief='solid',
                                   bd=1)
        start_date_entry.grid(row=2, column=1, sticky='w', padx=(10, 0), pady=5)
        
        # Date de fin
        tk.Label(fields_frame,
                text="Date de Fin (jj/mm/aaaa):",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=3, column=0, sticky='w', pady=5)
        
        end_date_entry = tk.Entry(fields_frame,
                                 textvariable=self.leave_vars['end_date'],
                                 font=('Segoe UI', 11),
                                 width=43,
                                 relief='solid',
                                 bd=1)
        end_date_entry.grid(row=3, column=1, sticky='w', padx=(10, 0), pady=5)
        
        # Notes
        tk.Label(fields_frame,
                text="Notes:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=4, column=0, sticky='w', pady=5)
        
        notes_entry = tk.Entry(fields_frame,
                              textvariable=self.leave_vars['notes'],
                              font=('Segoe UI', 11),
                              width=43,
                              relief='solid',
                              bd=1)
        notes_entry.grid(row=4, column=1, sticky='w', padx=(10, 0), pady=5)
        
        # Bouton d'enregistrement
        save_leave_btn = tk.Button(form_frame,
                                  text="💾 Enregistrer le Congé",
                                  font=('Segoe UI', 12, 'bold'),
                                  bg=self.colors['primary_green'],
                                  fg='white',
                                  relief='flat',
                                  bd=0,
                                  padx=20,
                                  pady=10,
                                  cursor='hand2',
                                  command=self.save_leave)
        save_leave_btn.pack(pady=10)
        
        # Liste des congés récents
        recent_frame = tk.LabelFrame(parent,
                                    text="📋 Congés Récents",
                                    font=('Segoe UI', 12, 'bold'),
                                    fg=self.colors['primary_green'],
                                    bg=self.colors['background'])
        recent_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Treeview pour les congés récents
        columns = ('Employé', 'Type', 'Début', 'Fin', 'Durée', 'Statut')
        self.recent_leaves_tree = ttk.Treeview(recent_frame,
                                              columns=columns,
                                              show='headings',
                                              style='Custom.Treeview',
                                              height=8)
        
        for col in columns:
            self.recent_leaves_tree.heading(col, text=col)
            self.recent_leaves_tree.column(col, width=120, anchor='center')
            
        # Scrollbar
        v_scrollbar = ttk.Scrollbar(recent_frame, orient='vertical', command=self.recent_leaves_tree.yview)
        self.recent_leaves_tree.configure(yscrollcommand=v_scrollbar.set)
        
        self.recent_leaves_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        v_scrollbar.pack(side='right', fill='y', padx=(0, 10), pady=10)
        
        # Charger les congés récents
        self.load_recent_leaves()
        
    def save_leave(self):
        """Enregistrer un congé"""
        # Validation
        if not all([
            self.leave_vars['employee'].get(),
            self.leave_vars['leave_type'].get(),
            self.leave_vars['start_date'].get(),
            self.leave_vars['end_date'].get()
        ]):
            messagebox.showerror("Erreur", "Tous les champs obligatoires doivent être remplis")
            return
            
        # Validation des dates
        start_date = self.leave_vars['start_date'].get()
        end_date = self.leave_vars['end_date'].get()
        
        if not self.validate_date_format(start_date) or not self.validate_date_format(end_date):
            messagebox.showerror("Erreur", "Format de date invalide. Utilisez jj/mm/aaaa")
            return
            
        # Calculer la durée
        try:
            start_dt = datetime.strptime(start_date, '%d/%m/%Y')
            end_dt = datetime.strptime(end_date, '%d/%m/%Y')
            
            if end_dt < start_dt:
                messagebox.showerror("Erreur", "La date de fin doit être postérieure à la date de début")
                return
                
            days_count = (end_dt - start_dt).days + 1
        except ValueError:
            messagebox.showerror("Erreur", "Dates invalides")
            return
            
        # Extraire l'ID de l'employé
        employee_text = self.leave_vars['employee'].get()
        try:
            employee_id = int(employee_text.split('ID: ')[1].split(')')[0])
        except:
            messagebox.showerror("Erreur", "Employé invalide")
            return
            
        # Récupérer l'ID du type de congé
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM leave_types WHERE name = ?', (self.leave_vars['leave_type'].get(),))
        leave_type_result = cursor.fetchone()
        
        if not leave_type_result:
            messagebox.showerror("Erreur", "Type de congé invalide")
            conn.close()
            return
            
        leave_type_id = leave_type_result[0]
        
        # Enregistrer le congé
        try:
            cursor.execute('''
                INSERT INTO leaves (employee_id, leave_type_id, start_date, end_date, days_count, notes)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                employee_id,
                leave_type_id,
                start_date,
                end_date,
                days_count,
                self.leave_vars['notes'].get()
            ))
            
            conn.commit()
            messagebox.showinfo("Succès", f"Congé enregistré avec succès ({days_count} jour(s))")
            
            # Réinitialiser le formulaire
            for var in self.leave_vars.values():
                var.set('')
                
            # Recharger la liste
            self.load_recent_leaves()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def load_recent_leaves(self):
        """Charger les congés récents"""
        if not hasattr(self, 'recent_leaves_tree'):
            return
            
        # Vider la liste
        for item in self.recent_leaves_tree.get_children():
            self.recent_leaves_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT e.first_name || ' ' || e.last_name, lt.name, l.start_date, l.end_date, l.days_count, l.status
            FROM leaves l
            JOIN employees e ON l.employee_id = e.id
            JOIN leave_types lt ON l.leave_type_id = lt.id
            ORDER BY l.created_at DESC
            LIMIT 20
        ''')
        
        leaves = cursor.fetchall()
        conn.close()
        
        for leave in leaves:
            employee_name, leave_type, start_date, end_date, days_count, status = leave
            
            self.recent_leaves_tree.insert('', 'end', values=(
                employee_name,
                leave_type,
                start_date,
                end_date,
                f"{days_count} jour(s)",
                status
            ))
            
    def create_leave_calendar_tab(self, parent):
        """Créer l'onglet calendrier des congés"""
        # Titre
        title = tk.Label(parent,
                        text="📆 Calendrier des Congés",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=20)
        
        # Navigation mensuelle
        nav_frame = tk.Frame(parent, bg=self.colors['background'])
        nav_frame.pack(pady=10)
        
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year
        
        prev_btn = tk.Button(nav_frame,
                            text="◀ Précédent",
                            font=('Segoe UI', 11),
                            bg=self.colors['accent_green'],
                            fg=self.colors['text_dark'],
                            relief='flat',
                            bd=0,
                            padx=15,
                            pady=5,
                            cursor='hand2',
                            command=self.prev_month)
        prev_btn.pack(side='left', padx=5)
        
        self.month_label = tk.Label(nav_frame,
                                   text="",
                                   font=('Segoe UI', 14, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'])
        self.month_label.pack(side='left', padx=20)
        
        next_btn = tk.Button(nav_frame,
                            text="Suivant ▶",
                            font=('Segoe UI', 11),
                            bg=self.colors['accent_green'],
                            fg=self.colors['text_dark'],
                            relief='flat',
                            bd=0,
                            padx=15,
                            pady=5,
                            cursor='hand2',
                            command=self.next_month)
        next_btn.pack(side='left', padx=5)
        
        # Zone du calendrier
        self.calendar_frame = tk.Frame(parent, bg=self.colors['background'])
        self.calendar_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Afficher le calendrier initial
        self.display_calendar()
        
    def prev_month(self):
        """Mois précédent"""
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.display_calendar()
        
    def next_month(self):
        """Mois suivant"""
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.display_calendar()
        
    def display_calendar(self):
        """Afficher le calendrier mensuel"""
        # Nettoyer le frame
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()
            
        # Mettre à jour le label du mois
        month_names = [
            "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
            "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
        ]
        self.month_label.configure(text=f"{month_names[self.current_month-1]} {self.current_year}")
        
        # Créer le calendrier
        cal = calendar.monthcalendar(self.current_year, self.current_month)
        
        # En-têtes des jours
        days = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
        for i, day in enumerate(days):
            label = tk.Label(self.calendar_frame,
                           text=day,
                           font=('Segoe UI', 11, 'bold'),
                           fg='white',
                           bg=self.colors['primary_green'],
                           width=12,
                           height=2)
            label.grid(row=0, column=i, padx=1, pady=1, sticky='nsew')
            
        # Récupérer les congés du mois
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Dates du mois
        first_day = f"01/{self.current_month:02d}/{self.current_year}"
        last_day = f"{calendar.monthrange(self.current_year, self.current_month)[1]:02d}/{self.current_month:02d}/{self.current_year}"
        
        cursor.execute('''
            SELECT l.start_date, l.end_date, e.first_name, e.last_name, lt.name
            FROM leaves l
            JOIN employees e ON l.employee_id = e.id
            JOIN leave_types lt ON l.leave_type_id = lt.id
            WHERE l.status = "Approved"
            AND ((l.start_date <= ? AND l.end_date >= ?) OR 
                 (l.start_date >= ? AND l.start_date <= ?))
        ''', (last_day, first_day, first_day, last_day))
        
        leaves_data = cursor.fetchall()
        conn.close()
        
        # Créer un dictionnaire des congés par date
        leaves_by_date = {}
        for leave in leaves_data:
            start_date, end_date, first_name, last_name, leave_type = leave
            try:
                start_dt = datetime.strptime(start_date, '%d/%m/%Y')
                end_dt = datetime.strptime(end_date, '%d/%m/%Y')
                
                current_dt = start_dt
                while current_dt <= end_dt:
                    if current_dt.month == self.current_month and current_dt.year == self.current_year:
                        day_key = current_dt.day
                        if day_key not in leaves_by_date:
                            leaves_by_date[day_key] = []
                        leaves_by_date[day_key].append(f"{first_name} {last_name}")
                    current_dt += timedelta(days=1)
            except ValueError:
                continue
                
        # Afficher les jours du calendrier
        for week_num, week in enumerate(cal):
            for day_num, day in enumerate(week):
                if day == 0:
                    # Jour vide
                    label = tk.Label(self.calendar_frame,
                                   text="",
                                   bg=self.colors['background'],
                                   width=12,
                                   height=4)
                else:
                    # Jour avec ou sans congés
                    bg_color = self.colors['white']
                    text_color = self.colors['text_dark']
                    day_text = str(day)
                    
                    if day in leaves_by_date:
                        bg_color = self.colors['light_green']
                        text_color = 'white'
                        employees_count = len(leaves_by_date[day])
                        if employees_count > 1:
                            day_text += f"\n({employees_count})"
                        else:
                            day_text += f"\n{leaves_by_date[day][0].split()[0]}"
                            
                    label = tk.Label(self.calendar_frame,
                                   text=day_text,
                                   font=('Segoe UI', 10),
                                   fg=text_color,
                                   bg=bg_color,
                                   width=12,
                                   height=4,
                                   relief='solid',
                                   bd=1,
                                   justify='center')
                    
                    # Tooltip pour les congés
                    if day in leaves_by_date:
                        tooltip_text = f"Congés le {day:02d}/{self.current_month:02d}/{self.current_year}:\n"
                        tooltip_text += "\n".join(leaves_by_date[day])
                        self.create_tooltip(label, tooltip_text)
                        
                label.grid(row=week_num+1, column=day_num, padx=1, pady=1, sticky='nsew')
                
        # Configurer les poids des colonnes et lignes
        for i in range(7):
            self.calendar_frame.grid_columnconfigure(i, weight=1)
        for i in range(len(cal)+1):
            self.calendar_frame.grid_rowconfigure(i, weight=1)
            
    def create_tooltip(self, widget, text):
        """Créer un tooltip pour un widget"""
        def on_enter(event):
            tooltip = tk.Toplevel()
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            
            label = tk.Label(tooltip,
                           text=text,
                           font=('Segoe UI', 9),
                           bg='#FFFFCC',
                           fg=self.colors['text_dark'],
                           relief='solid',
                           bd=1,
                           padx=5,
                           pady=3)
            label.pack()
            
            widget.tooltip = tooltip
            
        def on_leave(event):
            if hasattr(widget, 'tooltip'):
                widget.tooltip.destroy()
                del widget.tooltip
                
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
        
    def create_leave_config_tab(self, parent):
        """Créer l'onglet de configuration des congés"""
        # Titre
        title = tk.Label(parent,
                        text="⚙️ Configuration des Types de Congés",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=20)
        
        # Formulaire d'ajout de type de congé
        form_frame = tk.LabelFrame(parent,
                                  text="➕ Ajouter un Type de Congé",
                                  font=('Segoe UI', 12, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg=self.colors['background'])
        form_frame.pack(fill='x', padx=20, pady=10)
        
        # Variables du formulaire
        self.leave_type_vars = {
            'name': tk.StringVar(),
            'days_per_year': tk.StringVar(),
            'description': tk.StringVar()
        }
        
        # Champs
        fields_frame = tk.Frame(form_frame, bg=self.colors['background'])
        fields_frame.pack(padx=20, pady=15)
        
        tk.Label(fields_frame,
                text="Nom du Type:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=0, column=0, sticky='w', pady=5)
        
        tk.Entry(fields_frame,
                textvariable=self.leave_type_vars['name'],
                font=('Segoe UI', 11),
                width=30,
                relief='solid',
                bd=1).grid(row=0, column=1, padx=(10, 0), pady=5)
        
        tk.Label(fields_frame,
                text="Jours par An:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=1, column=0, sticky='w', pady=5)
        
        tk.Entry(fields_frame,
                textvariable=self.leave_type_vars['days_per_year'],
                font=('Segoe UI', 11),
                width=30,
                relief='solid',
                bd=1).grid(row=1, column=1, padx=(10, 0), pady=5)
        
        tk.Label(fields_frame,
                text="Description:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=2, column=0, sticky='w', pady=5)
        
        tk.Entry(fields_frame,
                textvariable=self.leave_type_vars['description'],
                font=('Segoe UI', 11),
                width=30,
                relief='solid',
                bd=1).grid(row=2, column=1, padx=(10, 0), pady=5)
        
        # Bouton d'ajout
        add_type_btn = tk.Button(form_frame,
                                text="➕ Ajouter",
                                font=('Segoe UI', 11, 'bold'),
                                bg=self.colors['primary_green'],
                                fg='white',
                                relief='flat',
                                bd=0,
                                padx=20,
                                pady=8,
                                cursor='hand2',
                                command=self.add_leave_type)
        add_type_btn.pack(pady=10)
        
        # Liste des types existants
        list_frame = tk.LabelFrame(parent,
                                  text="📋 Types de Congés Existants",
                                  font=('Segoe UI', 12, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg=self.colors['background'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Treeview
        columns = ('Nom', 'Jours/An', 'Description')
        self.leave_types_tree = ttk.Treeview(list_frame,
                                            columns=columns,
                                            show='headings',
                                            style='Custom.Treeview',
                                            height=8)
        
        for col in columns:
            self.leave_types_tree.heading(col, text=col)
            if col == 'Description':
                self.leave_types_tree.column(col, width=300, anchor='w')
            else:
                self.leave_types_tree.column(col, width=150, anchor='center')
                
        # Scrollbar
        v_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.leave_types_tree.yview)
        self.leave_types_tree.configure(yscrollcommand=v_scrollbar.set)
        
        self.leave_types_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        v_scrollbar.pack(side='right', fill='y', padx=(0, 10), pady=10)
        
        # Charger les types existants
        self.load_leave_types()
        
    def add_leave_type(self):
        """Ajouter un nouveau type de congé"""
        name = self.leave_type_vars['name'].get().strip()
        days_per_year = self.leave_type_vars['days_per_year'].get().strip()
        description = self.leave_type_vars['description'].get().strip()
        
        if not name:
            messagebox.showerror("Erreur", "Le nom du type de congé est obligatoire")
            return
            
        # Validation des jours par an
        try:
            days_per_year = int(days_per_year) if days_per_year else 0
        except ValueError:
            messagebox.showerror("Erreur", "Le nombre de jours par an doit être un nombre entier positif")
            return
            
        # Enregistrer en base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO leave_types (name, days_per_year, description)
                VALUES (?, ?, ?)
            ''', (name, days_per_year, description))
            
            conn.commit()
            messagebox.showinfo("Succès", "Type de congé ajouté avec succès")
            
            # Réinitialiser le formulaire
            for var in self.leave_type_vars.values():
                var.set('')
                
            # Recharger la liste
            self.load_leave_types()
            
        except sqlite3.IntegrityError:
            messagebox.showerror("Erreur", "Ce type de congé existe déjà")
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def load_leave_types(self):
        """Charger les types de congés"""
        if not hasattr(self, 'leave_types_tree'):
            return
            
        # Vider la liste
        for item in self.leave_types_tree.get_children():
            self.leave_types_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT name, days_per_year, description FROM leave_types ORDER BY name')
        
        leave_types = cursor.fetchall()
        conn.close()
        
        for leave_type in leave_types:
            name, days_per_year, description = leave_type
            self.leave_types_tree.insert('', 'end', values=(
                name,
                days_per_year,
                description or ""
            ))
            
    def show_mail_module(self):
        """Module de gestion des courriers - MISE À JOUR avec upload de fichiers"""
        self.clear_main_content()
        self.set_active_nav_button("📮 Gestion Courriers")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📮 Gestion des Courriers",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Barre d'outils
        toolbar = tk.Frame(self.main_content, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=(0, 10))
        
        # Boutons d'action
        add_btn = tk.Button(toolbar,
                           text="➕ Nouveau Courrier",
                           font=('Segoe UI', 11, 'bold'),
                           bg=self.colors['primary_green'],
                           fg='white',
                           relief='flat',
                           bd=0,
                           padx=15,
                           pady=8,
                           cursor='hand2',
                           command=self.add_new_mail)
        add_btn.pack(side='left', padx=(0, 10))
        
        refresh_btn = tk.Button(toolbar,
                               text="🔄 Rafraîchir",
                               font=('Segoe UI', 11),
                               bg=self.colors['accent_green'],
                               fg=self.colors['text_dark'],
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.show_mail_module)
        refresh_btn.pack(side='left')
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(self.main_content, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet Courriers d'Arrivée
        arrival_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(arrival_frame, text="📥 Courriers d'Arrivée")
        
        # Onglet Courriers de Départ
        departure_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(departure_frame, text="📤 Courriers de Départ")
        
        # Créer les listes pour chaque onglet
        self.create_mail_list(arrival_frame, 'arrivee')
        self.create_mail_list(departure_frame, 'depart')
        
    def create_mail_list(self, parent, mail_type):
        """Créer la liste des courriers pour un type donné - MISE À JOUR avec colonne fichier"""
        # Frame de recherche
        search_frame = tk.Frame(parent, bg=self.colors['background'])
        search_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(search_frame, text="🔍 Rechercher:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(side='left', padx=(0, 5))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        search_entry.pack(side='left', padx=(0, 10))
        
        search_btn = tk.Button(search_frame, text="Rechercher",
                              command=lambda: self.search_mail(mail_type, search_var.get()),
                              font=('Segoe UI', 10),
                              bg=self.colors['primary_green'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2')
        search_btn.pack(side='left')
        
        # Treeview pour la liste des courriers - AJOUT de la colonne Fichier
        columns = ('N° Ordre', 'Nb Pièces', 'Date', 'Expéditeur/Destinataire', 'Objet', 'N° Archive', 'Fichier')
        
        tree_frame = tk.Frame(parent, bg=self.colors['background'])
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                           style='Custom.Treeview', height=15)
        
        # Configuration des colonnes
        tree.heading('N° Ordre', text='N° Ordre')
        tree.heading('Nb Pièces', text='Nb Pièces')
        tree.heading('Date', text='Date')
        tree.heading('Expéditeur/Destinataire', text='Expéditeur' if mail_type == 'arrivee' else 'Destinataire')
        tree.heading('Objet', text='Objet')
        tree.heading('N° Archive', text='N° Archive')
        tree.heading('Fichier', text='Fichier')
        
        tree.column('N° Ordre', width=100)
        tree.column('Nb Pièces', width=80)
        tree.column('Date', width=100)
        tree.column('Expéditeur/Destinataire', width=200)
        tree.column('Objet', width=250)
        tree.column('N° Archive', width=100)
        tree.column('Fichier', width=80)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Charger les données
        self.load_mail_data(tree, mail_type)
        
        # Menu contextuel
        context_menu = tk.Menu(tree, tearoff=0)
        context_menu.add_command(label="✏️ Modifier",
                                command=lambda: self.edit_mail(tree))
        context_menu.add_command(label="👁️ Voir détails",
                                command=lambda: self.view_mail_details(tree))
        context_menu.add_separator()
        context_menu.add_command(label="📁 Ouvrir fichier",
                                command=lambda: self.open_mail_file(tree))
        context_menu.add_separator()
        context_menu.add_command(label="🗑️ Supprimer",
                                command=lambda: self.delete_mail(tree))
        
        def show_context_menu(event):
            try:
                context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                context_menu.grab_release()
        
        tree.bind("<Button-3>", show_context_menu)
        tree.bind("<Double-1>", lambda e: self.view_mail_details(tree))
        
        # Stocker la référence du tree pour chaque type
        if mail_type == 'arrivee':
            self.arrival_tree = tree
        else:
            self.departure_tree = tree
            
    def load_mail_data(self, tree, mail_type):
        """Charger les données des courriers dans le treeview - MISE À JOUR avec fichier"""
        # Vider le treeview
        for item in tree.get_children():
            tree.delete(item)
        
        # Récupérer les données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT numero_ordre, nombre_pieces, date_arrivee_expedition,
                   expediteur_destinataire, objet, numero_archive, file_path, id
            FROM courriers
            WHERE type_courrier = ?
            ORDER BY date_arrivee_expedition DESC
        ''', (mail_type,))
        
        for row in cursor.fetchall():
            # Formater la date
            date_str = row[2]
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = date_str
            
            # Indicateur de fichier
            file_indicator = "📄" if row[6] and os.path.exists(row[6]) else ""
            
            tree.insert('', 'end', values=(
                row[0], row[1], formatted_date, row[3], row[4], row[5] or '', file_indicator
            ), tags=(row[7],))  # Stocker l'ID dans les tags
            
        conn.close()
        
    def add_new_mail(self):
        """Ajouter un nouveau courrier"""
        self.show_mail_form()
        
    def show_mail_form(self, mail_id=None):
        """Afficher le formulaire de courrier (nouveau ou modification) - MISE À JOUR avec upload"""
        # Créer une nouvelle fenêtre
        form_window = tk.Toplevel(self.root)
        form_window.title("Nouveau Courrier" if not mail_id else "Modifier Courrier")
        form_window.geometry("600x600")  # Augmenté pour le champ fichier
        form_window.configure(bg=self.colors['background'])
        form_window.transient(self.root)
        form_window.grab_set()
        
        # Variables du formulaire
        numero_ordre_var = tk.StringVar()
        type_courrier_var = tk.StringVar(value='arrivee')
        nombre_pieces_var = tk.StringVar(value='1')
        date_var = tk.StringVar()
        expediteur_destinataire_var = tk.StringVar()
        objet_var = tk.StringVar()
        numero_archive_var = tk.StringVar()
        observation_var = tk.StringVar()
        
        # Variable pour le fichier sélectionné
        self.selected_mail_file = None
        
        # Si modification, charger les données existantes
        if mail_id:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM courriers WHERE id = ?", (mail_id,))
            mail_data = cursor.fetchone()
            conn.close()
            
            if mail_data:
                numero_ordre_var.set(mail_data[1])
                type_courrier_var.set(mail_data[2])
                nombre_pieces_var.set(str(mail_data[3]))
                # Convertir la date au format dd/mm/yyyy
                try:
                    date_obj = datetime.strptime(mail_data[4], '%Y-%m-%d')
                    date_var.set(date_obj.strftime('%d/%m/%Y'))
                except:
                    date_var.set(mail_data[4])
                expediteur_destinataire_var.set(mail_data[5])
                objet_var.set(mail_data[6])
                numero_archive_var.set(mail_data[7] or '')
                observation_var.set(mail_data[8] or '')
                self.selected_mail_file = mail_data[9]  # file_path
        
        # Titre
        title = tk.Label(form_window,
                        text="📮 " + ("Nouveau Courrier" if not mail_id else "Modifier Courrier"),
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Frame principal avec scrollbar
        canvas = tk.Canvas(form_window, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(form_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Formulaire
        form_frame = tk.Frame(scrollable_frame, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        row = 0
        
        # Numéro d'ordre
        tk.Label(form_frame, text="Numéro d'ordre *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        numero_entry = tk.Entry(form_frame, textvariable=numero_ordre_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        numero_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Type de courrier
        tk.Label(form_frame, text="Type de courrier *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        type_frame = tk.Frame(form_frame, bg=self.colors['background'])
        type_frame.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        tk.Radiobutton(type_frame, text="📥 Arrivée", variable=type_courrier_var, value='arrivee',
                      font=('Segoe UI', 11), bg=self.colors['background']).pack(side='left', padx=(0, 20))
        tk.Radiobutton(type_frame, text="📤 Départ", variable=type_courrier_var, value='depart',
                      font=('Segoe UI', 11), bg=self.colors['background']).pack(side='left')
        row += 1
        
        # Nombre de pièces
        tk.Label(form_frame, text="Nombre de pièces:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        nombre_entry = tk.Entry(form_frame, textvariable=nombre_pieces_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        nombre_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Date
        tk.Label(form_frame, text="Date (dd/mm/yyyy) *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        date_entry = tk.Entry(form_frame, textvariable=date_var,
                             font=('Segoe UI', 11), width=30,
                             relief='solid', bd=1)
        date_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Expéditeur/Destinataire
        expediteur_label = tk.Label(form_frame, text="Expéditeur *:",
                                   font=('Segoe UI', 11, 'bold'),
                                   fg=self.colors['text_dark'],
                                   bg=self.colors['background'])
        expediteur_label.grid(row=row, column=0, sticky='w', pady=5)
        expediteur_entry = tk.Entry(form_frame, textvariable=expediteur_destinataire_var,
                                   font=('Segoe UI', 11), width=30,
                                   relief='solid', bd=1)
        expediteur_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Objet
        tk.Label(form_frame, text="Objet *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        objet_entry = tk.Entry(form_frame, textvariable=objet_var,
                              font=('Segoe UI', 11), width=30,
                              relief='solid', bd=1)
        objet_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Numéro d'archive
        tk.Label(form_frame, text="Numéro d'archive:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        archive_entry = tk.Entry(form_frame, textvariable=numero_archive_var,
                                font=('Segoe UI', 11), width=30,
                                relief='solid', bd=1)
        archive_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # NOUVEAU: Section pour l'upload de fichier
        tk.Label(form_frame, text="Fichier joint:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        
        file_frame = tk.Frame(form_frame, bg=self.colors['background'])
        file_frame.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        select_file_btn = tk.Button(file_frame,
                                   text="📁 Sélectionner Fichier",
                                   font=('Segoe UI', 10),
                                   bg=self.colors['accent_green'],
                                   fg=self.colors['text_dark'],
                                   relief='flat',
                                   bd=0,
                                   padx=15,
                                   pady=5,
                                   cursor='hand2',
                                   command=self.select_mail_file)
        select_file_btn.pack(side='left')
        
        self.file_label = tk.Label(file_frame,
                                  text="Aucun fichier sélectionné" if not self.selected_mail_file else os.path.basename(self.selected_mail_file),
                                  font=('Segoe UI', 10),
                                  fg=self.colors['text_light'] if not self.selected_mail_file else self.colors['primary_green'],
                                  bg=self.colors['background'])
        self.file_label.pack(side='left', padx=(10, 0))
        row += 1
        
        # Observation
        tk.Label(form_frame, text="Observation:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='nw', pady=5)
        observation_text = tk.Text(form_frame, font=('Segoe UI', 11), width=30, height=4,
                                  relief='solid', bd=1)
        observation_text.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        # Lier le Text widget à la variable
        if observation_var.get():
            observation_text.insert('1.0', observation_var.get())
        
        form_frame.grid_columnconfigure(1, weight=1)
        
        # Boutons
        button_frame = tk.Frame(scrollable_frame, bg=self.colors['background'])
        button_frame.pack(fill='x', padx=20, pady=20)
        
        # Bouton Annuler
        cancel_btn = tk.Button(button_frame, text="❌ Annuler",
                              command=form_window.destroy,
                              font=('Segoe UI', 11),
                              bg=self.colors['error'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=8,
                              cursor='hand2')
        cancel_btn.pack(side='right', padx=(10, 0))
        
        # Bouton Enregistrer
        save_btn = tk.Button(button_frame, text="💾 Enregistrer",
                            command=lambda: self.save_mail(
                                form_window, mail_id,
                                numero_ordre_var.get(),
                                type_courrier_var.get(),
                                nombre_pieces_var.get(),
                                date_var.get(),
                                expediteur_destinataire_var.get(),
                                objet_var.get(),
                                numero_archive_var.get(),
                                observation_text.get('1.0', 'end-1c')
                            ),
                            font=('Segoe UI', 11, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=8,
                            cursor='hand2')
        save_btn.pack(side='right')
        
        canvas.pack(side="left", fill="both", expand=True, padx=(0, 20))
        scrollbar.pack(side="right", fill="y")
        
        # Focus sur le premier champ
        numero_entry.focus()
        
    def select_mail_file(self):
        """Sélectionner un fichier pour le courrier - NOUVELLE FONCTION"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier pour le courrier",
            filetypes=[
                ("Tous les fichiers", "*.*"),
                ("PDF", "*.pdf"),
                ("Images", "*.jpg *.jpeg *.png *.gif *.bmp"),
                ("Word", "*.docx *.doc"),
                ("Excel", "*.xlsx *.xls"),
                ("Texte", "*.txt")
            ]
        )
        
        if file_path:
            self.selected_mail_file = file_path
            filename = os.path.basename(file_path)
            self.file_label.configure(text=f"📄 {filename}", fg=self.colors['primary_green'])
            
    def save_mail(self, form_window, mail_id, numero_ordre, type_courrier, nombre_pieces,
                  date_str, expediteur_destinataire, objet, numero_archive, observation):
        """Enregistrer un courrier - MISE À JOUR avec gestion fichier"""
        try:
            # Validation
            if not numero_ordre.strip():
                messagebox.showerror("Erreur", "Le numéro d'ordre est obligatoire.")
                return
            
            if not date_str.strip():
                messagebox.showerror("Erreur", "La date est obligatoire.")
                return
            
            if not expediteur_destinataire.strip():
                messagebox.showerror("Erreur", "L'expéditeur/destinataire est obligatoire.")
                return
            
            if not objet.strip():
                messagebox.showerror("Erreur", "L'objet est obligatoire.")
                return
            
            # Valider et convertir la date
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                date_formatted = date_obj.strftime('%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Erreur", "Format de date invalide. Utilisez dd/mm/yyyy.")
                return
            
            # Valider le nombre de pièces
            try:
                nombre_pieces_int = int(nombre_pieces) if nombre_pieces.strip() else 1
                if nombre_pieces_int < 1:
                    raise ValueError()
            except ValueError:
                messagebox.showerror("Erreur", "Le nombre de pièces doit être un nombre entier positif.")
                return
            
            # Gérer le fichier joint
            file_path = None
            if self.selected_mail_file and os.path.exists(self.selected_mail_file):
                try:
                    # Copier le fichier dans le dossier courriers
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"courrier_{numero_ordre.replace('/', '_')}_{timestamp}_{os.path.basename(self.selected_mail_file)}"
                    file_path = os.path.join(self.courriers_folder, filename)
                    shutil.copy2(self.selected_mail_file, file_path)
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de la copie du fichier: {str(e)}")
                    return
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            if mail_id:
                # Modification
                cursor.execute('''
                    UPDATE courriers SET
                        numero_ordre = ?, type_courrier = ?, nombre_pieces = ?,
                        date_arrivee_expedition = ?, expediteur_destinataire = ?,
                        objet = ?, numero_archive = ?, observation = ?, file_path = ?
                    WHERE id = ?
                ''', (numero_ordre.strip(), type_courrier, nombre_pieces_int,
                     date_formatted, expediteur_destinataire.strip(),
                     objet.strip(), numero_archive.strip() or None,
                     observation.strip() or None, file_path, mail_id))
                
                messagebox.showinfo("Succès", "Courrier modifié avec succès!")
            else:
                # Vérifier l'unicité du numéro d'ordre
                cursor.execute("SELECT id FROM courriers WHERE numero_ordre = ?", (numero_ordre.strip(),))
                if cursor.fetchone():
                    messagebox.showerror("Erreur", "Ce numéro d'ordre existe déjà.")
                    conn.close()
                    return
                
                # Nouveau courrier
                cursor.execute('''
                    INSERT INTO courriers (numero_ordre, type_courrier, nombre_pieces,
                                         date_arrivee_expedition, expediteur_destinataire,
                                         objet, numero_archive, observation, file_path, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (numero_ordre.strip(), type_courrier, nombre_pieces_int,
                     date_formatted, expediteur_destinataire.strip(),
                     objet.strip(), numero_archive.strip() or None,
                     observation.strip() or None, file_path, self.current_user['username']))
                
                messagebox.showinfo("Succès", "Courrier enregistré avec succès!")
            
            conn.commit()
            conn.close()
            form_window.destroy()
            
            # Rafraîchir la liste
            self.show_mail_module()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
            
    def open_mail_file(self, tree):
        """Ouvrir le fichier joint d'un courrier - NOUVELLE FONCTION"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Récupérer le chemin du fichier
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM courriers WHERE id = ?", (mail_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result and result[0] and os.path.exists(result[0]):
            try:
                # Ouvrir le fichier avec l'application par défaut
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', result[0]))
                elif platform.system() == 'Windows':  # Windows
                    os.startfile(result[0])
                else:  # Linux
                    subprocess.call(('xdg-open', result[0]))
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
        else:
            messagebox.showinfo("Information", "Aucun fichier joint à ce courrier.")
            
    def search_mail(self, mail_type, search_term):
        """Rechercher des courriers"""
        tree = self.arrival_tree if mail_type == 'arrivee' else self.departure_tree
        
        # Vider le treeview
        for item in tree.get_children():
            tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        if search_term.strip():
            # Recherche avec terme
            cursor.execute('''
                SELECT numero_ordre, nombre_pieces, date_arrivee_expedition,
                       expediteur_destinataire, objet, numero_archive, file_path, id
                FROM courriers
                WHERE type_courrier = ? AND (
                    numero_ordre LIKE ? OR
                    expediteur_destinataire LIKE ? OR
                    objet LIKE ? OR
                    numero_archive LIKE ?
                )
                ORDER BY date_arrivee_expedition DESC
            ''', (mail_type, f'%{search_term}%', f'%{search_term}%',
                 f'%{search_term}%', f'%{search_term}%'))
        else:
            # Afficher tous
            self.load_mail_data(tree, mail_type)
            conn.close()
            return
        
        for row in cursor.fetchall():
            # Formater la date
            date_str = row[2]
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = date_str
            
            # Indicateur de fichier
            file_indicator = "📄" if row[6] and os.path.exists(row[6]) else ""
            
            tree.insert('', 'end', values=(
                row[0], row[1], formatted_date, row[3], row[4], row[5] or '', file_indicator
            ), tags=(row[7],))
            
        conn.close()
        
    def edit_mail(self, tree):
        """Modifier un courrier sélectionné"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier à modifier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if mail_id:
            self.show_mail_form(mail_id)
            
    def view_mail_details(self, tree):
        """Voir les détails d'un courrier - MISE À JOUR avec info fichier"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Récupérer les détails
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM courriers WHERE id = ?", (mail_id,))
        mail_data = cursor.fetchone()
        conn.close()
        
        if not mail_data:
            messagebox.showerror("Erreur", "Courrier introuvable.")
            return
        
        # Créer une fenêtre de détails
        details_window = tk.Toplevel(self.root)
        details_window.title("Détails du Courrier")
        details_window.geometry("500x500")  # Augmenté pour le fichier
        details_window.configure(bg=self.colors['background'])
        details_window.transient(self.root)
        details_window.grab_set()
        
        # Titre
        title = tk.Label(details_window,
                        text="📮 Détails du Courrier",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Détails
        details_frame = tk.Frame(details_window, bg=self.colors['background'])
        details_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Formater la date
        try:
            date_obj = datetime.strptime(mail_data[4], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
        except:
            formatted_date = mail_data[4]
        
        details_info = [
            ("Numéro d'ordre:", mail_data[1]),
            ("Type:", "📥 Arrivée" if mail_data[2] == 'arrivee' else "📤 Départ"),
            ("Nombre de pièces:", str(mail_data[3])),
            ("Date:", formatted_date),
            ("Expéditeur/Destinataire:", mail_data[5]),
            ("Objet:", mail_data[6]),
            ("Numéro d'archive:", mail_data[7] or "Non spécifié"),
            ("Observation:", mail_data[8] or "Aucune"),
            ("Fichier joint:", "Oui" if mail_data[9] and os.path.exists(mail_data[9]) else "Non")
        ]
        
        for i, (label, value) in enumerate(details_info):
            tk.Label(details_frame, text=label,
                    font=('Segoe UI', 11, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background']).grid(row=i, column=0, sticky='nw', pady=5)
            
            if label == "Observation:" and len(str(value)) > 50:
                # Utiliser un Text widget pour les longues observations
                text_widget = tk.Text(details_frame, font=('Segoe UI', 11), height=4, width=40,
                                     relief='solid', bd=1)
                text_widget.insert('1.0', str(value))
                text_widget.config(state='disabled')
                text_widget.grid(row=i, column=1, sticky='ew', padx=(10, 0), pady=5)
            else:
                tk.Label(details_frame, text=str(value),
                        font=('Segoe UI', 11),
                        fg=self.colors['text_dark'],
                        bg=self.colors['background'],
                        wraplength=300,
                        justify='left').grid(row=i, column=1, sticky='w', padx=(10, 0), pady=5)
        
        details_frame.grid_columnconfigure(1, weight=1)
        
        # Boutons
        buttons_frame = tk.Frame(details_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=20)
        
        # Bouton Ouvrir Fichier (si fichier existe)
        if mail_data[9] and os.path.exists(mail_data[9]):
            open_file_btn = tk.Button(buttons_frame, text="📁 Ouvrir Fichier",
                                     command=lambda: self.open_file_direct(mail_data[9]),
                                     font=('Segoe UI', 11),
                                     bg=self.colors['primary_green'],
                                     fg='white',
                                     relief='flat',
                                     bd=0,
                                     padx=20,
                                     pady=8,
                                     cursor='hand2')
            open_file_btn.pack(side='left')
        
        # Bouton Fermer
        close_btn = tk.Button(buttons_frame, text="✖️ Fermer",
                             command=details_window.destroy,
                             font=('Segoe UI', 11),
                             bg=self.colors['text_light'],
                             fg='white',
                             relief='flat',
                             bd=0,
                             padx=20,
                             pady=8,
                             cursor='hand2')
        close_btn.pack(side='right')
        
    def open_file_direct(self, file_path):
        """Ouvrir un fichier directement - NOUVELLE FONCTION"""
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', file_path))
            elif platform.system() == 'Windows':  # Windows
                os.startfile(file_path)
            else:  # Linux
                subprocess.call(('xdg-open', file_path))
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
            
    def delete_mail(self, tree):
        """Supprimer un courrier"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier à supprimer.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Confirmation
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer ce courrier ?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Récupérer le chemin du fichier pour le supprimer
                cursor.execute("SELECT file_path FROM courriers WHERE id = ?", (mail_id,))
                result = cursor.fetchone()
                file_path = result[0] if result else None
                
                # Supprimer le courrier de la base
                cursor.execute("DELETE FROM courriers WHERE id = ?", (mail_id,))
                conn.commit()
                conn.close()
                
                # Supprimer le fichier joint s'il existe
                if file_path and os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"Erreur lors de la suppression du fichier: {e}")
                
                messagebox.showinfo("Succès", "Courrier supprimé avec succès!")
                
                # Rafraîchir la liste
                self.show_mail_module()
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la suppression: {str(e)}")
                
    def show_reports_module(self):
        """Module de génération de rapports"""
        self.clear_main_content()
        self.set_active_nav_button("📊 Rapports")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📊 Génération de Rapports",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les boutons de rapports
        reports_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        reports_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Configuration des rapports disponibles
        reports_config = [
            {
                'title': '👥 Liste Complète du Personnel',
                'description': 'Rapport complet de tous les employés avec leurs informations principales',
                'icon': '📋',
                'command': self.generate_staff_list_report
            },
            {
                'title': '📄 Fiche Employé Détaillée',
                'description': 'Fiche complète et imprimable d\'un employé sélectionné',
                'icon': '👤',
                'command': self.generate_employee_sheet_report
            },
            {
                'title': '🏖️ Rapport Annuel des Congés',
                'description': 'Synthèse des congés pris et soldes restants par employé',
                'icon': '📅',
                'command': self.generate_annual_leave_report
            },
            {
                'title': '📈 Statistiques RH',
                'description': 'Tableau de bord avec les principales métriques RH',
                'icon': '📊',
                'command': self.generate_hr_statistics_report
            }
        ]
        
        # Créer les cartes de rapports
        for i, report in enumerate(reports_config):
            # Frame pour chaque rapport
            report_card = tk.Frame(reports_frame, 
                                  bg='white', 
                                  relief='solid', 
                                  bd=1,
                                  padx=20,
                                  pady=20)
            report_card.grid(row=i//2, column=i%2, padx=20, pady=15, sticky='ew')
            
            # Icône et titre
            header_frame = tk.Frame(report_card, bg='white')
            header_frame.pack(fill='x', pady=(0, 10))
            
            icon_label = tk.Label(header_frame,
                                 text=report['icon'],
                                 font=('Segoe UI', 24),
                                 bg='white')
            icon_label.pack(side='left')
            
            title_label = tk.Label(header_frame,
                                  text=report['title'],
                                  font=('Segoe UI', 14, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg='white')
            title_label.pack(side='left', padx=(10, 0))
            
            # Description
            desc_label = tk.Label(report_card,
                                 text=report['description'],
                                 font=('Segoe UI', 11),
                                 fg=self.colors['text_dark'],
                                 bg='white',
                                 wraplength=300,
                                 justify='left')
            desc_label.pack(fill='x', pady=(0, 15))
            
            # Boutons d'export
            buttons_frame = tk.Frame(report_card, bg='white')
            buttons_frame.pack(fill='x')
            
            # Bouton PDF
            pdf_btn = tk.Button(buttons_frame,
                               text="📄 Générer PDF",
                               font=('Segoe UI', 10, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=lambda cmd=report['command']: cmd('pdf'))
            pdf_btn.pack(side='left', padx=(0, 10))
            
            # Bouton Excel
            excel_btn = tk.Button(buttons_frame,
                                 text="📊 Générer Excel",
                                 font=('Segoe UI', 10, 'bold'),
                                 bg=self.colors['success'],
                                 fg='white',
                                 relief='flat',
                                 bd=0,
                                 padx=15,
                                 pady=8,
                                 cursor='hand2',
                                 command=lambda cmd=report['command']: cmd('excel'))
            excel_btn.pack(side='left')
            
        # Configurer les colonnes pour qu'elles s'étendent uniformément
        reports_frame.grid_columnconfigure(0, weight=1)
        reports_frame.grid_columnconfigure(1, weight=1)
        
    def generate_staff_list_report(self, format_type):
        """Générer le rapport de liste du personnel"""
        try:
            # Récupérer les données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT matricule, first_name, last_name, job_title, department, 
                       hire_date, contract_type, status, phone, email
                FROM employees 
                ORDER BY last_name, first_name
            ''')
            employees = cursor.fetchall()
            conn.close()
            
            if not employees:
                messagebox.showwarning("Attention", "Aucun employé trouvé")
                return
                
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"liste_personnel_{timestamp}.pdf"
                self.create_staff_list_pdf(employees, filename)
            else:  # excel
                filename = f"liste_personnel_{timestamp}.xlsx"
                self.create_staff_list_excel(employees, filename)
                
            messagebox.showinfo("Succès", f"Rapport généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
            
    def create_staff_list_pdf(self, employees, filename):
        """Créer le PDF de la liste du personnel"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph("Liste du Personnel - Mairie de Ngoundiane", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tableau des employés
        data = [['Matricule', 'Nom Complet', 'Poste', 'Département', 'Embauche', 'Statut']]
        
        for emp in employees:
            matricule, first_name, last_name, job_title, department, hire_date, contract_type, status, phone, email = emp
            data.append([
                matricule,
                f"{first_name} {last_name}",
                job_title or '',
                department or '',
                hire_date or '',
                status
            ])
            
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        
    def create_staff_list_excel(self, employees, filename):
        """Créer le fichier Excel de la liste du personnel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Liste Personnel"
        
        # Titre
        ws['A1'] = "Liste du Personnel - Mairie de Ngoundiane"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        # En-têtes
        headers = ['Matricule', 'Nom Complet', 'Poste', 'Département', 'Date Embauche', 'Statut', 'Téléphone', 'Email']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Données
        for row, emp in enumerate(employees, 5):
            matricule, first_name, last_name, job_title, department, hire_date, contract_type, status, phone, email = emp
            
            ws.cell(row=row, column=1, value=matricule)
            ws.cell(row=row, column=2, value=f"{first_name} {last_name}")
            ws.cell(row=row, column=3, value=job_title or '')
            ws.cell(row=row, column=4, value=department or '')
            ws.cell(row=row, column=5, value=hire_date or '')
            ws.cell(row=row, column=6, value=status)
            ws.cell(row=row, column=7, value=phone or '')
            ws.cell(row=row, column=8, value=email or '')
            
        # Ajuster les largeurs de colonnes
        # Ajuster les largeurs de colonnes (version corrigée)

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                # Ignorer les cellules fusionnées
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value:
                        # Ajouter 2 pour un peu d'espace
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Définir une largeur minimale et maximale
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)   
        wb.save(filename)
        
    def generate_employee_sheet_report(self, format_type):
        """Générer la fiche détaillée d'un employé"""
        # Sélectionner un employé
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id, first_name, last_name, matricule FROM employees ORDER BY last_name')
        employees = cursor.fetchall()
        conn.close()
        
        if not employees:
            messagebox.showwarning("Attention", "Aucun employé trouvé")
            return
            
        # Créer une fenêtre de sélection
        selection_window = tk.Toplevel(self.root)
        selection_window.title("Sélectionner un Employé")
        selection_window.geometry("400x300")
        selection_window.configure(bg=self.colors['background'])
        selection_window.transient(self.root)
        selection_window.grab_set()
        
        tk.Label(selection_window,
                text="Sélectionnez un employé:",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary_green'],
                bg=self.colors['background']).pack(pady=20)
        
        # Liste des employés
        listbox = tk.Listbox(selection_window,
                            font=('Segoe UI', 11),
                            height=10)
        listbox.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        for emp in employees:
            emp_id, first_name, last_name, matricule = emp
            listbox.insert(tk.END, f"{first_name} {last_name} (Matricule: {matricule})")
            
        # Boutons
        buttons_frame = tk.Frame(selection_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        def generate_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Attention", "Veuillez sélectionner un employé")
                return
                
            selected_emp = employees[selection[0]]
            selection_window.destroy()
            self.create_employee_sheet_report(selected_emp[0], format_type)
            
        tk.Button(buttons_frame,
                 text=f"Générer {format_type.upper()}",
                 font=('Segoe UI', 11, 'bold'),
                 bg=self.colors['primary_green'],
                 fg='white',
                 relief='flat',
                 bd=0,
                 padx=20,
                 pady=8,
                 cursor='hand2',
                 command=generate_selected).pack(side='right', padx=(10, 0))
        
        tk.Button(buttons_frame,
                 text="Annuler",
                 font=('Segoe UI', 11),
                 bg=self.colors['text_light'],
                 fg='white',
                 relief='flat',
                 bd=0,
                 padx=20,
                 pady=8,
                 cursor='hand2',
                 command=selection_window.destroy).pack(side='right')
                 
    def create_employee_sheet_report(self, employee_id, format_type):
        """Créer la fiche détaillée d'un employé"""
        try:
            # Récupérer les données complètes de l'employé
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM employees WHERE id = ?', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                messagebox.showerror("Erreur", "Employé non trouvé")
                return
                
            # Récupérer l'historique de carrière
            cursor.execute('''
                SELECT act_number, nature, subject, act_date, effective_date
                FROM career_history 
                WHERE employee_id = ?
                ORDER BY act_date DESC
            ''', (employee_id,))
            career_history = cursor.fetchall()
            
            # Récupérer les congés
            cursor.execute('''
                SELECT lt.name, l.start_date, l.end_date, l.days_count, l.status
                FROM leaves l
                JOIN leave_types lt ON l.leave_type_id = lt.id
                WHERE l.employee_id = ?
                ORDER BY l.start_date DESC
                LIMIT 10
            ''', (employee_id,))
            recent_leaves = cursor.fetchall()
            
            conn.close()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            employee_name = f"{employee[2]}_{employee[3]}"
            
            if format_type == 'pdf':
                filename = f"fiche_{employee_name}_{timestamp}.pdf"
                self.create_employee_sheet_pdf(employee, career_history, recent_leaves, filename)
            else:  # excel
                filename = f"fiche_{employee_name}_{timestamp}.xlsx"
                self.create_employee_sheet_excel(employee, career_history, recent_leaves, filename)
                
            messagebox.showinfo("Succès", f"Fiche employé générée: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_employee_sheet_pdf(self, employee, career_history, recent_leaves, filename):
        """Créer le PDF de la fiche employé"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph(f"Fiche Employé - {employee[2]} {employee[3]}", title_style))
        story.append(Spacer(1, 20))
        
        # Informations personnelles
        story.append(Paragraph("Informations Personnelles", styles['Heading2']))
        
        personal_data = [
            ['Matricule:', employee[1]],
            ['Nom Complet:', f"{employee[2]} {employee[3]}"],
            ['Genre:', employee[4] or ''],
            ['Date de Naissance:', employee[5] or ''],
            ['Lieu de Naissance:', employee[6] or ''],
            ['Adresse:', employee[7] or ''],
            ['Téléphone:', employee[8] or ''],
            ['Email:', employee[9] or ''],
            ['Situation Matrimoniale:', employee[10] or ''],
            ['Personnes à Charge:', str(employee[11]) if employee[11] else '0']
        ]
        
        personal_table = Table(personal_data, colWidths=[150, 300])
        personal_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(personal_table)
        story.append(Spacer(1, 20))
        
        # Informations contractuelles
        story.append(Paragraph("Informations Contractuelles", styles['Heading2']))
        
        contract_data = [
            ['Date d\'Embauche:', employee[14] or ''],
            ['Type de Contrat:', employee[15] or ''],
            ['Début de Contrat:', employee[16] or ''],
            ['Fin de Contrat:', employee[17] or ''],
            ['Département:', employee[18] or ''],
            ['Poste/Fonction:', employee[19] or ''],
            ['Statut:', employee[20] or '']
        ]
        
        contract_table = Table(contract_data, colWidths=[150, 300])
        contract_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(contract_table)
        story.append(Spacer(1, 20))
        
        # Historique de carrière (si disponible)
        if career_history:
            story.append(Paragraph("Historique de Carrière", styles['Heading2']))
            
            career_data = [['N° Acte', 'Nature', 'Date Acte', 'Date Effet']]
            for act in career_history[:5]:  # Limiter à 5 derniers actes
                career_data.append([
                    act[0] or '',
                    act[1] or '',
                    act[3] or '',
                    act[4] or ''
                ])
                
            career_table = Table(career_data)
            career_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(career_table)
            
        doc.build(story)
        
    def create_employee_sheet_excel(self, employee, career_history, recent_leaves, filename):
        """Créer le fichier Excel de la fiche employé"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fiche Employé"
        
        # Titre
        ws['A1'] = f"Fiche Employé - {employee[2]} {employee[3]}"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Informations personnelles
        ws[f'A{row}'] = "INFORMATIONS PERSONNELLES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        personal_fields = [
            ('Matricule:', employee[1]),
            ('Nom Complet:', f"{employee[2]} {employee[3]}"),
            ('Genre:', employee[4]),
            ('Date de Naissance:', employee[5]),
            ('Lieu de Naissance:', employee[6]),
            ('Adresse:', employee[7]),
            ('Téléphone:', employee[8]),
            ('Email:', employee[9]),
            ('Situation Matrimoniale:', employee[10]),
            ('Personnes à Charge:', employee[11])
        ]
        
        for label, value in personal_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value or ''
            row += 1
            
        row += 2
        
        # Informations contractuelles
        ws[f'A{row}'] = "INFORMATIONS CONTRACTUELLES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        contract_fields = [
            ('Date d\'Embauche:', employee[14]),
            ('Type de Contrat:', employee[15]),
            ('Début de Contrat:', employee[16]),
            ('Fin de Contrat:', employee[17]),
            ('Département:', employee[18]),
            ('Poste/Fonction:', employee[19]),
            ('Statut:', employee[20])
        ]
        
        for label, value in contract_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value or ''
            row += 1
            
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        wb.save(filename)
        
    def generate_annual_leave_report(self, format_type):
        """Générer le rapport annuel des congés"""
        try:
            current_year = datetime.now().year
            
            # Récupérer les données des congés
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT e.matricule, e.first_name, e.last_name, 
                       COUNT(l.id) as total_leaves,
                       SUM(l.days_count) as total_days,
                       GROUP_CONCAT(lt.name || ': ' || l.days_count || ' jours', '; ') as leave_details
                FROM employees e
                LEFT JOIN leaves l ON e.id = l.employee_id 
                    AND strftime('%Y', date(substr(l.start_date, 7, 4) || '-' || substr(l.start_date, 4, 2) || '-' || substr(l.start_date, 1, 2))) = ?
                LEFT JOIN leave_types lt ON l.leave_type_id = lt.id
                WHERE e.status = 'Active'
                GROUP BY e.id, e.matricule, e.first_name, e.last_name
                ORDER BY e.last_name, e.first_name
            ''', (str(current_year),))
            
            leave_data = cursor.fetchall()
            conn.close()
            
            if not leave_data:
                messagebox.showwarning("Attention", "Aucune donnée de congé trouvée")
                return
                
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"rapport_conges_{current_year}_{timestamp}.pdf"
                self.create_annual_leave_pdf(leave_data, current_year, filename)
            else:  # excel
                filename = f"rapport_conges_{current_year}_{timestamp}.xlsx"
                self.create_annual_leave_excel(leave_data, current_year, filename)
                
            messagebox.showinfo("Succès", f"Rapport des congés généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_annual_leave_pdf(self, leave_data, year, filename):
        """Créer le PDF du rapport annuel des congés"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph(f"Rapport Annuel des Congés - {year}", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tableau des congés
        data = [['Matricule', 'Employé', 'Nb Congés', 'Total Jours', 'Solde Restant']]
        
        for emp_data in leave_data:
            matricule, first_name, last_name, total_leaves, total_days, leave_details = emp_data
            
            # Calcul du solde (30 jours par défaut - jours pris)
            annual_allowance = 30
            days_taken = total_days or 0
            remaining_balance = annual_allowance - days_taken
            
            data.append([
                matricule,
                f"{first_name} {last_name}",
                str(total_leaves or 0),
                str(days_taken),
                str(remaining_balance)
            ])
            
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(table)
        doc.build(story)
        
    def create_annual_leave_excel(self, leave_data, year, filename):
        """Créer le fichier Excel du rapport annuel des congés"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Congés {year}"
        
        # Titre
        ws['A1'] = f"Rapport Annuel des Congés - {year}"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        # En-têtes
        headers = ['Matricule', 'Employé', 'Nb Congés Pris', 'Total Jours Pris', 'Solde Restant']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Données
        for row, emp_data in enumerate(leave_data, 5):
            matricule, first_name, last_name, total_leaves, total_days, leave_details = emp_data
            
            # Calcul du solde
            annual_allowance = 30
            days_taken = total_days or 0
            remaining_balance = annual_allowance - days_taken
            
            ws.cell(row=row, column=1, value=matricule)
            ws.cell(row=row, column=2, value=f"{first_name} {last_name}")
            ws.cell(row=row, column=3, value=total_leaves or 0)
            ws.cell(row=row, column=4, value=days_taken)
            ws.cell(row=row, column=5, value=remaining_balance)
            
            # Colorer en rouge si solde négatif
            if remaining_balance < 0:
                ws.cell(row=row, column=5).font = Font(color='FF0000', bold=True)
                
        # Ajuster les largeurs de colonnes
        # Ajuster les largeurs de colonnes (version corrigée)
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                # Ignorer les cellules fusionnées
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value:
                        # Ajouter 2 pour un peu d'espace
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Définir une largeur minimale et maximale
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        wb.save(filename)
        
    def generate_hr_statistics_report(self, format_type):
        """Générer le rapport de statistiques RH"""
        try:
            # Récupérer toutes les statistiques
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Statistiques générales
            cursor.execute('SELECT COUNT(*) FROM employees WHERE status = "Active"')
            total_active = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM employees')
            total_employees = cursor.fetchone()[0]
            
            # Répartition par département
            cursor.execute('''
                SELECT department, COUNT(*) 
                FROM employees 
                WHERE status = "Active" AND department IS NOT NULL AND department != ""
                GROUP BY department 
                ORDER BY COUNT(*) DESC
            ''')
            dept_stats = cursor.fetchall()
            
            # Répartition par type de contrat
            cursor.execute('''
                SELECT contract_type, COUNT(*) 
                FROM employees 
                WHERE status = "Active" AND contract_type IS NOT NULL AND contract_type != ""
                GROUP BY contract_type
            ''')
            contract_stats = cursor.fetchall()
            
            # Statistiques des congés (année courante)
            current_year = datetime.now().year
            cursor.execute('''
                SELECT COUNT(*), SUM(days_count)
                FROM leaves 
                WHERE strftime('%Y', date(substr(start_date, 7, 4) || '-' || substr(start_date, 4, 2) || '-' || substr(start_date, 1, 2))) = ?
            ''', (str(current_year),))
            leave_stats = cursor.fetchone()
            
            conn.close()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"statistiques_rh_{timestamp}.pdf"
                self.create_hr_statistics_pdf(
                    total_active, total_employees, dept_stats, 
                    contract_stats, leave_stats, filename
                )
            else:  # excel
                filename = f"statistiques_rh_{timestamp}.xlsx"
                self.create_hr_statistics_excel(
                    total_active, total_employees, dept_stats, 
                    contract_stats, leave_stats, filename
                )
                
            messagebox.showinfo("Succès", f"Rapport statistiques généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_hr_statistics_pdf(self, total_active, total_employees, dept_stats, contract_stats, leave_stats, filename):
        """Créer le PDF des statistiques RH"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph("Statistiques RH - Mairie de Ngoundiane", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Statistiques générales
        story.append(Paragraph("Statistiques Générales", styles['Heading2']))
        
        general_data = [
            ['Total Employés:', str(total_employees)],
            ['Employés Actifs:', str(total_active)],
            ['Taux d\'Activité:', f"{(total_active/total_employees*100):.1f}%" if total_employees > 0 else "0%"]
        ]
        
        general_table = Table(general_data, colWidths=[200, 100])
        general_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(general_table)
        story.append(Spacer(1, 20))
        
        # Répartition par département
        if dept_stats:
            story.append(Paragraph("Répartition par Département", styles['Heading2']))
            
            dept_data = [['Département', 'Nombre d\'Employés', 'Pourcentage']]
            for dept, count in dept_stats:
                percentage = (count / total_active * 100) if total_active > 0 else 0
                dept_data.append([dept, str(count), f"{percentage:.1f}%"])
                
            dept_table = Table(dept_data)
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(dept_table)
            story.append(Spacer(1, 20))
            
        # Statistiques des congés
        story.append(Paragraph(f"Statistiques des Congés {datetime.now().year}", styles['Heading2']))
        
        total_leave_requests = leave_stats[0] or 0
        total_leave_days = leave_stats[1] or 0
        
        leave_data = [
            ['Total Demandes de Congés:', str(total_leave_requests)],
            ['Total Jours de Congés:', str(total_leave_days)],
            ['Moyenne par Demande:', f"{(total_leave_days/total_leave_requests):.1f} jours" if total_leave_requests > 0 else "0 jours"]
        ]
        
        leave_table = Table(leave_data, colWidths=[200, 100])
        leave_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(leave_table)
        
        doc.build(story)
        
    def create_hr_statistics_excel(self, total_active, total_employees, dept_stats, contract_stats, leave_stats, filename):
        """Créer le fichier Excel des statistiques RH"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques RH"
        
        # Titre
        ws['A1'] = "Statistiques RH - Mairie de Ngoundiane"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        row = 4
        
        # Statistiques générales
        ws[f'A{row}'] = "STATISTIQUES GÉNÉRALES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        ws[f'A{row}'] = "Total Employés:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_employees
        row += 1
        
        ws[f'A{row}'] = "Employés Actifs:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_active
        row += 1
        
        ws[f'A{row}'] = "Taux d'Activité:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{(total_active/total_employees*100):.1f}%" if total_employees > 0 else "0%"
        row += 3
        
        # Répartition par département
        if dept_stats:
            ws[f'A{row}'] = "RÉPARTITION PAR DÉPARTEMENT"
            ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
            row += 2
            
            # En-têtes
            headers = ['Département', 'Nombre', 'Pourcentage']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            row += 1
            
            # Données
            for dept, count in dept_stats:
                percentage = (count / total_active * 100) if total_active > 0 else 0
                ws.cell(row=row, column=1, value=dept)
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
                
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        wb.save(filename)
        
    def show_settings_module(self):
        """Module de configuration"""
        self.clear_main_content()
        self.set_active_nav_button("⚙️ Configuration")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="⚙️ Configuration du Système",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les options de configuration
        config_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        config_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Section Utilisateurs
        users_frame = tk.LabelFrame(config_frame,
                                   text="👥 Gestion des Utilisateurs",
                                   font=('Segoe UI', 14, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'],
                                   padx=20,
                                   pady=20)
        users_frame.pack(fill='x', pady=(0, 20))
        
        # Boutons de gestion des utilisateurs
        users_buttons_frame = tk.Frame(users_frame, bg=self.colors['background'])
        users_buttons_frame.pack(fill='x')
        
        add_user_btn = tk.Button(users_buttons_frame,
                                text="➕ Ajouter Utilisateur",
                                font=('Segoe UI', 11, 'bold'),
                                bg=self.colors['primary_green'],
                                fg='white',
                                relief='flat',
                                bd=0,
                                padx=20,
                                pady=10,
                                cursor='hand2',
                                command=self.add_user)
        add_user_btn.pack(side='left', padx=(0, 10))
        
        change_password_btn = tk.Button(users_buttons_frame,
                                       text="🔑 Changer Mot de Passe",
                                       font=('Segoe UI', 11, 'bold'),
                                       bg=self.colors['accent_green'],
                                       fg=self.colors['text_dark'],
                                       relief='flat',
                                       bd=0,
                                       padx=20,
                                       pady=10,
                                       cursor='hand2',
                                       command=self.change_password)
        change_password_btn.pack(side='left')
        
        # Section Base de Données
        db_frame = tk.LabelFrame(config_frame,
                                text="💾 Base de Données",
                                font=('Segoe UI', 14, 'bold'),
                                fg=self.colors['primary_green'],
                                bg=self.colors['background'],
                                padx=20,
                                pady=20)
        db_frame.pack(fill='x', pady=(0, 20))
        
        # Boutons de gestion de la base de données
        db_buttons_frame = tk.Frame(db_frame, bg=self.colors['background'])
        db_buttons_frame.pack(fill='x')
        
        backup_btn = tk.Button(db_buttons_frame,
                              text="💾 Sauvegarder",
                              font=('Segoe UI', 11, 'bold'),
                              bg=self.colors['success'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=self.backup_database)
        backup_btn.pack(side='left', padx=(0, 10))
        
        restore_btn = tk.Button(db_buttons_frame,
                               text="📁 Restaurer",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['warning'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=20,
                               pady=10,
                               cursor='hand2',
                               command=self.restore_database)
        restore_btn.pack(side='left')
        
        # Section Informations Système
        info_frame = tk.LabelFrame(config_frame,
                                  text="ℹ️ Informations Système",
                                  font=('Segoe UI', 14, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg=self.colors['background'],
                                  padx=20,
                                  pady=20)
        info_frame.pack(fill='x')
        
        # Informations système
        info_text = tk.Text(info_frame,
                           font=('Segoe UI', 10),
                           bg=self.colors['light_gray'],
                           fg=self.colors['text_dark'],
                           relief='flat',
                           height=8,
                           wrap='word',
                           state='disabled')
        info_text.pack(fill='both', expand=True)
        
        # Remplir les informations système
        system_info = f"""
Version de l'Application: 1.0.0
Base de Données: SQLite ({self.db_path})
Dossier Documents: {os.path.abspath(self.documents_folder)}
Dossier Photos: {os.path.abspath(self.photos_folder)}
Dossier Courriers: {os.path.abspath(self.courriers_folder)}

Statistiques:
- Nombre total d'employés: {self.get_total_employees()}
- Nombre d'utilisateurs: {self.get_total_users()}
- Taille de la base de données: {self.get_db_size()}

Développé pour la Mairie de Ngoundiane
© 2024 - Système de Gestion RH
        """
        
        info_text.config(state='normal')
        info_text.insert('1.0', system_info.strip())
        info_text.config(state='disabled')
        
    def add_user(self):
        """Ajouter un nouvel utilisateur"""
        # Fenêtre de saisie
        user_window = tk.Toplevel(self.root)
        user_window.title("Ajouter Utilisateur")
        user_window.geometry("400x300")
        user_window.configure(bg=self.colors['background'])
        user_window.transient(self.root)
        user_window.grab_set()
        
        # Variables
        username_var = tk.StringVar()
        password_var = tk.StringVar()
        role_var = tk.StringVar(value='user')
        
        # Titre
        title = tk.Label(user_window,
                        text="👤 Nouvel Utilisateur",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(user_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=30)
        
        # Nom d'utilisateur
        tk.Label(form_frame,
                text="Nom d'utilisateur:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        username_entry = tk.Entry(form_frame,
                                 textvariable=username_var,
                                 font=('Segoe UI', 11),
                                 width=30,
                                 relief='solid',
                                 bd=1)
        username_entry.pack(fill='x', pady=(0, 15))
        
        # Mot de passe
        tk.Label(form_frame,
                text="Mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        password_entry = tk.Entry(form_frame,
                                 textvariable=password_var,
                                 font=('Segoe UI', 11),
                                 width=30,
                                 show='*',
                                 relief='solid',
                                 bd=1)
        password_entry.pack(fill='x', pady=(0, 15))
        
        # Rôle
        tk.Label(form_frame,
                text="Rôle:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        role_combo = ttk.Combobox(form_frame,
                                 textvariable=role_var,
                                 values=['user', 'admin'],
                                 font=('Segoe UI', 11),
                                 width=27,
                                 state='readonly')
        role_combo.pack(fill='x', pady=(0, 20))
        
        # Boutons
        buttons_frame = tk.Frame(user_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=30, pady=(0, 20))
        
        def save_user():
            username = username_var.get().strip()
            password = password_var.get().strip()
            role = role_var.get()
            
            if not username or not password:
                messagebox.showerror("Erreur", "Tous les champs sont obligatoires")
                return
                
            # Vérifier l'unicité
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
            
            if cursor.fetchone():
                messagebox.showerror("Erreur", "Ce nom d'utilisateur existe déjà")
                conn.close()
                return
                
            # Créer l'utilisateur
            try:
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                cursor.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                              (username, password_hash, role))
                conn.commit()
                messagebox.showinfo("Succès", "Utilisateur créé avec succès")
                user_window.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur lors de la création: {str(e)}")
            finally:
                conn.close()
                
        save_btn = tk.Button(buttons_frame,
                            text="💾 Créer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=save_user)
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=user_window.destroy)
        cancel_btn.pack(side='right')
        
        username_entry.focus()
        
    def change_password(self):
        """Changer le mot de passe de l'utilisateur actuel"""
        # Fenêtre de changement de mot de passe
        pwd_window = tk.Toplevel(self.root)
        pwd_window.title("Changer Mot de Passe")
        pwd_window.geometry("400x250")
        pwd_window.configure(bg=self.colors['background'])
        pwd_window.transient(self.root)
        pwd_window.grab_set()
        
        # Variables
        current_pwd_var = tk.StringVar()
        new_pwd_var = tk.StringVar()
        confirm_pwd_var = tk.StringVar()
        
        # Titre
        title = tk.Label(pwd_window,
                        text="🔑 Changer Mot de Passe",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(pwd_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=30)
        
        # Mot de passe actuel
        tk.Label(form_frame,
                text="Mot de passe actuel:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        current_pwd_entry = tk.Entry(form_frame,
                                    textvariable=current_pwd_var,
                                    font=('Segoe UI', 11),
                                    width=30,
                                    show='*',
                                    relief='solid',
                                    bd=1)
        current_pwd_entry.pack(fill='x', pady=(0, 10))
        
        # Nouveau mot de passe
        tk.Label(form_frame,
                text="Nouveau mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        new_pwd_entry = tk.Entry(form_frame,
                                textvariable=new_pwd_var,
                                font=('Segoe UI', 11),
                                width=30,
                                show='*',
                                relief='solid',
                                bd=1)
        new_pwd_entry.pack(fill='x', pady=(0, 10))
        
        # Confirmer nouveau mot de passe
        tk.Label(form_frame,
                text="Confirmer nouveau mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        confirm_pwd_entry = tk.Entry(form_frame,
                                    textvariable=confirm_pwd_var,
                                    font=('Segoe UI', 11),
                                    width=30,
                                    show='*',
                                    relief='solid',
                                    bd=1)
        confirm_pwd_entry.pack(fill='x', pady=(0, 15))
        
        # Boutons
        buttons_frame = tk.Frame(pwd_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=30, pady=(0, 20))
        
        def save_password():
            current_pwd = current_pwd_var.get()
            new_pwd = new_pwd_var.get()
            confirm_pwd = confirm_pwd_var.get()
            
            if not all([current_pwd, new_pwd, confirm_pwd]):
                messagebox.showerror("Erreur", "Tous les champs sont obligatoires")
                return
                
            if new_pwd != confirm_pwd:
                messagebox.showerror("Erreur", "Les nouveaux mots de passe ne correspondent pas")
                return
                
            if len(new_pwd) < 4:
                messagebox.showerror("Erreur", "Le mot de passe doit contenir au moins 4 caractères")
                return
                
            # Vérifier le mot de passe actuel
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            current_pwd_hash = hashlib.sha256(current_pwd.encode()).hexdigest()
            cursor.execute('SELECT id FROM users WHERE id = ? AND password_hash = ?',
                          (self.current_user['id'], current_pwd_hash))
            
            if not cursor.fetchone():
                messagebox.showerror("Erreur", "Mot de passe actuel incorrect")
                conn.close()
                return
                
            # Mettre à jour le mot de passe
            try:
                new_pwd_hash = hashlib.sha256(new_pwd.encode()).hexdigest()
                cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?',
                              (new_pwd_hash, self.current_user['id']))
                conn.commit()
                messagebox.showinfo("Succès", "Mot de passe modifié avec succès")
                pwd_window.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur lors de la modification: {str(e)}")
            finally:
                conn.close()
                
        save_btn = tk.Button(buttons_frame,
                            text="💾 Modifier",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=save_password)
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=pwd_window.destroy)
        cancel_btn.pack(side='right')
        
        current_pwd_entry.focus()
        
    def backup_database(self):
        """Sauvegarder la base de données"""
        try:
            # Sélectionner le dossier de destination
            backup_path = filedialog.asksaveasfilename(
                title="Sauvegarder la base de données",
                defaultextension=".db",
                filetypes=[("Base de données SQLite", "*.db"), ("Tous les fichiers", "*.*")],
                initialname=f"hr_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            )
            
            if backup_path:
                # Copier la base de données
                shutil.copy2(self.db_path, backup_path)
                messagebox.showinfo("Succès", f"Base de données sauvegardée avec succès:\n{backup_path}")
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la sauvegarde: {str(e)}")
            
    def restore_database(self):
        """Restaurer la base de données"""
        if messagebox.askyesno("Confirmation", 
                              "Êtes-vous sûr de vouloir restaurer la base de données ?\n\nCette action remplacera toutes les données actuelles."):
            
            try:
                # Sélectionner le fichier de sauvegarde
                backup_file = filedialog.askopenfilename(
                    title="Sélectionner la sauvegarde à restaurer",
                    filetypes=[("Base de données SQLite", "*.db"), ("Tous les fichiers", "*.*")]
                )
                
                if backup_file:
                    # Remplacer la base de données actuelle
                    shutil.copy2(backup_file, self.db_path)
                    messagebox.showinfo("Succès", "Base de données restaurée avec succès.\n\nL'application va redémarrer.")
                    
                    # Redémarrer l'application
                    self.root.quit()
                    
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la restauration: {str(e)}")
                
    def get_total_employees(self):
        """Obtenir le nombre total d'employés"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM employees')
            count = cursor.fetchone()[0]
            conn.close()
            return count
        except:
            return 0
            
    def get_total_users(self):
        """Obtenir le nombre total d'utilisateurs"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users')
            count = cursor.fetchone()[0]
            conn.close()
            return count
        except:
            return 0
            
    def get_db_size(self):
        """Obtenir la taille de la base de données"""
        try:
            size = os.path.getsize(self.db_path)
            if size < 1024:
                return f"{size} bytes"
            elif size < 1024 * 1024:
                return f"{size / 1024:.1f} KB"
            else:
                return f"{size / (1024 * 1024):.1f} MB"
        except:
            return "Inconnue"
            
    def clear_main_content(self):
        """Nettoyer le contenu principal"""
        for widget in self.main_content.winfo_children():
            widget.destroy()
            
    def logout(self):
        """Déconnexion"""
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir vous déconnecter ?"):
            self.current_user = None
            self.current_employee_id = None
            self.show_login_screen()
            
    def run(self):
        """Lancer l'application"""
        # Centrer la fenêtre
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
        
        # Démarrer la boucle principale
        self.root.mainloop()

# Point d'entrée de l'application
if __name__ == "__main__":
    try:
        app = HRManagementApp()
        app.run()
    except Exception as e:
        print(f"Erreur lors du démarrage de l'application: {str(e)}")
        input("Appuyez sur Entrée pour fermer...")