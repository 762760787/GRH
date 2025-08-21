import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import hashlib
import os
import shutil
import sys # Ajout important
from datetime import datetime, timedelta
import calendar
from PIL import Image, ImageTk
import subprocess
import platform
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import pytesseract
from pdf2image import convert_from_path

# --- CODE SPÉCIFIQUE À WINDOWS POUR L'ICÔNE DE LA BARRE DES TÂCHES ---
# Doit être exécuté avant la création de la fenêtre principale Tk()
if platform.system() == "Windows":
    try:
        import ctypes
        myappid = 'YoonuRH.App.1' 
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except (ImportError, AttributeError):
        pass # Ne fait rien si ctypes n'est pas disponible ou si l'appel échoue


class HRManagementApp:

    def __init__(self):
        """Initialisation de l'application"""
        self.root = tk.Tk()
        # Rétablir la barre de titre standard de Windows
        self.root.title("YoonuRH")

        # Déterminer le chemin de base de manière fiable
        if getattr(sys, 'frozen', False):
            # Si l'application est "gelée" (ex: avec PyInstaller)
            base_dir = os.path.dirname(sys.executable)
        else:
            # En mode script normal
            base_dir = os.path.dirname(os.path.abspath(__file__))

        # --- GESTION DE L'ICÔNE ---
        ico_icon_path = os.path.join(base_dir, 'mairie_icon.ico')
        try:
            if os.path.exists(ico_icon_path):
                self.root.iconbitmap(ico_icon_path)
        except Exception as e:
            print(f"Erreur lors du chargement de l'icône: {e}")

        # --- GESTION DE LA TAILLE ET DE LA RESPONSIVITÉ ---
        # Démarrer en mode maximisé pour une meilleure expérience
        if platform.system() == "Windows":
            self.root.state('zoomed')
        else: # Pour macOS et Linux
            self.root.geometry("{0}x{1}+0+0".format(self.root.winfo_screenwidth(), self.root.winfo_screenheight()))
        
        self.root.minsize(1200, 800)

        # Palette de couleurs
        self.colors = {
            'primary_green': '#2E7D32', 'light_green': '#4CAF50', 'dark_green': '#1B5E20',
            'accent_green': '#81C784', 'background': '#F8F9FA', 'white': '#FFFFFF',
            'light_gray': '#E8F5E8', 'text_dark': '#2C3E50', 'text_light': '#7F8C8D',
            'error': '#E74C3C', 'warning': '#F39C12', 'success': '#27AE60'
        }

        # --- CHEMINS ABSOLUS POUR LES FICHIERS ET DOSSIERS ---
        self.current_user = None
        self.current_employee_id = None
        self.db_path = os.path.join(base_dir, "hr_database.db")
        self.documents_folder = os.path.join(base_dir, "documents")
        self.photos_folder = os.path.join(base_dir, "photos")
        self.courriers_folder = os.path.join(base_dir, "courriers_files")
        
        try:
            pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
            # Test pour voir si Tesseract est accessible
            pytesseract.get_tesseract_version()
        except Exception as e:
            print("AVERTISSEMENT TESSERACT : Le moteur OCR n'a pas été trouvé ou configuré correctement.")
            print("Veuillez installer Tesseract-OCR et vérifier le chemin dans la fonction __init__.")
            print(f"Erreur: {e}")

        # Créer les dossiers nécessaires s'ils n'existent pas
        os.makedirs(self.documents_folder, exist_ok=True)
        os.makedirs(self.photos_folder, exist_ok=True)
        os.makedirs(self.courriers_folder, exist_ok=True)
        
        # Configuration du style
        self.setup_styles()
        
        # Initialisation de la base de données
        self.init_database()
        
        # Démarrage avec l'écran de connexion
        self.show_login_screen()

    def setup_styles(self):
        """Configuration des styles visuels modernes"""
        style = ttk.Style()
        style.theme_use('clam')
        
        style.configure('Primary.TButton', background=self.colors['primary_green'], foreground='white', borderwidth=0, focuscolor='none', padding=(20, 10))
        style.map('Primary.TButton', background=[('active', self.colors['light_green']), ('pressed', self.colors['dark_green'])])
        style.configure('Secondary.TButton', background=self.colors['accent_green'], foreground=self.colors['text_dark'], borderwidth=1, focuscolor='none', padding=(15, 8))
        style.configure('Title.TLabel', background=self.colors['background'], foreground=self.colors['primary_green'], font=('Segoe UI', 16, 'bold'))
        style.configure('Section.TLabel', background=self.colors['background'], foreground=self.colors['text_dark'], font=('Segoe UI', 12, 'bold'))
        style.configure('Custom.TNotebook', background=self.colors['background'], borderwidth=0)
        style.configure('Custom.TNotebook.Tab', background=self.colors['light_gray'], foreground=self.colors['text_dark'], padding=(20, 10), font=('Segoe UI', 10, 'bold'))
        style.map('Custom.TNotebook.Tab', background=[('selected', self.colors['primary_green']), ('active', self.colors['accent_green'])], foreground=[('selected', 'white')])
        style.configure('Custom.Treeview', background='white', foreground=self.colors['text_dark'], rowheight=30, fieldbackground='white', font=('Segoe UI', 10))
        style.configure('Custom.Treeview.Heading', background=self.colors['primary_green'], foreground='white', font=('Segoe UI', 10, 'bold'))

    def init_database(self):
        """Initialisation de la base de données SQLite"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Table des utilisateurs
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table des employés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS employees (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                matricule TEXT UNIQUE NOT NULL,
                first_name TEXT NOT NULL,
                last_name TEXT NOT NULL,
                gender TEXT,
                birth_date TEXT,
                birth_place TEXT,
                address TEXT,
                phone TEXT,
                email TEXT,
                marital_status TEXT,
                dependents INTEGER DEFAULT 0,
                social_security TEXT,
                bank_details TEXT,
                hire_date TEXT,
                contract_type TEXT,
                contract_start TEXT,
                contract_end TEXT,
                department TEXT,
                job_title TEXT,
                status TEXT DEFAULT 'Active',
                photo_path TEXT,
                numero_decision TEXT, -- Champ ajouté
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table de l'historique de carrière
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS career_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                act_number TEXT,
                nature TEXT,
                subject TEXT,
                act_date TEXT,
                effective_date TEXT,
                document_path TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id)
            )
        ''')
        
        # Table des documents
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                category TEXT,
                name TEXT,
                file_path TEXT,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id)
            )
        ''')
        
        # Table des types de congés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leave_types (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                days_per_year INTEGER DEFAULT 0,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Table des congés
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS leaves (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                leave_type_id INTEGER,
                start_date TEXT,
                end_date TEXT,
                days_count INTEGER,
                status TEXT DEFAULT 'Approved',
                notes TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES employees (id),
                FOREIGN KEY (leave_type_id) REFERENCES leave_types (id)
            )
        ''')
        
        # Table des courriers - MISE À JOUR avec colonne pour fichiers
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS courriers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                numero_ordre TEXT UNIQUE NOT NULL,
                type_courrier TEXT NOT NULL,
                nombre_pieces INTEGER DEFAULT 1,
                date_arrivee_expedition TEXT NOT NULL,
                expediteur_destinataire TEXT NOT NULL,
                objet TEXT NOT NULL,
                numero_archive TEXT,
                observation TEXT,
                file_path TEXT,
                date_creation TEXT DEFAULT CURRENT_TIMESTAMP,
                created_by TEXT
            )
        ''')
        # AJOUTEZ CE BLOC DE CODE ICI
        # Ajouter les colonnes CNI et nationalité si elles n'existent pas
        try:
            cursor.execute('ALTER TABLE employees ADD COLUMN cni TEXT')
            cursor.execute('ALTER TABLE employees ADD COLUMN nationalite TEXT')
        except sqlite3.OperationalError:
            # Les colonnes existent déjà
            pass
        # Ajouter la colonne file_path si elle n'existe pas déjà
        try:
            cursor.execute('ALTER TABLE courriers ADD COLUMN file_path TEXT')
        except sqlite3.OperationalError:
            # La colonne existe déjà
            pass
        try:
            cursor.execute('ALTER TABLE employees ADD COLUMN numero_decision TEXT')
        except sqlite3.OperationalError:
            # La colonne existe déjà
            pass

        # Insérer des utilisateurs par défaut
        try:
            admin_hash = hashlib.sha256('admin'.encode()).hexdigest()
            user_hash = hashlib.sha256('user'.encode()).hexdigest()
            
            cursor.execute('INSERT OR IGNORE INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                          ('admin', admin_hash, 'admin'))
            cursor.execute('INSERT OR IGNORE INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                          ('user', user_hash, 'user'))
        except sqlite3.IntegrityError:
            pass
        
        # Insérer des types de congés par défaut
        default_leave_types = [
            ('Congé Annuel', 30, 'Congé annuel réglementaire'),
            ('Congé Maladie', 0, 'Congé pour maladie'),
            ('Congé Maternité', 0, 'Congé de maternité'),
            ('Congé Paternité', 0, 'Congé de paternité'),
            ('Permission Exceptionnelle', 0, 'Permission pour événements familiaux')
        ]
        
        for leave_type in default_leave_types:
            cursor.execute('INSERT OR IGNORE INTO leave_types (name, days_per_year, description) VALUES (?, ?, ?)',
                          leave_type)
        
        conn.commit()
        conn.close()
        
    def show_login_screen(self):
        """Affichage de l'écran de connexion (design moderne)"""
        # Nettoyer la fenêtre
        for widget in self.root.winfo_children():
            widget.destroy()

        self.root.configure(bg=self.colors['background'])
        self.root.bind('<Return>', lambda e: self.login())

        # --- Conteneur Principal ---
        main_frame = tk.Frame(self.root, bg=self.colors['background'])
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- Panneau Gauche (Branding) ---
        left_panel = tk.Frame(main_frame, bg=self.colors['primary_green'])
        left_panel.place(relx=0, rely=0, relwidth=0.4, relheight=1)

        # Titre et description dans le panneau gauche
        tk.Label(left_panel, text="YoonuRH", font=('Segoe UI', 36, 'bold'), fg='white', bg=self.colors['primary_green']).pack(pady=(150, 20))
        tk.Label(left_panel, text="Système de Gestion\ndes Ressources Humaines", font=('Segoe UI', 18), fg=self.colors['light_gray'], bg=self.colors['primary_green']).pack(pady=10)
        tk.Label(left_panel, text="© 2024", font=('Segoe UI', 10), fg=self.colors['accent_green'], bg=self.colors['primary_green']).pack(side=tk.BOTTOM, pady=20)

        # --- Panneau Droit (Formulaire de Connexion) ---
        right_panel = tk.Frame(main_frame, bg=self.colors['background'])
        right_panel.place(relx=0.4, rely=0, relwidth=0.6, relheight=1)

        # Frame pour centrer le contenu
        login_container = tk.Frame(right_panel, bg=self.colors['background'])
        login_container.place(relx=0.5, rely=0.5, anchor='center')

        tk.Label(login_container, text="Authentification", font=('Segoe UI', 28, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background']).pack(pady=(0, 40))

        # --- Champ Nom d'utilisateur ---
        user_frame = tk.Frame(login_container, bg=self.colors['background'])
        user_frame.pack(pady=10)
        tk.Label(user_frame, text="👤", font=('Segoe UI', 14), fg=self.colors['text_light'], bg=self.colors['background']).pack(side=tk.LEFT, padx=(0,10))
        self.username_entry = tk.Entry(user_frame, font=('Segoe UI', 14), width=25, relief='flat', bg=self.colors['light_gray'])
        self.username_entry.pack(side=tk.LEFT)
        self.username_entry.insert(0, "Nom d'utilisateur")
        self.username_entry.config(fg=self.colors['text_light'])

        # --- Champ Mot de passe ---
        pass_frame = tk.Frame(login_container, bg=self.colors['background'])
        pass_frame.pack(pady=10)
        tk.Label(pass_frame, text="🔑", font=('Segoe UI', 14), fg=self.colors['text_light'], bg=self.colors['background']).pack(side=tk.LEFT, padx=(0,10))
        self.password_entry = tk.Entry(pass_frame, font=('Segoe UI', 14), width=25, relief='flat', bg=self.colors['light_gray'])
        self.password_entry.pack(side=tk.LEFT)
        self.password_entry.insert(0, "Mot de passe")
        self.password_entry.config(fg=self.colors['text_light'])

        # --- Logique pour les placeholders ---
        def on_user_click(event):
            if self.username_entry.get() == "Nom d'utilisateur":
                self.username_entry.delete(0, tk.END)
                self.username_entry.config(fg=self.colors['text_dark'])
        def on_user_leave(event):
            if self.username_entry.get() == '':
                self.username_entry.insert(0, "Nom d'utilisateur")
                self.username_entry.config(fg=self.colors['text_light'])

        def on_pass_click(event):
            if self.password_entry.get() == "Mot de passe":
                self.password_entry.delete(0, tk.END)
                self.password_entry.config(fg=self.colors['text_dark'], show='*')
        def on_pass_leave(event):
            if self.password_entry.get() == '':
                self.password_entry.insert(0, "Mot de passe")
                self.password_entry.config(fg=self.colors['text_light'], show='')

        self.username_entry.bind('<FocusIn>', on_user_click)
        self.username_entry.bind('<FocusOut>', on_user_leave)
        self.password_entry.bind('<FocusIn>', on_pass_click)
        self.password_entry.bind('<FocusOut>', on_pass_leave)

        # --- Bouton de connexion ---
        login_btn = tk.Button(login_container,
                             text="Se Connecter",
                             font=('Segoe UI', 14, 'bold'),
                             bg=self.colors['primary_green'],
                             fg='white',
                             relief='flat',
                             bd=0,
                             padx=40,
                             pady=12,
                             cursor='hand2',
                             command=self.login)
        login_btn.pack(pady=40)

        # --- Animation du bouton au survol ---
        def on_enter(e):
            login_btn.config(bg=self.colors['light_green'])
        def on_leave(e):
            login_btn.config(bg=self.colors['primary_green'])

        login_btn.bind("<Enter>", on_enter)
        login_btn.bind("<Leave>", on_leave)

        # --- Info utilisateurs ---
        # tk.Label(login_container,
        #          text="Utilisateurs par défaut: admin/admin ou user/user",
        #          font=('Segoe UI', 9),
        #          fg=self.colors['text_light'],
        #          bg=self.colors['background']).pack()

        # Focus initial
        # self.root.after(100, lambda: self.root.focus_force())

    def login(self):
        """Gestion de la connexion utilisateur"""
        username = self.username_entry.get().strip()
        password = self.password_entry.get().strip()
        
        if not username or not password:
            messagebox.showerror("Erreur", "Veuillez saisir le nom d'utilisateur et le mot de passe")
            return
            
        # Vérification dans la base de données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        cursor.execute('SELECT id, username, role FROM users WHERE username = ? AND password_hash = ?',
                      (username, password_hash))
        
        user = cursor.fetchone()
        conn.close()
        
        if user:
            self.current_user = {
                'id': user[0],
                'username': user[1],
                'role': user[2]
            }
            self.show_main_dashboard()
        else:
            messagebox.showerror("Erreur", "Nom d'utilisateur ou mot de passe incorrect")
            self.password_entry.delete(0, tk.END)
            
    def show_main_dashboard(self):
        """Affichage du tableau de bord principal"""
        # Nettoyer la fenêtre
        for widget in self.root.winfo_children():
            widget.destroy()
            
        self.root.configure(bg=self.colors['background'])
        
        # Frame principal
        main_container = tk.Frame(self.root, bg=self.colors['background'])
        main_container.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Header avec titre et info utilisateur
        header_frame = tk.Frame(main_container, bg=self.colors['primary_green'], height=80)
        header_frame.pack(fill='x', pady=(0, 10))
        header_frame.pack_propagate(False)
        
        # Titre principal
        title_label = tk.Label(header_frame,
                              text="🏛️ Système de Gestion RH ",
                              font=('Segoe UI', 20, 'bold'),
                              fg='white',
                              bg=self.colors['primary_green'])
        title_label.pack(side='left', padx=20, pady=20)
        
        # Info utilisateur et déconnexion
        user_frame = tk.Frame(header_frame, bg=self.colors['primary_green'])
        user_frame.pack(side='right', padx=20, pady=20)
        
        user_label = tk.Label(user_frame,
                             text=f"👤 {self.current_user['username']} ({self.current_user['role']})",
                             font=('Segoe UI', 12),
                             fg='white',
                             bg=self.colors['primary_green'])
        user_label.pack(side='left', padx=(0, 15))
        
        logout_btn = tk.Button(user_frame,
                              text="Déconnexion",
                              font=('Segoe UI', 10),
                              bg=self.colors['error'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2',
                              command=self.logout)
        logout_btn.pack(side='right')
        
        # Container pour le contenu principal
        content_container = tk.Frame(main_container, bg=self.colors['background'])
        content_container.pack(fill='both', expand=True)
        
        # Sidebar pour la navigation
        sidebar = tk.Frame(content_container, bg=self.colors['white'], width=250)
        sidebar.pack(side='left', fill='y', padx=(0, 10))
        sidebar.pack_propagate(False)
        
        # Titre sidebar
        sidebar_title = tk.Label(sidebar,
                                text="📋 MODULES",
                                font=('Segoe UI', 14, 'bold'),
                                fg=self.colors['primary_green'],
                                bg=self.colors['white'])
        sidebar_title.pack(pady=(20, 10))
        
        # Boutons de navigation
        nav_buttons = [
            ("👥 Tableau de Bord", self.show_dashboard_content),
            ("📁 Gestion Employés", self.show_employees_module),
            ("🏖️ Gestion Congés", self.show_leaves_module),
            ("📮 Gestion Courriers", self.show_mail_module),
            ("✍️ OCR - Extraire Texte", self.show_ocr_module), # <--- AJOUTER CETTE LIGNE
            ("📊 Rapports", self.show_reports_module),
            ("⚙️ Configuration", self.show_settings_module)
        ]
        
        self.nav_buttons = {}
        for text, command in nav_buttons:
            btn = tk.Button(sidebar,
                           text=text,
                           font=('Segoe UI', 11),
                           bg=self.colors['light_gray'],
                           fg=self.colors['text_dark'],
                           relief='flat',
                           bd=0,
                           padx=20,
                           pady=12,
                           width=25,
                           anchor='w',
                           cursor='hand2',
                           command=command)
            btn.pack(fill='x', padx=10, pady=2)
            self.nav_buttons[text] = btn
            
            # Effet hover
            def on_enter(e, button=btn):
                if button['bg'] != self.colors['primary_green']:
                    button.configure(bg=self.colors['accent_green'])
            def on_leave(e, button=btn):
                if button['bg'] != self.colors['primary_green']:
                    button.configure(bg=self.colors['light_gray'])
                    
            btn.bind("<Enter>", on_enter)
            btn.bind("<Leave>", on_leave)
        
        # Zone de contenu principal
        self.main_content = tk.Frame(content_container, bg=self.colors['background'])
        self.main_content.pack(side='right', fill='both', expand=True)
        
        # Afficher le contenu du tableau de bord par défaut
        self.show_dashboard_content()
        self.set_active_nav_button("👥 Tableau de Bord")
        
    def set_active_nav_button(self, active_text):
        """Met en évidence le bouton de navigation actif"""
        for text, btn in self.nav_buttons.items():
            if text == active_text:
                btn.configure(bg=self.colors['primary_green'], fg='white')
            else:
                btn.configure(bg=self.colors['light_gray'], fg=self.colors['text_dark'])

    def show_dashboard_content(self):
        """Affichage du contenu du tableau de bord - MISE À JOUR avec statistiques mensuelles"""
        self.clear_main_content()
        self.set_active_nav_button("👥 Tableau de Bord")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📊 Tableau de Bord",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les statistiques
        stats_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        stats_frame.pack(fill='x', padx=20)
        
        # Récupérer les statistiques
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Total employés actifs
        cursor.execute('SELECT COUNT(*) FROM employees WHERE status = "Active"')
        total_employees = cursor.fetchone()[0]
        
        # --- CORRECTION : Employés en congé ce mois ---
        now = datetime.now()
        # Début du mois au format YYYY-MM-DD pour une comparaison correcte
        month_start_iso = now.replace(day=1).strftime('%Y-%m-%d')
        
        # Fin du mois au format YYYY-MM-DD
        # On trouve le premier jour du mois suivant, puis on enlève un jour
        next_month_start_dt = (now.replace(day=28) + timedelta(days=4)).replace(day=1)
        month_end_dt = next_month_start_dt - timedelta(days=1)
        month_end_iso = month_end_dt.strftime('%Y-%m-%d')

        # La requête est corrigée pour utiliser la fonction date() de SQLite
        # ce qui garantit une comparaison chronologique correcte des dates.
        cursor.execute('''
            SELECT COUNT(DISTINCT e.id) FROM employees e
            JOIN leaves l ON e.id = l.employee_id
            WHERE l.status = "Approved"
            AND date(substr(l.start_date, 7, 4) || '-' || substr(l.start_date, 4, 2) || '-' || substr(l.start_date, 1, 2)) <= ?
            AND date(substr(l.end_date, 7, 4) || '-' || substr(l.end_date, 4, 2) || '-' || substr(l.end_date, 1, 2)) >= ?
        ''', (month_end_iso, month_start_iso))
        employees_on_leave_month = cursor.fetchone()[0]
        
        # Anniversaires ce mois
        current_month_str = now.strftime('%m')
        cursor.execute('''
            SELECT COUNT(*) FROM employees 
            WHERE substr(birth_date, 4, 2) = ? AND status = "Active"
        ''', (current_month_str,))
        birthdays_this_month = cursor.fetchone()[0]
        
        # Total courriers arrivée
        cursor.execute("SELECT COUNT(*) FROM courriers WHERE type_courrier = 'arrivee'")
        total_arrival_mail = cursor.fetchone()[0]

        # Total courriers départ
        cursor.execute("SELECT COUNT(*) FROM courriers WHERE type_courrier = 'depart'")
        total_departure_mail = cursor.fetchone()[0]
        
        conn.close()
        
        # Cartes de statistiques mises à jour
        stats_data = [
            ("👥 Total Employés Actifs", total_employees, self.colors['primary_green']),
            ("🏖️ En Congé ce Mois", employees_on_leave_month, self.colors['warning']),
            ("🎂 Anniversaires ce Mois", birthdays_this_month, '#ff6b6b'),
            ("📥 Courriers Arrivée", total_arrival_mail, '#6f42c1'),
            ("📤 Courriers Départ", total_departure_mail, '#17a2b8'),
        ]
        
        for i, (card_title_text, value, color) in enumerate(stats_data):
            col = i % 3
            row = i // 3
            card = tk.Frame(stats_frame, bg='white', relief='solid', bd=1, highlightbackground=self.colors['light_gray'], highlightthickness=1)
            card.grid(row=row, column=col, padx=15, pady=10, sticky='nsew')
            
            card_title = tk.Label(card,
                                 text=card_title_text,
                                 font=('Segoe UI', 12, 'bold'),
                                 fg=self.colors['text_dark'],
                                 bg='white')
            card_title.pack(pady=(15, 5), padx=10)
            
            card_value = tk.Label(card,
                                 text=str(value),
                                 font=('Segoe UI', 28, 'bold'),
                                 fg=color,
                                 bg='white')
            card_value.pack(pady=(0, 20), padx=10)
            
        # Configurer les colonnes pour qu'elles s'étendent uniformément
        for i in range(3):
            stats_frame.grid_columnconfigure(i, weight=1)
            
        # Section des alertes
        alerts_frame = tk.LabelFrame(self.main_content,
                                   text="🚨 Alertes et Notifications",
                                   font=('Segoe UI', 14, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'],
                                   padx=10,
                                   pady=10)
        alerts_frame.pack(fill='both', expand=True, padx=20, pady=(30, 20))
        
        alerts_text = tk.Text(alerts_frame,
                             font=('Segoe UI', 11),
                             bg='white',
                             fg=self.colors['text_dark'],
                             relief='flat',
                             wrap='word',
                             state='disabled')
        alerts_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        alerts_content = []
        
        if birthdays_this_month > 0:
            alerts_content.append(f"🎂 {birthdays_this_month} anniversaire(s) à souhaiter ce mois-ci.")
            
        if employees_on_leave_month > 0:
            alerts_content.append(f"🏖️ {employees_on_leave_month} employé(s) sont en congé durant ce mois.")
            
        if not alerts_content:
            alerts_content.append("✅ Aucune alerte importante pour le moment.")
            
        alerts_text.config(state='normal')
        alerts_text.delete('1.0', tk.END)
        for alert in alerts_content:
            alerts_text.insert(tk.END, "• " + alert + "\n\n")
        alerts_text.config(state='disabled')

    def show_employees_module(self):
        """Module de gestion des employés"""
        self.clear_main_content()
        self.set_active_nav_button("📁 Gestion Employés")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="👥 Gestion des Employés",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Barre d'outils
        toolbar = tk.Frame(self.main_content, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=(0, 10))
        
        # Boutons d'action
        add_btn = tk.Button(toolbar,
                           text="➕ Nouvel Employé",
                           font=('Segoe UI', 11, 'bold'),
                           bg=self.colors['primary_green'],
                           fg='white',
                           relief='flat',
                           bd=0,
                           padx=15,
                           pady=8,
                           cursor='hand2',
                           command=self.add_new_employee)
        add_btn.pack(side='left', padx=(0, 10))
        
        # Champ de recherche
        search_frame = tk.Frame(toolbar, bg=self.colors['background'])
        search_frame.pack(side='right')
        
        tk.Label(search_frame,
                text="🔍 Rechercher:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(side='left', padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.filter_employees)
        search_entry = tk.Entry(search_frame,
                               textvariable=self.search_var,
                               font=('Segoe UI', 11),
                               width=25,
                               relief='solid',
                               bd=1)
        search_entry.pack(side='left')
        
        # Liste des employés
        list_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
         # NOUVELLE VERSION
        # NOUVELLE VERSION CORRIGÉE
        columns = ('Photo', 'Matricule', 'Nom Complet', "Corps de l'agent", 'Division', 'Statut')
        self.employees_tree = ttk.Treeview(list_frame,
                                          columns=columns,
                                          show='headings',
                                          style='Custom.Treeview',
                                          height=15)
        
        # Configuration des colonnes
        self.employees_tree.heading('Photo', text='Photo')
        self.employees_tree.heading('Matricule', text='Matricule')
        self.employees_tree.heading('Nom Complet', text='Nom Complet')
        self.employees_tree.heading("Corps de l'agent", text="Corps de l'agent")
        self.employees_tree.heading('Division', text='Division')
        self.employees_tree.heading('Statut', text='Statut')
        
        self.employees_tree.column('Photo', width=80, anchor='center')
        self.employees_tree.column('Matricule', width=100, anchor='center')
        self.employees_tree.column('Nom Complet', width=200, anchor='w')
        self.employees_tree.column("Corps de l'agent", width=150, anchor='w')
        self.employees_tree.column('Division', width=150, anchor='w')
        self.employees_tree.column('Statut', width=100, anchor='center')
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.employees_tree.yview)
        h_scrollbar = ttk.Scrollbar(list_frame, orient='horizontal', command=self.employees_tree.xview)
        self.employees_tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Placement
        self.employees_tree.grid(row=0, column=0, sticky='nsew')
        v_scrollbar.grid(row=0, column=1, sticky='ns')
        h_scrollbar.grid(row=1, column=0, sticky='ew')
        
        list_frame.grid_rowconfigure(0, weight=1)
        list_frame.grid_columnconfigure(0, weight=1)
        
        # Double-clic pour ouvrir le dossier employé
        self.employees_tree.bind('<Double-1>', self.open_employee_file)
        
        # Menu contextuel
        self.create_employee_context_menu()
        
        # Charger les employés
        self.load_employees()
        
    def create_employee_context_menu(self):
        """Créer le menu contextuel pour la liste des employés"""
        self.employee_context_menu = tk.Menu(self.root, tearoff=0)
        self.employee_context_menu.add_command(label="📂 Ouvrir le dossier", command=self.open_employee_file)
        self.employee_context_menu.add_command(label="✏️ Modifier", command=self.edit_employee)
        self.employee_context_menu.add_separator()
        self.employee_context_menu.add_command(label="🗑️ Supprimer", command=self.delete_employee)
        
        def show_context_menu(event):
            try:
                self.employee_context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                self.employee_context_menu.grab_release()
                
        self.employees_tree.bind("<Button-3>", show_context_menu)  # Clic droit
        
    def load_employees(self):
        """Charger la liste des employés"""
        # Vider la liste actuelle
        for item in self.employees_tree.get_children():
            self.employees_tree.delete(item)
            
        # Récupérer les employés de la base de données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        search_term = self.search_var.get() if hasattr(self, 'search_var') else ""
        
        if search_term:
            cursor.execute('''
                SELECT id, matricule, first_name, last_name, job_title, department, status, photo_path
                FROM employees 
                WHERE first_name LIKE ? OR last_name LIKE ? OR matricule LIKE ?
                ORDER BY last_name, first_name
            ''', (f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'))
        else:
            cursor.execute('''
                SELECT id, matricule, first_name, last_name, job_title, department, status, photo_path
                FROM employees 
                ORDER BY last_name, first_name
            ''')
            
        employees = cursor.fetchall()
        conn.close()
        
        # Ajouter les employés à la liste
        for emp in employees:
            emp_id, matricule, first_name, last_name, job_title, department, status, photo_path = emp
            full_name = f"{first_name} {last_name}"
            
            # Indicateur photo
            photo_indicator = "📷" if photo_path and os.path.exists(photo_path) else "👤"
            
            # Couleur selon le statut
            tags = []
            if status == "Active":
                tags = ['active']
            elif status == "En Congé":
                tags = ['on_leave']
            else:
                tags = ['inactive']
                
            self.employees_tree.insert('', 'end',
                                     values=(photo_indicator, matricule, full_name, 
                                           job_title or '', department or '', status),
                                     tags=tags)
        
        # Configuration des couleurs par tag
        self.employees_tree.tag_configure('active', background='#E8F5E8')
        self.employees_tree.tag_configure('on_leave', background='#FFF3E0')
        self.employees_tree.tag_configure('inactive', background='#FFEBEE')
        
    def filter_employees(self, *args):
        """Filtrer les employés selon la recherche"""
        self.load_employees()
        
    def add_new_employee(self):
        """Ajouter un nouvel employé"""
        self.current_employee_id = None
        self.show_employee_form()
        
    def edit_employee(self):
        """Modifier un employé sélectionné"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé à modifier")
            return
            
        # Récupérer l'ID de l'employé
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            self.current_employee_id = result[0]
            self.show_employee_form()
            
    def delete_employee(self):
        """Supprimer un employé"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé à supprimer")
            return
            
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        full_name = item['values'][2]
        
        if messagebox.askyesno("Confirmation", 
                              f"Êtes-vous sûr de vouloir supprimer l'employé {full_name} (Matricule: {matricule}) ?\n\nCette action est irréversible."):
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Récupérer l'ID
            cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
            result = cursor.fetchone()
            
            if result:
                emp_id = result[0]
                
                # Supprimer les données liées
                cursor.execute('DELETE FROM career_history WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM documents WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM leaves WHERE employee_id = ?', (emp_id,))
                cursor.execute('DELETE FROM employees WHERE id = ?', (emp_id,))
                
                conn.commit()
                messagebox.showinfo("Succès", "Employé supprimé avec succès")
                self.load_employees()
            
            conn.close()
            
    def open_employee_file(self, event=None):
        """Ouvrir le dossier complet d'un employé"""
        selection = self.employees_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un employé")
            return
            
        item = self.employees_tree.item(selection[0])
        matricule = item['values'][1]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM employees WHERE matricule = ?', (matricule,))
        result = cursor.fetchone()
        conn.close()
        
        if result:
            self.current_employee_id = result[0]
            self.show_employee_details()
            
    def show_employee_form(self):
        """Afficher le formulaire d'ajout/modification d'employé"""
        # Créer une nouvelle fenêtre
        form_window = tk.Toplevel(self.root)
        form_window.title("Nouvel Employé" if not self.current_employee_id else "Modifier Employé")
        form_window.geometry("800x700")
        form_window.configure(bg=self.colors['background'])
        form_window.transient(self.root)
        form_window.grab_set()
        
        # Variables pour les champs
        self.form_vars = {}
        self.form_text_widgets = {} # <-- LIGNE A AJOUTER
        # NOUVELLE VERSION
        fields = [
            'matricule', 'first_name', 'last_name', 'gender', 'birth_date', 'birth_place',
            'address', 'phone', 'email', 'marital_status', 'dependents', 'social_security',
            'bank_details', 'hire_date', 'contract_type', 'contract_start', 'contract_end',
            'department', 'job_title', 'status', 'cni', 'nationalite', 'numero_decision'
        ]
        
        for field in fields:
            self.form_vars[field] = tk.StringVar()
            
        # Si modification, charger les données existantes
        if self.current_employee_id:
            self.load_employee_data()
            
        # Titre
        title = tk.Label(form_window,
                        text="📝 " + ("Nouvel Employé" if not self.current_employee_id else "Modifier Employé"),
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(form_window, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet Informations Personnelles
        personal_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(personal_frame, text="👤 Informations Personnelles")
        
        # Onglet Informations Contractuelles
        contract_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(contract_frame, text="📋 Informations Contractuelles")
        
        # Remplir l'onglet personnel
        self.create_personal_info_tab(personal_frame)
        
        # Remplir l'onglet contractuel
        self.create_contract_info_tab(contract_frame)
        
        # Boutons d'action
        buttons_frame = tk.Frame(form_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        save_btn = tk.Button(buttons_frame,
                            text="💾 Enregistrer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=lambda: self.save_employee(form_window))
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=form_window.destroy)
        cancel_btn.pack(side='right')
        
    def create_personal_info_tab(self, parent):
        """Créer l'onglet des informations personnelles"""
        # Frame avec scrollbar
        canvas = tk.Canvas(parent, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Champs du formulaire
        fields_config = [
            ("Matricule *", 'matricule', 'entry'),
            ("Prénom *", 'first_name', 'entry'),
            ("Nom *", 'last_name', 'entry'),
            ("CNI (Carte Nationale d'Identité)", 'cni', 'entry'),      # <-- LIGNE A AJOUTER
            ("Nationalité", 'nationalite', 'entry'), 
            ("Genre", 'gender', 'combo', ['Masculin', 'Féminin']),
            ("Date de Naissance (jj/mm/aaaa)", 'birth_date', 'entry'),
            ("Lieu de Naissance", 'birth_place', 'entry'),
            ("Adresse", 'address', 'text'),
            ("Téléphone", 'phone', 'entry'),
            ("Email", 'email', 'entry'),
            ("Situation Matrimoniale", 'marital_status', 'combo', ['Célibataire', 'Marié(e)', 'Divorcé(e)', 'Veuf/Veuve']),
            ("Nombre de Personnes à Charge", 'dependents', 'entry'),
            ("Numéro de Sécurité Sociale", 'social_security', 'entry'),
            ("RIB/Détails Bancaires", 'bank_details', 'text')
        ]
        
        row = 0
        for field_config in fields_config:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(scrollable_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', padx=20, pady=(10, 5))
            
            # Champ de saisie
            if field_type == 'entry':
                widget = tk.Entry(scrollable_frame,
                                textvariable=self.form_vars[var_name],
                                font=('Segoe UI', 11),
                                width=40,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(scrollable_frame,
                                    textvariable=self.form_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=37,
                                    state='readonly')
            elif field_type == 'text':
                widget = tk.Text(scrollable_frame,
                               font=('Segoe UI', 11),
                               width=40,
                               height=3,
                               relief='solid',
                               bd=1)
                # On stocke une référence au widget pour le manipuler plus tard
                self.form_text_widgets[var_name] = widget
                    
            widget.grid(row=row+1, column=0, sticky='w', padx=20, pady=(0, 5))
            row += 2
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def create_contract_info_tab(self, parent):
        """Créer l'onglet des informations contractuelles"""
        # Frame avec scrollbar
        canvas = tk.Canvas(parent, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Champs contractuels
        # NOUVELLE VERSION
        fields_config = [
            ("Date d'Embauche (jj/mm/aaaa) *", 'hire_date', 'entry'),
            ("Type d'engagement", 'contract_type', 'combo', ["Décision d'engagements", 'CDD', 'Stage', 'Consultant']),
            ("Numéro décision", 'numero_decision', 'entry'),
            ("Début de Contrat (jj/mm/aaaa)", 'contract_start', 'entry'),
            ("Fin de Contrat (jj/mm/aaaa)", 'contract_end', 'entry'),
            ("Division", 'department', 'entry'),
            ("Corps de l'agent *", 'job_title', 'entry'),
            ("Statut", 'status', 'combo', ['Active', 'En Congé', 'Suspendu', 'Retraité', 'Démissionné'])
        ]
        
        row = 0
        for field_config in fields_config:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(scrollable_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', padx=20, pady=(10, 5))
            
            # Champ de saisie
            if field_type == 'entry':
                widget = tk.Entry(scrollable_frame,
                                textvariable=self.form_vars[var_name],
                                font=('Segoe UI', 11),
                                width=40,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(scrollable_frame,
                                    textvariable=self.form_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=37,
                                    state='readonly')
                    
            widget.grid(row=row+1, column=0, sticky='w', padx=20, pady=(0, 5))
            row += 2
            
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
    def load_employee_data(self):
        """Charger les données d'un employé existant"""
        if not self.current_employee_id:
            return
            
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row # Permet d'accéder aux colonnes par leur nom
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM employees WHERE id = ?', (self.current_employee_id,))
        employee_data = cursor.fetchone()
        conn.close()
        
        if employee_data:
            # 1. Remplir les champs standards (Entry, Combobox)
            for field_name, var in self.form_vars.items():
                if field_name in employee_data.keys() and field_name not in self.form_text_widgets:
                    value = employee_data[field_name] or ''
                    var.set(value)

            # 2. Remplir manuellement les champs Text (Adresse et RIB)
            for field_name, text_widget in self.form_text_widgets.items():
                if field_name in employee_data.keys():
                    value = employee_data[field_name] or ''
                    text_widget.delete('1.0', tk.END)  # Vider le champ
                    text_widget.insert('1.0', value)   # Insérer la nouvelle valeur
                        
    def save_employee(self, form_window):
        """Enregistrer les données de l'employé"""
        # Validation des champs obligatoires
        # NOUVELLE VERSION
        required_fields = {
            'matricule': 'Matricule',
            'first_name': 'Prénom',
            'last_name': 'Nom',
            'hire_date': 'Date d\'embauche',
            'job_title': "Corps de l'agent"
        }
        
        for field, label in required_fields.items():
            if not self.form_vars[field].get().strip():
                messagebox.showerror("Erreur", f"Le champ '{label}' est obligatoire")
                return
                
        # Validation du format des dates
        date_fields = ['birth_date', 'hire_date', 'contract_start', 'contract_end']
        for field in date_fields:
            date_value = self.form_vars[field].get().strip()
            if date_value and not self.validate_date_format(date_value):
                messagebox.showerror("Erreur", f"Format de date invalide pour '{field}'. Utilisez jj/mm/aaaa")
                return
                
        # Vérifier l'unicité du matricule
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        if self.current_employee_id:
            cursor.execute('SELECT id FROM employees WHERE matricule = ? AND id != ?',
                          (self.form_vars['matricule'].get(), self.current_employee_id))
        else:
            cursor.execute('SELECT id FROM employees WHERE matricule = ?',
                          (self.form_vars['matricule'].get(),))
                          
        if cursor.fetchone():
            messagebox.showerror("Erreur", "Ce matricule existe déjà")
            conn.close()
            return
            
        # Préparer les données
        data = []
        # NOUVELLE VERSION
        fields = [
            'matricule', 'first_name', 'last_name', 'gender', 'birth_date', 'birth_place',
            'address', 'phone', 'email', 'marital_status', 'dependents', 'social_security',
            'bank_details', 'hire_date', 'contract_type', 'contract_start', 'contract_end',
            'department', 'job_title', 'status', 'cni', 'nationalite', 'numero_decision'
        ]
        
        for field in fields:
            # Si le champ est un widget Text
            if field in self.form_text_widgets:
                value = self.form_text_widgets[field].get('1.0', tk.END).strip()
                data.append(value if value else None)
            # Sinon (c'est un Entry ou Combobox)
            else:
                value = self.form_vars[field].get().strip()
                data.append(value if value else None)
            
        try:
            # NOUVELLE VERSION
            if self.current_employee_id:
                # Mise à jour
                sql = '''UPDATE employees SET 
                        matricule=?, first_name=?, last_name=?, gender=?, birth_date=?, birth_place=?,
                        address=?, phone=?, email=?, marital_status=?, dependents=?, social_security=?,
                        bank_details=?, hire_date=?, contract_type=?, contract_start=?, contract_end=?,
                        department=?, job_title=?, status=?, cni=?, nationalite=?, numero_decision=?, updated_at=CURRENT_TIMESTAMP
                        WHERE id=?'''
                data.append(self.current_employee_id)
                cursor.execute(sql, data)
                message = "Employé modifié avec succès"
            else:
                # Insertion
                sql = '''INSERT INTO employees 
                        (matricule, first_name, last_name, gender, birth_date, birth_place,
                        address, phone, email, marital_status, dependents, social_security,
                        bank_details, hire_date, contract_type, contract_start, contract_end,
                        department, job_title, status, cni, nationalite, numero_decision) 
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)'''
                cursor.execute(sql, data)
                message = "Employé ajouté avec succès"
                
            conn.commit()
            messagebox.showinfo("Succès", message)
            form_window.destroy()
            self.load_employees()  # Recharger la liste
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def validate_date_format(self, date_string):
        """Valider le format de date jj/mm/aaaa"""
        try:
            datetime.strptime(date_string, '%d/%m/%Y')
            return True
        except ValueError:
            return False
            
    def show_employee_details(self):
        """Afficher les détails complets d'un employé"""
        if not self.current_employee_id:
            return
            
        # Créer une nouvelle fenêtre
        details_window = tk.Toplevel(self.root)
        details_window.title("Dossier Employé")
        details_window.geometry("1000x800")
        details_window.configure(bg=self.colors['background'])
        details_window.transient(self.root)
        
        # Récupérer les données de l'employé en utilisant les noms de colonnes
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row # Important: permet d'accéder aux colonnes par leur nom
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM employees WHERE id = ?', (self.current_employee_id,))
        employee = cursor.fetchone()
        conn.close()
        
        if not employee:
            messagebox.showerror("Erreur", "Employé non trouvé")
            details_window.destroy()
            return
            
        # Titre avec nom de l'employé
        title = tk.Label(details_window,
                        text=f"📁 Dossier de {employee['first_name']} {employee['last_name']}",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(details_window, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglets
        self.create_employee_info_tab(notebook, employee)
        self.create_career_history_tab(notebook)
        self.create_documents_tab(notebook)
        self.create_leaves_history_tab(notebook)

    def create_employee_info_tab(self, notebook, employee):
        """Créer l'onglet des informations de l'employé"""
        info_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(info_frame, text="👤 Informations")
        
        # Frame principal avec scrollbar
        canvas = tk.Canvas(info_frame, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(info_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Section Photo
        photo_frame = tk.LabelFrame(scrollable_frame,
                                   text="📷 Photo",
                                   font=('Segoe UI', 12, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'])
        photo_frame.pack(fill='x', padx=20, pady=10)
        
        photo_label = tk.Label(photo_frame,
                              bg=self.colors['light_gray'],
                              relief='solid',
                              bd=1)
        photo_label.pack(side='left', padx=10, pady=10)
        
        self.display_photo(photo_label, employee['photo_path'])
        
        photo_buttons = tk.Frame(photo_frame, bg=self.colors['background'])
        photo_buttons.pack(side='left', padx=20, pady=10)
        
        upload_photo_btn = tk.Button(photo_buttons,
                                    text="📁 Charger Photo",
                                    font=('Segoe UI', 10),
                                    bg=self.colors['accent_green'],
                                    fg=self.colors['text_dark'],
                                    relief='flat',
                                    bd=0,
                                    padx=15,
                                    pady=5,
                                    cursor='hand2',
                                    command=lambda: self.upload_employee_photo(photo_label))
        upload_photo_btn.pack(pady=2)
        
        # Informations personnelles
        personal_frame = tk.LabelFrame(scrollable_frame,
                                      text="👤 Informations Personnelles",
                                      font=('Segoe UI', 12, 'bold'),
                                      fg=self.colors['primary_green'],
                                      bg=self.colors['background'])
        personal_frame.pack(fill='x', padx=20, pady=10)
        
        personal_fields = [
            ("Matricule:", employee['matricule']),
            ("Nom Complet:", f"{employee['first_name']} {employee['last_name']}"),
            ("CNI:", employee['cni']),
            ("Nationalité:", employee['nationalite']),
            ("Genre:", employee['gender']),
            ("Date de Naissance:", employee['birth_date']),
            ("Lieu de Naissance:", employee['birth_place']),
            ("Adresse:", employee['address']),
            ("Téléphone:", employee['phone']),
            ("Email:", employee['email']),
            ("Situation Matrimoniale:", employee['marital_status']),
            ("Personnes à Charge:", employee['dependents']),
            ("Sécurité Sociale:", employee['social_security'])
        ]
        
        for i, (label, value) in enumerate(personal_fields):
            row_frame = tk.Frame(personal_frame, bg=self.colors['background'])
            row_frame.pack(fill='x', padx=10, pady=2)
            
            tk.Label(row_frame,
                    text=label,
                    font=('Segoe UI', 10, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background'],
                    width=20,
                    anchor='w').pack(side='left')
                    
            tk.Label(row_frame,
                    text=value or 'Non renseigné',
                    font=('Segoe UI', 10),
                    fg=self.colors['text_dark'] if value else self.colors['text_light'],
                    bg=self.colors['background'],
                    anchor='w').pack(side='left', padx=(10, 0))
        
        # Informations contractuelles
        contract_frame = tk.LabelFrame(scrollable_frame,
                                      text="📋 Informations Contractuelles",
                                      font=('Segoe UI', 12, 'bold'),
                                      fg=self.colors['primary_green'],
                                      bg=self.colors['background'])
        contract_frame.pack(fill='x', padx=20, pady=10)
        
        contract_fields = [
            ("Date d'Embauche:", employee['hire_date']),
            ("Type d'engagement:", employee['contract_type']),
            ("Numéro décision:", employee['numero_decision']), # <-- CORRIGÉ
            ("Début de Contrat:", employee['contract_start']),
            ("Fin de Contrat:", employee['contract_end']),
            ("Division:", employee['department']),
            ("Corps de l'agent:", employee['job_title']),
            ("Statut:", employee['status'])
        ]
        
        for label, value in contract_fields:
            row_frame = tk.Frame(contract_frame, bg=self.colors['background'])
            row_frame.pack(fill='x', padx=10, pady=2)
            
            tk.Label(row_frame,
                    text=label,
                    font=('Segoe UI', 10, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background'],
                    width=20,
                    anchor='w').pack(side='left')
                    
            tk.Label(row_frame,
                    text=value or 'Non renseigné',
                    font=('Segoe UI', 10),
                    fg=self.colors['text_dark'] if value else self.colors['text_light'],
                    bg=self.colors['background'],
                    anchor='w').pack(side='left', padx=(10, 0))
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def upload_employee_photo(self, photo_label):
        """Charger une photo pour l'employé"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner une photo",
            filetypes=[("Images", "*.jpg *.jpeg *.png *.gif *.bmp")]
        )
        
        if file_path:
            try:
                # Copier le fichier dans le dossier photos
                filename = f"emp_{self.current_employee_id}_{os.path.basename(file_path)}"
                dest_path = os.path.join(self.photos_folder, filename)
                shutil.copy2(file_path, dest_path)
                
                # Mettre à jour la base de données
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute('UPDATE employees SET photo_path = ? WHERE id = ?',
                              (dest_path, self.current_employee_id))
                conn.commit()
                conn.close()
                
                # Afficher la photo
                self.display_photo(photo_label, dest_path)
                
                messagebox.showinfo("Succès", "Photo chargée avec succès")
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors du chargement de la photo: {str(e)}")
                
    def display_photo(self, label, photo_path):
        """Afficher une photo dans un label"""
        try:
            if photo_path and os.path.exists(photo_path):
                # Charger et redimensionner l'image
                image = Image.open(photo_path)
                # MODIFIEZ LA TAILLE ICI (largeur, hauteur)
                image = image.resize((200, 210), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(image)
                
                label.configure(image=photo, text="")
                label.image = photo  # Garder une référence pour éviter qu'elle disparaisse
            else:
                # Si aucune photo n'est trouvée, on vide le label
                label.configure(image="", text="Aucune photo")
        except Exception as e:
            label.configure(image="", text="Erreur photo")
            print(f"Erreur lors de l'affichage de la photo : {e}")
            
    def create_career_history_tab(self, notebook):
        """Créer l'onglet historique de carrière"""
        history_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(history_frame, text="📈 Historique Carrière")
        
        # Barre d'outils
        toolbar = tk.Frame(history_frame, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=10)
        
        add_act_btn = tk.Button(toolbar,
                               text="➕ Nouvel Acte",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.add_career_act)
        add_act_btn.pack(side='left')
        
        # Liste des actes
        columns = ('N° Acte', 'Nature', 'Objet', 'Date Acte', 'Date Effet', 'Document')
        self.career_tree = ttk.Treeview(history_frame,
                                       columns=columns,
                                       show='headings',
                                       style='Custom.Treeview',
                                       height=12)
        
        # Configuration des colonnes
        for col in columns:
            self.career_tree.heading(col, text=col)
            self.career_tree.column(col, width=120, anchor='center')
            
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(history_frame, orient='vertical', command=self.career_tree.yview)
        self.career_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Placement
        self.career_tree.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=(0, 20))
        v_scrollbar.pack(side='right', fill='y', padx=(0, 20), pady=(0, 20))
        
        # Charger l'historique
        self.load_career_history()
        
    def add_career_act(self):
        """Ajouter un acte administratif"""
        # Fenêtre de saisie
        act_window = tk.Toplevel(self.root)
        act_window.title("Nouvel Acte Administratif")
        act_window.geometry("600x500")
        act_window.configure(bg=self.colors['background'])
        act_window.transient(self.root)
        act_window.grab_set()
        
        # Variables
        act_vars = {
            'act_number': tk.StringVar(),
            'nature': tk.StringVar(),
            'subject': tk.StringVar(),
            'act_date': tk.StringVar(),
            'effective_date': tk.StringVar()
        }
        
        # Titre
        title = tk.Label(act_window,
                        text="📝 Nouvel Acte Administratif",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(act_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20)
        
        fields = [
            ("Numéro d'Acte:", 'act_number', 'entry'),
            ("Nature:", 'nature', 'combo', ['Nomination', 'Promotion', 'Mutation', 'Sanction', 'Formation', 'Autre']),
            ("Objet:", 'subject', 'text'),
            ("Date de l'Acte (jj/mm/aaaa):", 'act_date', 'entry'),
            ("Date d'Effet (jj/mm/aaaa):", 'effective_date', 'entry')
        ]
        
        widgets = {}
        row = 0
        for field_config in fields:
            label_text = field_config[0]
            var_name = field_config[1]
            field_type = field_config[2]
            
            # Label
            label = tk.Label(form_frame,
                           text=label_text,
                           font=('Segoe UI', 11),
                           fg=self.colors['text_dark'],
                           bg=self.colors['background'],
                           anchor='w')
            label.grid(row=row, column=0, sticky='w', pady=(10, 5))
            
            # Widget
            if field_type == 'entry':
                widget = tk.Entry(form_frame,
                                textvariable=act_vars[var_name],
                                font=('Segoe UI', 11),
                                width=50,
                                relief='solid',
                                bd=1)
            elif field_type == 'combo':
                widget = ttk.Combobox(form_frame,
                                    textvariable=act_vars[var_name],
                                    values=field_config[3],
                                    font=('Segoe UI', 11),
                                    width=47,
                                    state='readonly')
            elif field_type == 'text':
                widget = tk.Text(form_frame,
                               font=('Segoe UI', 11),
                               width=50,
                               height=4,
                               relief='solid',
                               bd=1)
                widgets[var_name] = widget  # Stocker pour récupérer le contenu plus tard
                
            widget.grid(row=row+1, column=0, sticky='w', pady=(0, 5))
            row += 2
            
        # Bouton pour attacher un document
        doc_frame = tk.Frame(form_frame, bg=self.colors['background'])
        doc_frame.grid(row=row, column=0, sticky='w', pady=10)
        
        self.selected_doc_path = None
        
        attach_btn = tk.Button(doc_frame,
                              text="📎 Attacher Document",
                              font=('Segoe UI', 10),
                              bg=self.colors['accent_green'],
                              fg=self.colors['text_dark'],
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2',
                              command=self.select_document)
        attach_btn.pack(side='left')
        
        self.doc_label = tk.Label(doc_frame,
                                 text="Aucun document sélectionné",
                                 font=('Segoe UI', 10),
                                 fg=self.colors['text_light'],
                                 bg=self.colors['background'])
        self.doc_label.pack(side='left', padx=(10, 0))
        
        # Boutons d'action
        buttons_frame = tk.Frame(act_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=20)
        
        save_btn = tk.Button(buttons_frame,
                            text="💾 Enregistrer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=lambda: self.save_career_act(act_vars, widgets, act_window))
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=act_window.destroy)
        cancel_btn.pack(side='right')
        
    def select_document(self):
        """Sélectionner un document à attacher"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner un document",
            filetypes=[("Tous les fichiers", "*.*"), ("PDF", "*.pdf"), ("Images", "*.jpg *.jpeg *.png"), ("Word", "*.docx *.doc")]
        )
        
        if file_path:
            self.selected_doc_path = file_path
            filename = os.path.basename(file_path)
            self.doc_label.configure(text=f"📄 {filename}", fg=self.colors['primary_green'])
            
    def save_career_act(self, act_vars, widgets, window):
        """Enregistrer un acte administratif"""
        # Validation
        required_fields = ['act_number', 'nature', 'act_date']
        for field in required_fields:
            if not act_vars[field].get().strip():
                messagebox.showerror("Erreur", f"Le champ '{field}' est obligatoire")
                return
                
        # Validation des dates
        for date_field in ['act_date', 'effective_date']:
            date_value = act_vars[date_field].get().strip()
            if date_value and not self.validate_date_format(date_value):
                messagebox.showerror("Erreur", f"Format de date invalide pour '{date_field}'. Utilisez jj/mm/aaaa")
                return
                
        # Récupérer le contenu du champ texte
        subject = widgets['subject'].get('1.0', tk.END).strip() if 'subject' in widgets else act_vars['subject'].get()
        
        # Copier le document si sélectionné
        doc_path = None
        if self.selected_doc_path:
            try:
                filename = f"act_{self.current_employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(self.selected_doc_path)}"
                doc_path = os.path.join(self.documents_folder, filename)
                shutil.copy2(self.selected_doc_path, doc_path)
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la copie du document: {str(e)}")
                return
                
        # Enregistrer en base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                INSERT INTO career_history 
                (employee_id, act_number, nature, subject, act_date, effective_date, document_path)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                self.current_employee_id,
                act_vars['act_number'].get(),
                act_vars['nature'].get(),
                subject,
                act_vars['act_date'].get(),
                act_vars['effective_date'].get() or None,
                doc_path
            ))
            
            conn.commit()
            messagebox.showinfo("Succès", "Acte administratif enregistré avec succès")
            window.destroy()
            self.load_career_history()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
        finally:
            conn.close()
            
    def load_career_history(self):
        """Charger l'historique de carrière"""
        if not hasattr(self, 'career_tree'):
            return
            
        # Vider la liste
        for item in self.career_tree.get_children():
            self.career_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT act_number, nature, subject, act_date, effective_date, document_path
            FROM career_history 
            WHERE employee_id = ?
            ORDER BY act_date DESC
        ''', (self.current_employee_id,))
        
        acts = cursor.fetchall()
        conn.close()
        
        for act in acts:
            act_number, nature, subject, act_date, effective_date, document_path = act
            doc_indicator = "📄" if document_path and os.path.exists(document_path) else ""
            
            self.career_tree.insert('', 'end', values=(
                act_number,
                nature,
                subject[:30] + "..." if len(subject) > 30 else subject,
                act_date,
                effective_date or "",
                doc_indicator
            ))
            
    def create_documents_tab(self, notebook):
        """Créer l'onglet gestion des documents avec filtres, modification et suppression."""
        docs_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(docs_frame, text="📄 Documents")

        # --- Barre d'outils avec filtres et bouton d'ajout ---
        toolbar = tk.Frame(docs_frame, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=10)

        category_frame = tk.Frame(toolbar, bg=self.colors['background'])
        category_frame.pack(side='left')

        categories = ["Tous", "Décision", "Correspondance", "Note de service", "État Civil", "Autres"]
        
        for category in categories:
            btn = tk.Button(category_frame,
                           text=category,
                           font=('Segoe UI', 10),
                           bg=self.colors['light_gray'],
                           fg=self.colors['text_dark'],
                           relief='flat',
                           bd=0,
                           padx=12,
                           pady=6,
                           cursor='hand2',
                           command=lambda cat=category: self.load_documents(category=cat))
            btn.pack(side='left', padx=3)

        add_doc_btn = tk.Button(toolbar,
                               text="📁 Ajouter Document",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.add_document)
        add_doc_btn.pack(side='right')

        # --- Liste des documents (Treeview) ---
        tree_container = tk.Frame(docs_frame)
        tree_container.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        columns = ('category', 'name', 'uploaded_at')
        self.docs_tree = ttk.Treeview(tree_container,
                                     columns=columns,
                                     show='headings',
                                     style='Custom.Treeview',
                                     height=12)

        self.docs_tree.heading('category', text='Catégorie')
        self.docs_tree.heading('name', text='Nom du Document')
        self.docs_tree.heading('uploaded_at', text="Date d'Ajout")

        self.docs_tree.column('category', width=150, anchor='w')
        self.docs_tree.column('name', width=300, anchor='w')
        self.docs_tree.column('uploaded_at', width=120, anchor='center')

        v_scrollbar = ttk.Scrollbar(tree_container, orient='vertical', command=self.docs_tree.yview)
        self.docs_tree.configure(yscrollcommand=v_scrollbar.set)

        self.docs_tree.pack(side='left', fill='both', expand=True)
        v_scrollbar.pack(side='right', fill='y')

        # --- NOUVEAU : Menu contextuel pour Modifier/Supprimer ---
        self.doc_context_menu = tk.Menu(self.root, tearoff=0)
        self.doc_context_menu.add_command(label="👁️ Ouvrir le fichier", command=self.open_document)
        self.doc_context_menu.add_command(label="✏️ Modifier les informations", command=self.edit_document)
        self.doc_context_menu.add_separator()
        self.doc_context_menu.add_command(label="🗑️ Supprimer le document", command=self.delete_document)

        def show_doc_context_menu(event):
            # Sélectionner l'item sous le curseur avant d'afficher le menu
            item_id = self.docs_tree.identify_row(event.y)
            if item_id:
                self.docs_tree.selection_set(item_id)
                self.doc_context_menu.tk_popup(event.x_root, event.y_root)

        self.docs_tree.bind('<Double-1>', self.open_document)
        self.docs_tree.bind("<Button-3>", show_doc_context_menu) # Clic droit

        self.load_documents()

    def add_document(self):
        """Ajouter un document pour l'employé actuel."""
        file_path = filedialog.askopenfilename(
            title="Sélectionner un document à joindre",
            filetypes=[("Tous les fichiers", "*.*"), ("PDF", "*.pdf"), ("Images", "*.jpg *.png"), ("Word", "*.docx")]
        )
        if not file_path:
            return

        categories = ["Décision", "Correspondance", "Note de service", "État Civil", "Diplômes", "Autres"]
        
        # Fenêtre de dialogue pour nom et catégorie
        dialog = tk.Toplevel(self.root)
        dialog.title("Informations du document")
        dialog.geometry("400x250")
        dialog.configure(bg=self.colors['background'])
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text="Nom du document:", bg=self.colors['background'], font=('Segoe UI', 11)).pack(pady=(10,0))
        name_var = tk.StringVar(value=os.path.basename(file_path))
        tk.Entry(dialog, textvariable=name_var, width=50).pack(pady=5, padx=10)

        tk.Label(dialog, text="Catégorie:", bg=self.colors['background'], font=('Segoe UI', 11)).pack(pady=(10,0))
        category_var = tk.StringVar()
        ttk.Combobox(dialog, textvariable=category_var, values=categories, state='readonly', width=47).pack(pady=5, padx=10)

        def on_submit():
            doc_name = name_var.get().strip()
            category = category_var.get()
            if not doc_name or not category:
                messagebox.showerror("Erreur", "Le nom et la catégorie sont obligatoires.", parent=dialog)
                return
            
            dialog.destroy()
            self._save_document_to_db(doc_name, category, file_path)

        submit_btn = tk.Button(dialog, text="Enregistrer", command=on_submit, bg=self.colors['primary_green'], fg='white', relief='flat')
        submit_btn.pack(pady=20)

    def _save_document_to_db(self, doc_name, category, file_path, doc_id=None):
        """Logique interne pour sauvegarder le document dans la base de données."""
        try:
            # Copier le fichier physique
            filename = f"doc_{self.current_employee_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{os.path.basename(file_path)}"
            dest_path = os.path.join(self.documents_folder, filename)
            shutil.copy2(file_path, dest_path)

            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            if doc_id: # Mode mise à jour
                # On ne met à jour que les métadonnées, pas le fichier lui-même pour l'instant
                cursor.execute('''
                    UPDATE documents SET name = ?, category = ? WHERE id = ?
                ''', (doc_name, category, doc_id))
                message = "Informations du document modifiées avec succès."
            else: # Mode création
                cursor.execute('''
                    INSERT INTO documents (employee_id, category, name, file_path)
                    VALUES (?, ?, ?, ?)
                ''', (self.current_employee_id, category, doc_name, dest_path))
                message = "Document ajouté avec succès."

            conn.commit()
            conn.close()
            messagebox.showinfo("Succès", message)
            self.load_documents()

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement du document: {e}")

    def load_documents(self, category="Tous"):
        """Charger la liste des documents en filtrant par catégorie."""
        if not hasattr(self, 'docs_tree'):
            return

        for item in self.docs_tree.get_children():
            self.docs_tree.delete(item)

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        query = 'SELECT id, category, name, uploaded_at, file_path FROM documents WHERE employee_id = ?'
        params = (self.current_employee_id,)

        if category != "Tous":
            query += ' AND category = ?'
            params = (self.current_employee_id, category)
        
        query += ' ORDER BY uploaded_at DESC'

        cursor.execute(query, params)
        documents = cursor.fetchall()
        conn.close()

        for doc in documents:
            doc_id, doc_category, name, uploaded_at, file_path = doc
            try:
                date_obj = datetime.strptime(uploaded_at, '%Y-%m-%d %H:%M:%S')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except (ValueError, TypeError):
                formatted_date = uploaded_at

            # On utilise l'ID du document comme iid
            self.docs_tree.insert('', 'end', iid=doc_id, values=(
                doc_category,
                name,
                formatted_date
            ))

    def open_document(self, event=None):
        """Ouvrir un document sélectionné."""
        selection = self.docs_tree.selection()
        if not selection:
            return
            
        doc_id = int(selection[0])
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT file_path FROM documents WHERE id = ?', (doc_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result and os.path.exists(result[0]):
            try:
                if platform.system() == 'Windows':
                    os.startfile(result[0])
                else:
                    subprocess.call(('open' if platform.system() == 'Darwin' else 'xdg-open', result[0]))
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'ouvrir le document: {e}")
        else:
            messagebox.showerror("Erreur", "Fichier du document non trouvé sur le disque.")

    def edit_document(self):
        """Modifier les informations d'un document sélectionné."""
        selection = self.docs_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un document à modifier.")
            return

        doc_id = int(selection[0])

        # Récupérer les informations actuelles
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT name, category FROM documents WHERE id = ?', (doc_id,))
        current_data = cursor.fetchone()
        conn.close()

        if not current_data:
            messagebox.showerror("Erreur", "Document non trouvé dans la base de données.")
            return
        
        current_name, current_category = current_data
        categories = ["Décision", "Correspondance", "Note de service", "État Civil", "Diplômes", "Autres"]

        # Fenêtre de dialogue pour la modification
        dialog = tk.Toplevel(self.root)
        dialog.title("Modifier les informations")
        dialog.geometry("400x250")
        dialog.configure(bg=self.colors['background'])
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text="Nom du document:", bg=self.colors['background'], font=('Segoe UI', 11)).pack(pady=(10,0))
        name_var = tk.StringVar(value=current_name)
        tk.Entry(dialog, textvariable=name_var, width=50).pack(pady=5, padx=10)

        tk.Label(dialog, text="Catégorie:", bg=self.colors['background'], font=('Segoe UI', 11)).pack(pady=(10,0))
        category_var = tk.StringVar(value=current_category)
        combo = ttk.Combobox(dialog, textvariable=category_var, values=categories, state='readonly', width=47)
        combo.pack(pady=5, padx=10)

        def on_submit():
            new_name = name_var.get().strip()
            new_category = category_var.get()
            if not new_name or not new_category:
                messagebox.showerror("Erreur", "Le nom et la catégorie sont obligatoires.", parent=dialog)
                return
            
            dialog.destroy()
            
            # Mettre à jour la base de données
            try:
                conn_update = sqlite3.connect(self.db_path)
                cursor_update = conn_update.cursor()
                cursor_update.execute('UPDATE documents SET name = ?, category = ? WHERE id = ?', (new_name, new_category, doc_id))
                conn_update.commit()
                conn_update.close()
                messagebox.showinfo("Succès", "Document mis à jour.")
                self.load_documents()
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur de base de données : {e}")

        submit_btn = tk.Button(dialog, text="Enregistrer", command=on_submit, bg=self.colors['primary_green'], fg='white', relief='flat')
        submit_btn.pack(pady=20)

    def delete_document(self):
        """Supprimer un document sélectionné (enregistrement et fichier)."""
        selection = self.docs_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un document à supprimer.")
            return
        
        doc_id = int(selection[0])
        
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer ce document ?\nLe fichier associé sera également effacé définitivement."):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # 1. Récupérer le chemin du fichier avant de supprimer l'enregistrement
                cursor.execute('SELECT file_path FROM documents WHERE id = ?', (doc_id,))
                result = cursor.fetchone()
                file_path_to_delete = result[0] if result else None
                
                # 2. Supprimer l'enregistrement de la base de données
                cursor.execute('DELETE FROM documents WHERE id = ?', (doc_id,))
                conn.commit()
                conn.close()
                
                # 3. Supprimer le fichier physique s'il existe
                if file_path_to_delete and os.path.exists(file_path_to_delete):
                    os.remove(file_path_to_delete)
                
                messagebox.showinfo("Succès", "Document supprimé avec succès.")
                self.load_documents() # Rafraîchir la liste
                
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur de base de données : {e}")
            except OSError as e:
                messagebox.showerror("Erreur", f"Erreur lors de la suppression du fichier : {e}")


    def create_leaves_history_tab(self, notebook):
        """Créer l'onglet historique des congés"""
        leaves_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(leaves_frame, text="🏖️ Historique Congés")
        
        # Informations sur les soldes de congés
        balance_frame = tk.LabelFrame(leaves_frame,
                                     text="📊 Soldes de Congés",
                                     font=('Segoe UI', 12, 'bold'),
                                     fg=self.colors['primary_green'],
                                     bg=self.colors['background'])
        balance_frame.pack(fill='x', padx=20, pady=10)
        
        # Calculer les soldes (simplifié)
        current_year = datetime.now().year
        annual_leave_balance = 30  # Exemple: 30 jours par an
        
        balance_label = tk.Label(balance_frame,
                                text=f"Congés Annuels {current_year}: {annual_leave_balance} jours disponibles",
                                font=('Segoe UI', 11),
                                fg=self.colors['text_dark'],
                                bg=self.colors['background'])
        balance_label.pack(padx=10, pady=10)
        
        # Liste des congés
        columns = ('Type', 'Début', 'Fin', 'Durée', 'Statut', 'Notes')
        self.leaves_tree = ttk.Treeview(leaves_frame,
                                       columns=columns,
                                       show='headings',
                                       style='Custom.Treeview',
                                       height=10)
        
        # Configuration des colonnes
        for col in columns:
            self.leaves_tree.heading(col, text=col)
            self.leaves_tree.column(col, width=120, anchor='center')
            
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(leaves_frame, orient='vertical', command=self.leaves_tree.yview)
        self.leaves_tree.configure(yscrollcommand=v_scrollbar.set)
        
        # Placement
        self.leaves_tree.pack(side='left', fill='both', expand=True, padx=(20, 0), pady=(0, 20))
        v_scrollbar.pack(side='right', fill='y', padx=(0, 20), pady=(0, 20))
        
        # Charger l'historique des congés
        self.load_leaves_history()
        
    def load_leaves_history(self):
        """Charger l'historique des congés"""
        if not hasattr(self, 'leaves_tree'):
            return
            
        # Vider la liste
        for item in self.leaves_tree.get_children():
            self.leaves_tree.delete(item)
            
        # Charger depuis la base
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT lt.name, l.start_date, l.end_date, l.days_count, l.status, l.notes
            FROM leaves l
            JOIN leave_types lt ON l.leave_type_id = lt.id
            WHERE l.employee_id = ?
            ORDER BY l.start_date DESC
        ''', (self.current_employee_id,))
        
        leaves = cursor.fetchall()
        conn.close()
        
        for leave in leaves:
            leave_type, start_date, end_date, days_count, status, notes = leave
            
            self.leaves_tree.insert('', 'end', values=(
                leave_type,
                start_date,
                end_date,
                f"{days_count} jour(s)",
                status,
                notes or ""
            ))
            
    def show_leaves_module(self):
        """Module de gestion des congés"""
        self.clear_main_content()
        self.set_active_nav_button("🏖️ Gestion Congés")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="🏖️ Gestion des Congés et Absences",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Notebook pour les sous-modules
        notebook = ttk.Notebook(self.main_content, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet planification des congés
        planning_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(planning_frame, text="📅 Planification")
        
        # Onglet calendrier
        calendar_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(calendar_frame, text="📆 Calendrier")
        
        # Onglet configuration
        config_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(config_frame, text="⚙️ Configuration")
        
        # Remplir les onglets
        self.create_leave_planning_tab(planning_frame)
        self.create_leave_calendar_tab(calendar_frame)
        self.create_leave_config_tab(config_frame)

    def create_leave_planning_tab(self, parent):
        """Créer l'onglet de planification des congés avec défilement vertical et recherche."""
        # Frame principal pour cet onglet
        main_frame = tk.Frame(parent, bg=self.colors['background'])
        main_frame.pack(fill='both', expand=True)

        # Barre d'outils en haut
        toolbar = tk.Frame(main_frame, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=10)

        # Sélecteur d'année
        tk.Label(toolbar, text="Année:", font=('Segoe UI', 11, 'bold'), bg=self.colors['background']).pack(side='left', padx=(0, 5))
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        year_combo = ttk.Combobox(toolbar, textvariable=self.year_var, values=[str(y) for y in range(datetime.now().year - 5, datetime.now().year + 5)], width=8)
        year_combo.pack(side='left', padx=5)
        year_combo.bind('<<ComboboxSelected>>', lambda e: self.display_yearly_leave_plan())

        # Boutons d'action
        add_btn = tk.Button(toolbar, text="➕ Planifier Congé", font=('Segoe UI', 11, 'bold'), bg=self.colors['primary_green'], fg='white', relief='flat', command=self.open_leave_planning_form)
        add_btn.pack(side='left', padx=10)
        modify_btn = tk.Button(toolbar, text="✏️ Modifier", font=('Segoe UI', 11), bg=self.colors['accent_green'], fg=self.colors['text_dark'], relief='flat', command=self.modify_planned_leave)
        modify_btn.pack(side='left', padx=5)
        delete_btn = tk.Button(toolbar, text="🗑️ Supprimer", font=('Segoe UI', 11), bg=self.colors['error'], fg='white', relief='flat', command=self.delete_planned_leave)
        delete_btn.pack(side='left', padx=5)

        # Champ de recherche
        search_frame = tk.Frame(toolbar, bg=self.colors['background'])
        search_frame.pack(side='right', padx=10)
        
        tk.Label(search_frame, text="Rechercher Employé:", font=('Segoe UI', 11), bg=self.colors['background']).pack(side='left')
        self.leave_search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=self.leave_search_var, font=('Segoe UI', 11), width=25)
        search_entry.pack(side='left', padx=5)
        search_btn = tk.Button(search_frame, text="🔍", font=('Segoe UI', 11, 'bold'), bg=self.colors['primary_green'], fg='white', relief='flat', command=self.display_yearly_leave_plan)
        search_btn.pack(side='left')
        search_entry.bind('<Return>', lambda e: self.display_yearly_leave_plan())

        # --- MODIFIÉ : Canvas pour le défilement VERTICAL ---
        container_frame = tk.Frame(main_frame, bg=self.colors['background'])
        container_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.leave_planning_canvas = tk.Canvas(container_frame, bg=self.colors['background'], highlightthickness=0)
        v_scrollbar = ttk.Scrollbar(container_frame, orient="vertical", command=self.leave_planning_canvas.yview)
        self.leave_planning_canvas.configure(yscrollcommand=v_scrollbar.set)

        v_scrollbar.pack(side="right", fill="y")
        self.leave_planning_canvas.pack(side="left", fill="both", expand=True)

        # Frame qui contiendra tous les mois et qui défilera
        self.months_container = tk.Frame(self.leave_planning_canvas, bg=self.colors['background'])
        self.leave_planning_canvas.create_window((0, 0), window=self.months_container, anchor="nw")

        self.months_container.bind(
            "<Configure>",
            lambda e: self.leave_planning_canvas.configure(scrollregion=self.leave_planning_canvas.bbox("all"))
        )

        # Création des 12 mois en grille
        month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        self.monthly_leave_trees = []
        self.month_frames = [] # Pour stocker les frames des mois pour le défilement
        self.active_leave_tree = None
        columns = ('Employé', 'Poste', 'Division', 'Début', 'Fin')

        # On affiche 2 mois par ligne
        months_per_row = 2
        for i, month_name in enumerate(month_names):
            row, col = divmod(i, months_per_row)
            
            month_frame = tk.LabelFrame(self.months_container, text=month_name, font=('Segoe UI', 11, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'], padx=10, pady=10)
            month_frame.grid(row=row, column=col, padx=10, pady=10, sticky='nsew')
            self.months_container.grid_columnconfigure(col, weight=1) # Les colonnes s'étendent
            self.month_frames.append(month_frame)

            tree = ttk.Treeview(month_frame, columns=columns, show='headings', style='Custom.Treeview', height=8) # Hauteur réduite
            for c in columns:
                tree.heading(c, text=c)
                tree.column(c, width=150, anchor='w', minwidth=100)
            
            tree.bind("<Button-1>", lambda e, t=tree: self.set_active_tree(t))
            
            tree_v_scrollbar = ttk.Scrollbar(month_frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=tree_v_scrollbar.set)
            
            tree.pack(side='left', fill='both', expand=True)
            tree_v_scrollbar.pack(side='right', fill='y')
            
            self.monthly_leave_trees.append(tree)

        self.display_yearly_leave_plan()

    def display_yearly_leave_plan(self):
        """Récupère et affiche les congés, et défile jusqu'au résultat si une recherche est effectuée."""
        try:
            year = self.year_var.get()
            search_term = self.leave_search_var.get().strip()
        except AttributeError:
            year = str(datetime.now().year)
            search_term = ""

        for tree in self.monthly_leave_trees:
            for item in tree.get_children():
                tree.delete(item)

        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        base_query = '''
            SELECT l.id, l.start_date, l.end_date, e.first_name, e.last_name, e.job_title, e.department
            FROM leaves l
            JOIN employees e ON l.employee_id = e.id
            WHERE strftime('%Y', date(substr(l.start_date, 7, 4) || '-' || substr(l.start_date, 4, 2) || '-' || substr(l.start_date, 1, 2))) = ?
            AND l.status = 'Approved'
        '''
        params = [year]

        if search_term:
            base_query += " AND (e.first_name LIKE ? OR e.last_name LIKE ? OR (e.first_name || ' ' || e.last_name) LIKE ?)"
            params.extend([f'%{search_term}%', f'%{search_term}%', f'%{search_term}%'])

        base_query += " ORDER BY l.start_date"
        
        cursor.execute(base_query, params)
        all_leaves = cursor.fetchall()
        conn.close()

        if not all_leaves and search_term:
            messagebox.showinfo("Recherche", f"Aucun congé trouvé pour '{search_term}' en {year}.")

        first_month_index = None # Pour savoir vers quel mois défiler

        for leave in all_leaves:
            try:
                leave_id, start_date_str, end_date_str, first_name, last_name, job_title, department = leave
                
                if not isinstance(leave_id, int):
                    continue

                start_dt = datetime.strptime(start_date_str, '%d/%m/%Y')
                month_index = start_dt.month - 1

                if 0 <= month_index < len(self.monthly_leave_trees):
                    # On garde en mémoire le premier mois trouvé lors d'une recherche
                    if search_term and first_month_index is None:
                        first_month_index = month_index

                    tree = self.monthly_leave_trees[month_index]
                    tree.insert('', 'end', iid=leave_id, values=(
                        f"{first_name} {last_name}",
                        job_title or '',
                        department or '',
                        start_date_str,
                        end_date_str
                    ))
            except (ValueError, IndexError, TypeError) as e:
                print(f"Avertissement : Impossible d'afficher un congé. Données invalides. Erreur: {e}")
                continue

        # --- NOUVEAU : Défilement automatique vers le résultat de la recherche ---
        if search_term and first_month_index is not None:
            # On attend que l'interface se mette à jour pour avoir les bonnes coordonnées
            self.root.after(100, lambda: self.scroll_to_month(first_month_index))

    def scroll_to_month(self, month_index):
        """Fait défiler le canvas vertical pour afficher le mois spécifié."""
        try:
            # Forcer la mise à jour de la géométrie des widgets
            self.months_container.update_idletasks()
            
            target_frame = self.month_frames[month_index]
            
            # Coordonnée Y du haut du frame cible par rapport au canvas
            frame_y = target_frame.winfo_y()
            
            # Hauteur totale de la zone de défilement
            scroll_region_height = self.leave_planning_canvas.bbox("all")[3]
            
            if scroll_region_height > 0:
                # Calculer la position relative (0.0 à 1.0)
                scroll_fraction = frame_y / scroll_region_height
                self.leave_planning_canvas.yview_moveto(scroll_fraction)
        except (IndexError, tk.TclError) as e:
            print(f"Erreur lors du défilement automatique : {e}")

    def logout(self):
        """Déconnexion"""
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir vous déconnecter ?"):
            self.current_user = None
            self.current_employee_id = None
            self.show_login_screen()

    def run(self):
        """Lancer l'application et la centrer correctement"""
        # Forcer la mise à jour de la fenêtre pour avoir les dimensions
        self.root.update_idletasks()
        
        # Obtenir la taille de l'écran
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Taille de la fenêtre (définie dans __init__)
        window_width = 1400
        window_height = 900
        
        # Calculer la position pour centrer
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        
        # Appliquer la géométrie
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # Démarrer la boucle principale
        self.root.mainloop()

    

    def set_active_tree(self, tree_widget):
        """Garde en mémoire le dernier Treeview sur lequel l'utilisateur a cliqué."""
        self.active_leave_tree = tree_widget

    def _get_selected_leave_id(self):
        """Méthode interne pour récupérer et valider l'ID du congé sélectionné."""
        if not self.active_leave_tree:
            messagebox.showwarning("Attention", "Veuillez d'abord cliquer sur un congé dans le calendrier pour le sélectionner.")
            return None

        selection = self.active_leave_tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un congé dans la liste.")
            return None
        
        try:
            # On récupère l'ID (qui a été stocké comme iid) et on le convertit en nombre
            leave_id = int(selection[0])
            return leave_id
        except (ValueError, IndexError):
            messagebox.showerror("Erreur", "Impossible de récupérer l'identifiant du congé sélectionné.\nL'identifiant est invalide.")
            return None

    def modify_planned_leave(self):
        """Ouvre le formulaire pour modifier le congé sélectionné."""
        leave_id = self._get_selected_leave_id()
        if leave_id is not None:
            self.open_leave_planning_form(leave_id=leave_id)

    def delete_planned_leave(self):
        """Supprime le congé sélectionné de la base de données."""
        leave_id = self._get_selected_leave_id()
        if leave_id is None:
            return
        
        if messagebox.askyesno("Confirmation", f"Êtes-vous sûr de vouloir supprimer ce congé (ID: {leave_id}) ?\nCette action est irréversible."):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                cursor.execute("DELETE FROM leaves WHERE id = ?", (leave_id,))
                conn.commit()
                conn.close()
                
                messagebox.showinfo("Succès", "Le congé a été supprimé avec succès.")
                self.display_yearly_leave_plan()
                
            except sqlite3.Error as e:
                messagebox.showerror("Erreur de base de données", f"Une erreur est survenue : {e}")

    def open_leave_planning_form(self, leave_id=None):
        """Ouvre une fenêtre pour ajouter ou modifier un congé."""
        form_window = tk.Toplevel(self.root)
        title = "Modifier un Congé Planifié" if leave_id else "Planifier un Nouveau Congé"
        form_window.title(title)
        form_window.geometry("500x450")
        form_window.configure(bg=self.colors['background'])
        form_window.transient(self.root)
        form_window.grab_set()

        self.leave_vars = {'employee': tk.StringVar(), 'leave_type': tk.StringVar(), 'start_date': tk.StringVar(), 'end_date': tk.StringVar(), 'notes': tk.StringVar()}
        form_frame = tk.Frame(form_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20, pady=20)

        fields_config = [("Employé:", 'employee', 'combo_employee'), ("Type de Congé:", 'leave_type', 'combo_leave_type'), ("Date de Début (jj/mm/aaaa):", 'start_date', 'entry'), ("Date de Fin (jj/mm/aaaa):", 'end_date', 'entry'), ("Notes:", 'notes', 'entry')]
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        for i, (label_text, var_name, field_type) in enumerate(fields_config):
            tk.Label(form_frame, text=label_text, font=('Segoe UI', 11), bg=self.colors['background']).grid(row=i, column=0, sticky='w', pady=5)
            
            if field_type == 'entry':
                widget = tk.Entry(form_frame, textvariable=self.leave_vars[var_name], font=('Segoe UI', 11), width=40)
            elif field_type == 'combo_employee':
                cursor.execute('SELECT id, first_name, last_name FROM employees WHERE status = "Active" ORDER BY last_name')
                employees = cursor.fetchall()
                employee_choices = [f"{emp[1]} {emp[2]} (ID: {emp[0]})" for emp in employees]
                widget = ttk.Combobox(form_frame, textvariable=self.leave_vars[var_name], values=employee_choices, font=('Segoe UI', 11), width=37, state='readonly')
            elif field_type == 'combo_leave_type':
                cursor.execute('SELECT name FROM leave_types ORDER BY name')
                leave_types = [row[0] for row in cursor.fetchall()]
                widget = ttk.Combobox(form_frame, textvariable=self.leave_vars[var_name], values=leave_types, font=('Segoe UI', 11), width=37, state='readonly')

            widget.grid(row=i, column=1, sticky='w', padx=(10, 0), pady=5)
        
        if leave_id is not None:
            cursor.execute('''
                SELECT l.employee_id, lt.name, l.start_date, l.end_date, l.notes, e.first_name, e.last_name
                FROM leaves l JOIN leave_types lt ON l.leave_type_id = lt.id JOIN employees e ON l.employee_id = e.id
                WHERE l.id = ?
            ''', (leave_id,))
            data = cursor.fetchone()
            if data:
                employee_id, leave_type, start_date, end_date, notes, first_name, last_name = data
                self.leave_vars['employee'].set(f"{first_name} {last_name} (ID: {employee_id})")
                self.leave_vars['leave_type'].set(leave_type)
                self.leave_vars['start_date'].set(start_date)
                self.leave_vars['end_date'].set(end_date)
                self.leave_vars['notes'].set(notes or '')
        
        conn.close()

        buttons_frame = tk.Frame(form_frame, bg=self.colors['background'])
        buttons_frame.grid(row=len(fields_config), column=0, columnspan=2, pady=20)
        save_btn = tk.Button(buttons_frame, text="💾 Enregistrer", font=('Segoe UI', 12, 'bold'), bg=self.colors['primary_green'], fg='white', relief='flat', command=lambda: self.save_leave(form_window, leave_id=leave_id))
        save_btn.pack(side='right', padx=(10, 0))
        cancel_btn = tk.Button(buttons_frame, text="❌ Annuler", font=('Segoe UI', 12), bg=self.colors['text_light'], fg='white', relief='flat', command=form_window.destroy)
        cancel_btn.pack(side='right')

    def save_leave(self, form_window, leave_id=None):
        """Enregistre un nouveau congé ou met à jour un congé existant."""
        if not all([self.leave_vars['employee'].get(), self.leave_vars['leave_type'].get(), self.leave_vars['start_date'].get(), self.leave_vars['end_date'].get()]):
            messagebox.showerror("Erreur", "Tous les champs obligatoires doivent être remplis", parent=form_window)
            return
            
        start_date = self.leave_vars['start_date'].get()
        end_date = self.leave_vars['end_date'].get()
        
        if not self.validate_date_format(start_date) or not self.validate_date_format(end_date):
            messagebox.showerror("Erreur", "Format de date invalide. Utilisez jj/mm/aaaa", parent=form_window)
            return
            
        try:
            start_dt = datetime.strptime(start_date, '%d/%m/%Y')
            end_dt = datetime.strptime(end_date, '%d/%m/%Y')
            if end_dt < start_dt:
                messagebox.showerror("Erreur", "La date de fin doit être postérieure à la date de début", parent=form_window)
                return
            days_count = (end_dt - start_dt).days + 1
        except ValueError:
            messagebox.showerror("Erreur", "Dates invalides", parent=form_window)
            return
            
        try:
            employee_text = self.leave_vars['employee'].get()
            employee_id = int(employee_text.split('ID: ')[1].split(')')[0])
        except (IndexError, ValueError):
            messagebox.showerror("Erreur", "Employé invalide", parent=form_window)
            return
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM leave_types WHERE name = ?', (self.leave_vars['leave_type'].get(),))
        leave_type_result = cursor.fetchone()
        
        if not leave_type_result:
            messagebox.showerror("Erreur", "Type de congé invalide", parent=form_window)
            conn.close()
            return
            
        leave_type_id = leave_type_result[0]
        
        try:
            notes = self.leave_vars['notes'].get()
            if leave_id is not None:
                cursor.execute('''
                    UPDATE leaves SET
                        employee_id = ?, leave_type_id = ?, start_date = ?,
                        end_date = ?, days_count = ?, notes = ?
                    WHERE id = ?
                ''', (employee_id, leave_type_id, start_date, end_date, days_count, notes, leave_id))
                message = "Congé modifié avec succès."
            else:
                cursor.execute('''
                    INSERT INTO leaves (employee_id, leave_type_id, start_date, end_date, days_count, notes)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (employee_id, leave_type_id, start_date, end_date, days_count, notes))
                message = f"Congé enregistré avec succès ({days_count} jour(s))."
            
            conn.commit()
            messagebox.showinfo("Succès", message)
            form_window.destroy()
            self.display_yearly_leave_plan()
            
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}", parent=form_window)
        finally:
            conn.close()

    def create_leave_calendar_tab(self, parent):
        """Créer l'onglet calendrier des congés"""
        title = tk.Label(parent, text="📆 Calendrier des Congés", font=('Segoe UI', 16, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        title.pack(pady=20)
        
        nav_frame = tk.Frame(parent, bg=self.colors['background'])
        nav_frame.pack(pady=10)
        
        self.current_month = datetime.now().month
        self.current_year = datetime.now().year
        
        prev_btn = tk.Button(nav_frame, text="◀ Précédent", font=('Segoe UI', 11), bg=self.colors['accent_green'], fg=self.colors['text_dark'], relief='flat', bd=0, padx=15, pady=5, cursor='hand2', command=self.prev_month)
        prev_btn.pack(side='left', padx=5)
        
        self.month_label = tk.Label(nav_frame, text="", font=('Segoe UI', 14, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        self.month_label.pack(side='left', padx=20)
        
        next_btn = tk.Button(nav_frame, text="Suivant ▶", font=('Segoe UI', 11), bg=self.colors['accent_green'], fg=self.colors['text_dark'], relief='flat', bd=0, padx=15, pady=5, cursor='hand2', command=self.next_month)
        next_btn.pack(side='left', padx=5)
        
        self.calendar_frame = tk.Frame(parent, bg=self.colors['background'])
        self.calendar_frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        self.display_calendar()
        
    def prev_month(self):
        """Mois précédent"""
        if self.current_month == 1:
            self.current_month = 12
            self.current_year -= 1
        else:
            self.current_month -= 1
        self.display_calendar()
        
    def next_month(self):
        """Mois suivant"""
        if self.current_month == 12:
            self.current_month = 1
            self.current_year += 1
        else:
            self.current_month += 1
        self.display_calendar()
        
    def display_calendar(self):
        """Afficher le calendrier mensuel"""
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()
            
        month_names = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        self.month_label.configure(text=f"{month_names[self.current_month-1]} {self.current_year}")
        
        cal = calendar.monthcalendar(self.current_year, self.current_month)
        
        days = ['Lun', 'Mar', 'Mer', 'Jeu', 'Ven', 'Sam', 'Dim']
        for i, day in enumerate(days):
            label = tk.Label(self.calendar_frame, text=day, font=('Segoe UI', 11, 'bold'), fg='white', bg=self.colors['primary_green'], width=12, height=2)
            label.grid(row=0, column=i, padx=1, pady=1, sticky='nsew')
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT l.start_date, l.end_date, e.first_name, e.last_name
            FROM leaves l
            JOIN employees e ON l.employee_id = e.id
            WHERE l.status = 'Approved'
        ''')
        leaves_data = cursor.fetchall()
        conn.close()
        
        leaves_by_date = {}
        for start_str, end_str, first_name, last_name in leaves_data:
            try:
                start_dt = datetime.strptime(start_str, '%d/%m/%Y')
                end_dt = datetime.strptime(end_str, '%d/%m/%Y')
                current_dt = start_dt
                while current_dt <= end_dt:
                    if current_dt.month == self.current_month and current_dt.year == self.current_year:
                        day_key = current_dt.day
                        if day_key not in leaves_by_date:
                            leaves_by_date[day_key] = []
                        leaves_by_date[day_key].append(f"{first_name} {last_name}")
                    current_dt += timedelta(days=1)
            except ValueError:
                continue
                
        for week_num, week in enumerate(cal):
            for day_num, day in enumerate(week):
                if day == 0:
                    label = tk.Label(self.calendar_frame, text="", bg=self.colors['background'], width=12, height=4)
                else:
                    bg_color = self.colors['white']
                    text_color = self.colors['text_dark']
                    day_text = str(day)
                    
                    if day in leaves_by_date:
                        bg_color = self.colors['light_green']
                        text_color = 'white'
                        employees_count = len(leaves_by_date[day])
                        if employees_count > 1:
                            day_text += f"\n({employees_count})"
                        else:
                            day_text += f"\n{leaves_by_date[day][0].split()[0]}"
                            
                    label = tk.Label(self.calendar_frame, text=day_text, font=('Segoe UI', 10), fg=text_color, bg=bg_color, width=12, height=4, relief='solid', bd=1, justify='center')
                    
                    if day in leaves_by_date:
                        tooltip_text = f"Congés le {day:02d}/{self.current_month:02d}/{self.current_year}:\n" + "\n".join(leaves_by_date[day])
                        self.create_tooltip(label, tooltip_text)
                        
                label.grid(row=week_num+1, column=day_num, padx=1, pady=1, sticky='nsew')
                
        for i in range(7):
            self.calendar_frame.grid_columnconfigure(i, weight=1)
        for i in range(len(cal)+1):
            self.calendar_frame.grid_rowconfigure(i, weight=1)
            
    def create_tooltip(self, widget, text):
        """Créer un tooltip pour un widget"""
        tooltip = None
        def on_enter(event):
            nonlocal tooltip
            tooltip = tk.Toplevel(widget)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
            label = tk.Label(tooltip, text=text, font=('Segoe UI', 9), bg='#FFFFCC', fg=self.colors['text_dark'], relief='solid', bd=1, padx=5, pady=3)
            label.pack()
            
        def on_leave(event):
            nonlocal tooltip
            if tooltip:
                tooltip.destroy()
                tooltip = None
                
        widget.bind('<Enter>', on_enter)
        widget.bind('<Leave>', on_leave)
        
    def create_leave_config_tab(self, parent):
        """Créer l'onglet de configuration des congés"""
        title = tk.Label(parent, text="⚙️ Configuration des Types de Congés", font=('Segoe UI', 16, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        title.pack(pady=20)
        
        form_frame = tk.LabelFrame(parent, text="➕ Ajouter un Type de Congé", font=('Segoe UI', 12, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        form_frame.pack(fill='x', padx=20, pady=10)
        
        self.leave_type_vars = {'name': tk.StringVar(), 'days_per_year': tk.StringVar(), 'description': tk.StringVar()}
        
        fields_frame = tk.Frame(form_frame, bg=self.colors['background'])
        fields_frame.pack(padx=20, pady=15)
        
        tk.Label(fields_frame, text="Nom du Type:", font=('Segoe UI', 11), fg=self.colors['text_dark'], bg=self.colors['background']).grid(row=0, column=0, sticky='w', pady=5)
        tk.Entry(fields_frame, textvariable=self.leave_type_vars['name'], font=('Segoe UI', 11), width=30, relief='solid', bd=1).grid(row=0, column=1, padx=(10, 0), pady=5)
        
        tk.Label(fields_frame, text="Jours par An:", font=('Segoe UI', 11), fg=self.colors['text_dark'], bg=self.colors['background']).grid(row=1, column=0, sticky='w', pady=5)
        tk.Entry(fields_frame, textvariable=self.leave_type_vars['days_per_year'], font=('Segoe UI', 11), width=30, relief='solid', bd=1).grid(row=1, column=1, padx=(10, 0), pady=5)
        
        tk.Label(fields_frame, text="Description:", font=('Segoe UI', 11), fg=self.colors['text_dark'], bg=self.colors['background']).grid(row=2, column=0, sticky='w', pady=5)
        tk.Entry(fields_frame, textvariable=self.leave_type_vars['description'], font=('Segoe UI', 11), width=30, relief='solid', bd=1).grid(row=2, column=1, padx=(10, 0), pady=5)
        
        add_type_btn = tk.Button(form_frame, text="➕ Ajouter", font=('Segoe UI', 11, 'bold'), bg=self.colors['primary_green'], fg='white', relief='flat', bd=0, padx=20, pady=8, cursor='hand2', command=self.add_leave_type)
        add_type_btn.pack(pady=10)
        
        list_frame = tk.LabelFrame(parent, text="📋 Types de Congés Existants", font=('Segoe UI', 12, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        list_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        columns = ('Nom', 'Jours/An', 'Description')
        self.leave_types_tree = ttk.Treeview(list_frame, columns=columns, show='headings', style='Custom.Treeview', height=8)
        
        for col in columns:
            self.leave_types_tree.heading(col, text=col)
            self.leave_types_tree.column(col, width=300 if col == 'Description' else 150, anchor='w' if col == 'Description' else 'center')
                
        v_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.leave_types_tree.yview)
        self.leave_types_tree.configure(yscrollcommand=v_scrollbar.set)
        
        self.leave_types_tree.pack(side='left', fill='both', expand=True, padx=10, pady=10)
        v_scrollbar.pack(side='right', fill='y', padx=(0, 10), pady=10)
        
        self.load_leave_types()
        
    def add_leave_type(self):
        """Ajouter un nouveau type de congé"""
        name = self.leave_type_vars['name'].get().strip()
        days_str = self.leave_type_vars['days_per_year'].get().strip()
        description = self.leave_type_vars['description'].get().strip()
        
        if not name:
            messagebox.showerror("Erreur", "Le nom du type de congé est obligatoire")
            return
        
        try:
            days = int(days_str) if days_str else 0
        except ValueError:
            messagebox.showerror("Erreur", "Le nombre de jours doit être un entier.")
            return
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        try:
            cursor.execute('INSERT INTO leave_types (name, days_per_year, description) VALUES (?, ?, ?)', (name, days, description))
            conn.commit()
            messagebox.showinfo("Succès", "Type de congé ajouté.")
            for var in self.leave_type_vars.values():
                var.set('')
            self.load_leave_types()
        except sqlite3.IntegrityError:
            messagebox.showerror("Erreur", "Ce type de congé existe déjà.")
        except sqlite3.Error as e:
            messagebox.showerror("Erreur", f"Erreur de base de données: {e}")
        finally:
            conn.close()
            
    def load_leave_types(self):
        """Charger les types de congés"""
        if not hasattr(self, 'leave_types_tree'):
            return

        for item in self.leave_types_tree.get_children():
            self.leave_types_tree.delete(item)
            
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT name, days_per_year, description FROM leave_types ORDER BY name')
        
        for name, days, desc in cursor.fetchall():
            self.leave_types_tree.insert('', 'end', values=(name, days, desc or ""))
        conn.close()

    def show_mail_module(self):
        """Module de gestion des courriers - MISE À JOUR avec upload de fichiers"""
        self.clear_main_content()
        self.set_active_nav_button("📮 Gestion Courriers")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📮 Gestion des Courriers",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Barre d'outils
        toolbar = tk.Frame(self.main_content, bg=self.colors['background'])
        toolbar.pack(fill='x', padx=20, pady=(0, 10))
        
        # Boutons d'action
        add_btn = tk.Button(toolbar,
                           text="➕ Nouveau Courrier",
                           font=('Segoe UI', 11, 'bold'),
                           bg=self.colors['primary_green'],
                           fg='white',
                           relief='flat',
                           bd=0,
                           padx=15,
                           pady=8,
                           cursor='hand2',
                           command=self.add_new_mail)
        add_btn.pack(side='left', padx=(0, 10))
        
        refresh_btn = tk.Button(toolbar,
                               text="🔄 Rafraîchir",
                               font=('Segoe UI', 11),
                               bg=self.colors['accent_green'],
                               fg=self.colors['text_dark'],
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=self.show_mail_module)
        refresh_btn.pack(side='left')
        
        # Notebook pour les onglets
        notebook = ttk.Notebook(self.main_content, style='Custom.TNotebook')
        notebook.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        # Onglet Courriers d'Arrivée
        arrival_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(arrival_frame, text="📥 Courriers d'Arrivée")
        
        # Onglet Courriers de Départ
        departure_frame = tk.Frame(notebook, bg=self.colors['background'])
        notebook.add(departure_frame, text="📤 Courriers de Départ")
        
        # Créer les listes pour chaque onglet
        self.create_mail_list(arrival_frame, 'arrivee')
        self.create_mail_list(departure_frame, 'depart')
        
    def create_mail_list(self, parent, mail_type):
        """Créer la liste des courriers pour un type donné - MISE À JOUR avec colonne fichier"""
        # Frame de recherche
        search_frame = tk.Frame(parent, bg=self.colors['background'])
        search_frame.pack(fill='x', padx=10, pady=10)
        
        tk.Label(search_frame, text="🔍 Rechercher:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(side='left', padx=(0, 5))
        
        search_var = tk.StringVar()
        search_entry = tk.Entry(search_frame, textvariable=search_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        search_entry.pack(side='left', padx=(0, 10))
        
        search_btn = tk.Button(search_frame, text="Rechercher",
                              command=lambda: self.search_mail(mail_type, search_var.get()),
                              font=('Segoe UI', 10),
                              bg=self.colors['primary_green'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=15,
                              pady=5,
                              cursor='hand2')
        search_btn.pack(side='left')
        
        # Treeview pour la liste des courriers - AJOUT de la colonne Fichier
        columns = ('N° Ordre', 'Nb Pièces', 'Date', 'Expéditeur/Destinataire', 'Objet', 'N° Archive', 'Fichier')
        
        tree_frame = tk.Frame(parent, bg=self.colors['background'])
        tree_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                           style='Custom.Treeview', height=15)
        
        # Configuration des colonnes
        tree.heading('N° Ordre', text='N° Ordre')
        tree.heading('Nb Pièces', text='Nb Pièces')
        tree.heading('Date', text='Date')
        tree.heading('Expéditeur/Destinataire', text='Expéditeur' if mail_type == 'arrivee' else 'Destinataire')
        tree.heading('Objet', text='Objet')
        tree.heading('N° Archive', text='N° Archive')
        tree.heading('Fichier', text='Fichier')
        
        tree.column('N° Ordre', width=100)
        tree.column('Nb Pièces', width=80)
        tree.column('Date', width=100)
        tree.column('Expéditeur/Destinataire', width=200)
        tree.column('Objet', width=250)
        tree.column('N° Archive', width=100)
        tree.column('Fichier', width=80)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient='vertical', command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Charger les données
        self.load_mail_data(tree, mail_type)
        
        # Menu contextuel
        context_menu = tk.Menu(tree, tearoff=0)
        context_menu.add_command(label="✏️ Modifier",
                                command=lambda: self.edit_mail(tree))
        context_menu.add_command(label="👁️ Voir détails",
                                command=lambda: self.view_mail_details(tree))
        context_menu.add_separator()
        context_menu.add_command(label="📁 Ouvrir fichier",
                                command=lambda: self.open_mail_file(tree))
        context_menu.add_separator()
        context_menu.add_command(label="🗑️ Supprimer",
                                command=lambda: self.delete_mail(tree))
        
        def show_context_menu(event):
            try:
                context_menu.tk_popup(event.x_root, event.y_root)
            finally:
                context_menu.grab_release()
        
        tree.bind("<Button-3>", show_context_menu)
        tree.bind("<Double-1>", lambda e: self.view_mail_details(tree))
        
        # Stocker la référence du tree pour chaque type
        if mail_type == 'arrivee':
            self.arrival_tree = tree
        else:
            self.departure_tree = tree
            
    def load_mail_data(self, tree, mail_type):
        """Charger les données des courriers dans le treeview - MISE À JOUR avec fichier"""
        # Vider le treeview
        for item in tree.get_children():
            tree.delete(item)
        
        # Récupérer les données
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT numero_ordre, nombre_pieces, date_arrivee_expedition,
                   expediteur_destinataire, objet, numero_archive, file_path, id
            FROM courriers
            WHERE type_courrier = ?
            ORDER BY date_arrivee_expedition DESC
        ''', (mail_type,))
        
        for row in cursor.fetchall():
            # Formater la date
            date_str = row[2]
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = date_str
            
            # Indicateur de fichier
            file_indicator = "📄" if row[6] and os.path.exists(row[6]) else ""
            
            tree.insert('', 'end', values=(
                row[0], row[1], formatted_date, row[3], row[4], row[5] or '', file_indicator
            ), tags=(row[7],))  # Stocker l'ID dans les tags
            
        conn.close()
        
    def add_new_mail(self):
        """Ajouter un nouveau courrier"""
        self.show_mail_form()
        
    def show_mail_form(self, mail_id=None):
        """Afficher le formulaire de courrier (nouveau ou modification) - MISE À JOUR avec upload"""
        # Créer une nouvelle fenêtre
        form_window = tk.Toplevel(self.root)
        form_window.title("Nouveau Courrier" if not mail_id else "Modifier Courrier")
        form_window.geometry("600x600")  # Augmenté pour le champ fichier
        form_window.configure(bg=self.colors['background'])
        form_window.transient(self.root)
        form_window.grab_set()
        
        # Variables du formulaire
        numero_ordre_var = tk.StringVar()
        type_courrier_var = tk.StringVar(value='arrivee')
        nombre_pieces_var = tk.StringVar(value='1')
        date_var = tk.StringVar()
        expediteur_destinataire_var = tk.StringVar()
        objet_var = tk.StringVar()
        numero_archive_var = tk.StringVar()
        observation_var = tk.StringVar()
        
        # Variable pour le fichier sélectionné
        self.selected_mail_file = None
        
        # Si modification, charger les données existantes
        if mail_id:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM courriers WHERE id = ?", (mail_id,))
            mail_data = cursor.fetchone()
            conn.close()
            
            if mail_data:
                numero_ordre_var.set(mail_data[1])
                type_courrier_var.set(mail_data[2])
                nombre_pieces_var.set(str(mail_data[3]))
                # Convertir la date au format dd/mm/yyyy
                try:
                    date_obj = datetime.strptime(mail_data[4], '%Y-%m-%d')
                    date_var.set(date_obj.strftime('%d/%m/%Y'))
                except:
                    date_var.set(mail_data[4])
                expediteur_destinataire_var.set(mail_data[5])
                objet_var.set(mail_data[6])
                numero_archive_var.set(mail_data[7] or '')
                observation_var.set(mail_data[8] or '')
                self.selected_mail_file = mail_data[9]  # file_path
        
        # Titre
        title = tk.Label(form_window,
                        text="📮 " + ("Nouveau Courrier" if not mail_id else "Modifier Courrier"),
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Frame principal avec scrollbar
        canvas = tk.Canvas(form_window, bg=self.colors['background'])
        scrollbar = ttk.Scrollbar(form_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=self.colors['background'])
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Formulaire
        form_frame = tk.Frame(scrollable_frame, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        row = 0
        
        # Numéro d'ordre
        tk.Label(form_frame, text="Numéro d'ordre *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        numero_entry = tk.Entry(form_frame, textvariable=numero_ordre_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        numero_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Type de courrier
        tk.Label(form_frame, text="Type de courrier *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        type_frame = tk.Frame(form_frame, bg=self.colors['background'])
        type_frame.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        tk.Radiobutton(type_frame, text="📥 Arrivée", variable=type_courrier_var, value='arrivee',
                      font=('Segoe UI', 11), bg=self.colors['background']).pack(side='left', padx=(0, 20))
        tk.Radiobutton(type_frame, text="📤 Départ", variable=type_courrier_var, value='depart',
                      font=('Segoe UI', 11), bg=self.colors['background']).pack(side='left')
        row += 1
        
        # Nombre de pièces
        tk.Label(form_frame, text="Nombre de pièces:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        nombre_entry = tk.Entry(form_frame, textvariable=nombre_pieces_var,
                               font=('Segoe UI', 11), width=30,
                               relief='solid', bd=1)
        nombre_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Date
        tk.Label(form_frame, text="Date (dd/mm/yyyy) *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        date_entry = tk.Entry(form_frame, textvariable=date_var,
                             font=('Segoe UI', 11), width=30,
                             relief='solid', bd=1)
        date_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Expéditeur/Destinataire
        expediteur_label = tk.Label(form_frame, text="Expéditeur *:",
                                   font=('Segoe UI', 11, 'bold'),
                                   fg=self.colors['text_dark'],
                                   bg=self.colors['background'])
        expediteur_label.grid(row=row, column=0, sticky='w', pady=5)
        expediteur_entry = tk.Entry(form_frame, textvariable=expediteur_destinataire_var,
                                   font=('Segoe UI', 11), width=30,
                                   relief='solid', bd=1)
        expediteur_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Objet
        tk.Label(form_frame, text="Objet *:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        objet_entry = tk.Entry(form_frame, textvariable=objet_var,
                              font=('Segoe UI', 11), width=30,
                              relief='solid', bd=1)
        objet_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # Numéro d'archive
        tk.Label(form_frame, text="Numéro d'archive:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        archive_entry = tk.Entry(form_frame, textvariable=numero_archive_var,
                                font=('Segoe UI', 11), width=30,
                                relief='solid', bd=1)
        archive_entry.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        row += 1
        
        # NOUVEAU: Section pour l'upload de fichier
        tk.Label(form_frame, text="Fichier joint:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='w', pady=5)
        
        file_frame = tk.Frame(form_frame, bg=self.colors['background'])
        file_frame.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        select_file_btn = tk.Button(file_frame,
                                   text="📁 Sélectionner Fichier",
                                   font=('Segoe UI', 10),
                                   bg=self.colors['accent_green'],
                                   fg=self.colors['text_dark'],
                                   relief='flat',
                                   bd=0,
                                   padx=15,
                                   pady=5,
                                   cursor='hand2',
                                   command=self.select_mail_file)
        select_file_btn.pack(side='left')
        
        self.file_label = tk.Label(file_frame,
                                  text="Aucun fichier sélectionné" if not self.selected_mail_file else os.path.basename(self.selected_mail_file),
                                  font=('Segoe UI', 10),
                                  fg=self.colors['text_light'] if not self.selected_mail_file else self.colors['primary_green'],
                                  bg=self.colors['background'])
        self.file_label.pack(side='left', padx=(10, 0))
        row += 1
        
        # Observation
        tk.Label(form_frame, text="Observation:",
                font=('Segoe UI', 11, 'bold'),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).grid(row=row, column=0, sticky='nw', pady=5)
        observation_text = tk.Text(form_frame, font=('Segoe UI', 11), width=30, height=4,
                                  relief='solid', bd=1)
        observation_text.grid(row=row, column=1, sticky='ew', padx=(10, 0), pady=5)
        
        # Lier le Text widget à la variable
        if observation_var.get():
            observation_text.insert('1.0', observation_var.get())
        
        form_frame.grid_columnconfigure(1, weight=1)
        
        # Boutons
        button_frame = tk.Frame(scrollable_frame, bg=self.colors['background'])
        button_frame.pack(fill='x', padx=20, pady=20)
        
        # Bouton Annuler
        cancel_btn = tk.Button(button_frame, text="❌ Annuler",
                              command=form_window.destroy,
                              font=('Segoe UI', 11),
                              bg=self.colors['error'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=8,
                              cursor='hand2')
        cancel_btn.pack(side='right', padx=(10, 0))
        
        # Bouton Enregistrer
        save_btn = tk.Button(button_frame, text="💾 Enregistrer",
                            command=lambda: self.save_mail(
                                form_window, mail_id,
                                numero_ordre_var.get(),
                                type_courrier_var.get(),
                                nombre_pieces_var.get(),
                                date_var.get(),
                                expediteur_destinataire_var.get(),
                                objet_var.get(),
                                numero_archive_var.get(),
                                observation_text.get('1.0', 'end-1c')
                            ),
                            font=('Segoe UI', 11, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=8,
                            cursor='hand2')
        save_btn.pack(side='right')
        
        canvas.pack(side="left", fill="both", expand=True, padx=(0, 20))
        scrollbar.pack(side="right", fill="y")
        
        # Focus sur le premier champ
        numero_entry.focus()
        
    def select_mail_file(self):
        """Sélectionner un fichier pour le courrier - NOUVELLE FONCTION"""
        file_path = filedialog.askopenfilename(
            title="Sélectionner un fichier pour le courrier",
            filetypes=[
                ("Tous les fichiers", "*.*"),
                ("PDF", "*.pdf"),
                ("Images", "*.jpg *.jpeg *.png *.gif *.bmp"),
                ("Word", "*.docx *.doc"),
                ("Excel", "*.xlsx *.xls"),
                ("Texte", "*.txt")
            ]
        )
        
        if file_path:
            self.selected_mail_file = file_path
            filename = os.path.basename(file_path)
            self.file_label.configure(text=f"📄 {filename}", fg=self.colors['primary_green'])
            
    def save_mail(self, form_window, mail_id, numero_ordre, type_courrier, nombre_pieces,
                  date_str, expediteur_destinataire, objet, numero_archive, observation):
        """Enregistrer un courrier - MISE À JOUR avec gestion fichier"""
        try:
            # Validation
            if not numero_ordre.strip():
                messagebox.showerror("Erreur", "Le numéro d'ordre est obligatoire.")
                return
            
            if not date_str.strip():
                messagebox.showerror("Erreur", "La date est obligatoire.")
                return
            
            if not expediteur_destinataire.strip():
                messagebox.showerror("Erreur", "L'expéditeur/destinataire est obligatoire.")
                return
            
            if not objet.strip():
                messagebox.showerror("Erreur", "L'objet est obligatoire.")
                return
            
            # Valider et convertir la date
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                date_formatted = date_obj.strftime('%Y-%m-%d')
            except ValueError:
                messagebox.showerror("Erreur", "Format de date invalide. Utilisez dd/mm/yyyy.")
                return
            
            # Valider le nombre de pièces
            try:
                nombre_pieces_int = int(nombre_pieces) if nombre_pieces.strip() else 1
                if nombre_pieces_int < 1:
                    raise ValueError()
            except ValueError:
                messagebox.showerror("Erreur", "Le nombre de pièces doit être un nombre entier positif.")
                return
            
            # Gérer le fichier joint
            file_path = None
            if self.selected_mail_file and os.path.exists(self.selected_mail_file):
                try:
                    # Copier le fichier dans le dossier courriers
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"courrier_{numero_ordre.replace('/', '_')}_{timestamp}_{os.path.basename(self.selected_mail_file)}"
                    file_path = os.path.join(self.courriers_folder, filename)
                    shutil.copy2(self.selected_mail_file, file_path)
                except Exception as e:
                    messagebox.showerror("Erreur", f"Erreur lors de la copie du fichier: {str(e)}")
                    return
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            if mail_id:
                # Modification
                cursor.execute('''
                    UPDATE courriers SET
                        numero_ordre = ?, type_courrier = ?, nombre_pieces = ?,
                        date_arrivee_expedition = ?, expediteur_destinataire = ?,
                        objet = ?, numero_archive = ?, observation = ?, file_path = ?
                    WHERE id = ?
                ''', (numero_ordre.strip(), type_courrier, nombre_pieces_int,
                     date_formatted, expediteur_destinataire.strip(),
                     objet.strip(), numero_archive.strip() or None,
                     observation.strip() or None, file_path, mail_id))
                
                messagebox.showinfo("Succès", "Courrier modifié avec succès!")
            else:
                # Vérifier l'unicité du numéro d'ordre
                cursor.execute("SELECT id FROM courriers WHERE numero_ordre = ?", (numero_ordre.strip(),))
                if cursor.fetchone():
                    messagebox.showerror("Erreur", "Ce numéro d'ordre existe déjà.")
                    conn.close()
                    return
                
                # Nouveau courrier
                cursor.execute('''
                    INSERT INTO courriers (numero_ordre, type_courrier, nombre_pieces,
                                         date_arrivee_expedition, expediteur_destinataire,
                                         objet, numero_archive, observation, file_path, created_by)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (numero_ordre.strip(), type_courrier, nombre_pieces_int,
                     date_formatted, expediteur_destinataire.strip(),
                     objet.strip(), numero_archive.strip() or None,
                     observation.strip() or None, file_path, self.current_user['username']))
                
                messagebox.showinfo("Succès", "Courrier enregistré avec succès!")
            
            conn.commit()
            conn.close()
            form_window.destroy()
            
            # Rafraîchir la liste
            self.show_mail_module()
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'enregistrement: {str(e)}")
            
    def open_mail_file(self, tree):
        """Ouvrir le fichier joint d'un courrier - NOUVELLE FONCTION"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Récupérer le chemin du fichier
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT file_path FROM courriers WHERE id = ?", (mail_id,))
        result = cursor.fetchone()
        conn.close()
        
        if result and result[0] and os.path.exists(result[0]):
            try:
                # Ouvrir le fichier avec l'application par défaut
                if platform.system() == 'Darwin':  # macOS
                    subprocess.call(('open', result[0]))
                elif platform.system() == 'Windows':  # Windows
                    os.startfile(result[0])
                else:  # Linux
                    subprocess.call(('xdg-open', result[0]))
            except Exception as e:
                messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
        else:
            messagebox.showinfo("Information", "Aucun fichier joint à ce courrier.")
            
    def search_mail(self, mail_type, search_term):
        """Rechercher des courriers"""
        tree = self.arrival_tree if mail_type == 'arrivee' else self.departure_tree
        
        # Vider le treeview
        for item in tree.get_children():
            tree.delete(item)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        if search_term.strip():
            # Recherche avec terme
            cursor.execute('''
                SELECT numero_ordre, nombre_pieces, date_arrivee_expedition,
                       expediteur_destinataire, objet, numero_archive, file_path, id
                FROM courriers
                WHERE type_courrier = ? AND (
                    numero_ordre LIKE ? OR
                    expediteur_destinataire LIKE ? OR
                    objet LIKE ? OR
                    numero_archive LIKE ?
                )
                ORDER BY date_arrivee_expedition DESC
            ''', (mail_type, f'%{search_term}%', f'%{search_term}%',
                 f'%{search_term}%', f'%{search_term}%'))
        else:
            # Afficher tous
            self.load_mail_data(tree, mail_type)
            conn.close()
            return
        
        for row in cursor.fetchall():
            # Formater la date
            date_str = row[2]
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = date_str
            
            # Indicateur de fichier
            file_indicator = "📄" if row[6] and os.path.exists(row[6]) else ""
            
            tree.insert('', 'end', values=(
                row[0], row[1], formatted_date, row[3], row[4], row[5] or '', file_indicator
            ), tags=(row[7],))
            
        conn.close()
        
    def edit_mail(self, tree):
        """Modifier un courrier sélectionné"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier à modifier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if mail_id:
            self.show_mail_form(mail_id)
            
    def view_mail_details(self, tree):
        """Voir les détails d'un courrier - MISE À JOUR avec info fichier"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Récupérer les détails
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM courriers WHERE id = ?", (mail_id,))
        mail_data = cursor.fetchone()
        conn.close()
        
        if not mail_data:
            messagebox.showerror("Erreur", "Courrier introuvable.")
            return
        
        # Créer une fenêtre de détails
        details_window = tk.Toplevel(self.root)
        details_window.title("Détails du Courrier")
        details_window.geometry("500x500")  # Augmenté pour le fichier
        details_window.configure(bg=self.colors['background'])
        details_window.transient(self.root)
        details_window.grab_set()
        
        # Titre
        title = tk.Label(details_window,
                        text="📮 Détails du Courrier",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Détails
        details_frame = tk.Frame(details_window, bg=self.colors['background'])
        details_frame.pack(fill='both', expand=True, padx=20, pady=10)
        
        # Formater la date
        try:
            date_obj = datetime.strptime(mail_data[4], '%Y-%m-%d')
            formatted_date = date_obj.strftime('%d/%m/%Y')
        except:
            formatted_date = mail_data[4]
        
        details_info = [
            ("Numéro d'ordre:", mail_data[1]),
            ("Type:", "📥 Arrivée" if mail_data[2] == 'arrivee' else "📤 Départ"),
            ("Nombre de pièces:", str(mail_data[3])),
            ("Date:", formatted_date),
            ("Expéditeur/Destinataire:", mail_data[5]),
            ("Objet:", mail_data[6]),
            ("Numéro d'archive:", mail_data[7] or "Non spécifié"),
            ("Observation:", mail_data[8] or "Aucune"),
            ("Fichier joint:", "Oui" if mail_data[9] and os.path.exists(mail_data[9]) else "Non")
        ]
        
        for i, (label, value) in enumerate(details_info):
            tk.Label(details_frame, text=label,
                    font=('Segoe UI', 11, 'bold'),
                    fg=self.colors['text_dark'],
                    bg=self.colors['background']).grid(row=i, column=0, sticky='nw', pady=5)
            
            if label == "Observation:" and len(str(value)) > 50:
                # Utiliser un Text widget pour les longues observations
                text_widget = tk.Text(details_frame, font=('Segoe UI', 11), height=4, width=40,
                                     relief='solid', bd=1)
                text_widget.insert('1.0', str(value))
                text_widget.config(state='disabled')
                text_widget.grid(row=i, column=1, sticky='ew', padx=(10, 0), pady=5)
            else:
                tk.Label(details_frame, text=str(value),
                        font=('Segoe UI', 11),
                        fg=self.colors['text_dark'],
                        bg=self.colors['background'],
                        wraplength=300,
                        justify='left').grid(row=i, column=1, sticky='w', padx=(10, 0), pady=5)
        
        details_frame.grid_columnconfigure(1, weight=1)
        
        # Boutons
        buttons_frame = tk.Frame(details_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=20)
        
        # Bouton Ouvrir Fichier (si fichier existe)
        if mail_data[9] and os.path.exists(mail_data[9]):
            open_file_btn = tk.Button(buttons_frame, text="📁 Ouvrir Fichier",
                                     command=lambda: self.open_file_direct(mail_data[9]),
                                     font=('Segoe UI', 11),
                                     bg=self.colors['primary_green'],
                                     fg='white',
                                     relief='flat',
                                     bd=0,
                                     padx=20,
                                     pady=8,
                                     cursor='hand2')
            open_file_btn.pack(side='left')
        
        # Bouton Fermer
        close_btn = tk.Button(buttons_frame, text="✖️ Fermer",
                             command=details_window.destroy,
                             font=('Segoe UI', 11),
                             bg=self.colors['text_light'],
                             fg='white',
                             relief='flat',
                             bd=0,
                             padx=20,
                             pady=8,
                             cursor='hand2')
        close_btn.pack(side='right')
        
    def open_file_direct(self, file_path):
        """Ouvrir un fichier directement - NOUVELLE FONCTION"""
        try:
            if platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', file_path))
            elif platform.system() == 'Windows':  # Windows
                os.startfile(file_path)
            else:  # Linux
                subprocess.call(('xdg-open', file_path))
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible d'ouvrir le fichier: {str(e)}")
            
    def delete_mail(self, tree):
        """Supprimer un courrier"""
        selection = tree.selection()
        if not selection:
            messagebox.showwarning("Attention", "Veuillez sélectionner un courrier à supprimer.")
            return
        
        item = tree.item(selection[0])
        mail_id = item['tags'][0] if item['tags'] else None
        
        if not mail_id:
            return
        
        # Confirmation
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir supprimer ce courrier ?"):
            try:
                conn = sqlite3.connect(self.db_path)
                cursor = conn.cursor()
                
                # Récupérer le chemin du fichier pour le supprimer
                cursor.execute("SELECT file_path FROM courriers WHERE id = ?", (mail_id,))
                result = cursor.fetchone()
                file_path = result[0] if result else None
                
                # Supprimer le courrier de la base
                cursor.execute("DELETE FROM courriers WHERE id = ?", (mail_id,))
                conn.commit()
                conn.close()
                
                # Supprimer le fichier joint s'il existe
                if file_path and os.path.exists(file_path):
                    try:
                        os.remove(file_path)
                    except Exception as e:
                        print(f"Erreur lors de la suppression du fichier: {e}")
                
                messagebox.showinfo("Succès", "Courrier supprimé avec succès!")
                
                # Rafraîchir la liste
                self.show_mail_module()
                
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la suppression: {str(e)}")
                
    def show_reports_module(self):
        """Module de génération de rapports"""
        self.clear_main_content()
        self.set_active_nav_button("📊 Rapports")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="📊 Génération de Rapports",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les boutons de rapports
        reports_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        reports_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Configuration des rapports disponibles
        reports_config = [
            {
                'title': '👥 Liste Complète du Personnel',
                'description': 'Rapport complet de tous les employés avec leurs informations principales',
                'icon': '📋',
                'command': self.generate_staff_list_report
            },
            {
                'title': '📄 Fiche Employé Détaillée',
                'description': 'Fiche complète et imprimable d\'un employé sélectionné',
                'icon': '👤',
                'command': self.generate_employee_sheet_report
            },
            {
                'title': '🏖️ Rapport Annuel des Congés',
                'description': 'Synthèse des congés pris et soldes restants par employé',
                'icon': '📅',
                'command': self.generate_annual_leave_report
            },
            {
                'title': '📈 Statistiques RH',
                'description': 'Tableau de bord avec les principales métriques RH',
                'icon': '📊',
                'command': self.generate_hr_statistics_report
            }
        ]
        
        # Créer les cartes de rapports
        for i, report in enumerate(reports_config):
            # Frame pour chaque rapport
            report_card = tk.Frame(reports_frame, 
                                  bg='white', 
                                  relief='solid', 
                                  bd=1,
                                  padx=20,
                                  pady=20)
            report_card.grid(row=i//2, column=i%2, padx=20, pady=15, sticky='ew')
            
            # Icône et titre
            header_frame = tk.Frame(report_card, bg='white')
            header_frame.pack(fill='x', pady=(0, 10))
            
            icon_label = tk.Label(header_frame,
                                 text=report['icon'],
                                 font=('Segoe UI', 24),
                                 bg='white')
            icon_label.pack(side='left')
            
            title_label = tk.Label(header_frame,
                                  text=report['title'],
                                  font=('Segoe UI', 14, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg='white')
            title_label.pack(side='left', padx=(10, 0))
            
            # Description
            desc_label = tk.Label(report_card,
                                 text=report['description'],
                                 font=('Segoe UI', 11),
                                 fg=self.colors['text_dark'],
                                 bg='white',
                                 wraplength=300,
                                 justify='left')
            desc_label.pack(fill='x', pady=(0, 15))
            
            # Boutons d'export
            buttons_frame = tk.Frame(report_card, bg='white')
            buttons_frame.pack(fill='x')
            
            # Bouton PDF
            pdf_btn = tk.Button(buttons_frame,
                               text="📄 Générer PDF",
                               font=('Segoe UI', 10, 'bold'),
                               bg=self.colors['primary_green'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=15,
                               pady=8,
                               cursor='hand2',
                               command=lambda cmd=report['command']: cmd('pdf'))
            pdf_btn.pack(side='left', padx=(0, 10))
            
            # Bouton Excel
            excel_btn = tk.Button(buttons_frame,
                                 text="📊 Générer Excel",
                                 font=('Segoe UI', 10, 'bold'),
                                 bg=self.colors['success'],
                                 fg='white',
                                 relief='flat',
                                 bd=0,
                                 padx=15,
                                 pady=8,
                                 cursor='hand2',
                                 command=lambda cmd=report['command']: cmd('excel'))
            excel_btn.pack(side='left')
            
        # Configurer les colonnes pour qu'elles s'étendent uniformément
        reports_frame.grid_columnconfigure(0, weight=1)
        reports_frame.grid_columnconfigure(1, weight=1)
        
    def generate_staff_list_report(self, format_type):
        """Générer le rapport de liste du personnel"""
        try:
            # Récupérer les données
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT matricule, first_name, last_name, job_title, department, 
                       hire_date, contract_type, status, phone, email
                FROM employees 
                ORDER BY last_name, first_name
            ''')
            employees = cursor.fetchall()
            conn.close()
            
            if not employees:
                messagebox.showwarning("Attention", "Aucun employé trouvé")
                return
                
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"liste_personnel_{timestamp}.pdf"
                self.create_staff_list_pdf(employees, filename)
            else:  # excel
                filename = f"liste_personnel_{timestamp}.xlsx"
                self.create_staff_list_excel(employees, filename)
                
            messagebox.showinfo("Succès", f"Rapport généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération du rapport: {str(e)}")
            
    def create_staff_list_pdf(self, employees, filename):
        """Créer le PDF de la liste du personnel"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph("Liste du Personnel ", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tableau des employés
        data = [['Matricule', 'Nom Complet', 'Poste', 'Département', 'Embauche', 'Statut']]
        
        for emp in employees:
            matricule, first_name, last_name, job_title, department, hire_date, contract_type, status, phone, email = emp
            data.append([
                matricule,
                f"{first_name} {last_name}",
                job_title or '',
                department or '',
                hire_date or '',
                status
            ])
            
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        
    def create_staff_list_excel(self, employees, filename):
        """Créer le fichier Excel de la liste du personnel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Liste Personnel"
        
        # Titre
        ws['A1'] = "Liste du Personnel"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:F1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        # En-têtes
        headers = ['Matricule', 'Nom Complet', 'Poste', 'Département', 'Date Embauche', 'Statut', 'Téléphone', 'Email']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Données
        for row, emp in enumerate(employees, 5):
            matricule, first_name, last_name, job_title, department, hire_date, contract_type, status, phone, email = emp
            
            ws.cell(row=row, column=1, value=matricule)
            ws.cell(row=row, column=2, value=f"{first_name} {last_name}")
            ws.cell(row=row, column=3, value=job_title or '')
            ws.cell(row=row, column=4, value=department or '')
            ws.cell(row=row, column=5, value=hire_date or '')
            ws.cell(row=row, column=6, value=status)
            ws.cell(row=row, column=7, value=phone or '')
            ws.cell(row=row, column=8, value=email or '')
            
        # Ajuster les largeurs de colonnes
        # Ajuster les largeurs de colonnes (version corrigée)

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                # Ignorer les cellules fusionnées
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value:
                        # Ajouter 2 pour un peu d'espace
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Définir une largeur minimale et maximale
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)   
        wb.save(filename)
        
    def generate_employee_sheet_report(self, format_type):
        """Générer la fiche détaillée d'un employé"""
        # Sélectionner un employé
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT id, first_name, last_name, matricule FROM employees ORDER BY last_name')
        employees = cursor.fetchall()
        conn.close()
        
        if not employees:
            messagebox.showwarning("Attention", "Aucun employé trouvé")
            return
            
        # Créer une fenêtre de sélection
        selection_window = tk.Toplevel(self.root)
        selection_window.title("Sélectionner un Employé")
        selection_window.geometry("400x300")
        selection_window.configure(bg=self.colors['background'])
        selection_window.transient(self.root)
        selection_window.grab_set()
        
        tk.Label(selection_window,
                text="Sélectionnez un employé:",
                font=('Segoe UI', 12, 'bold'),
                fg=self.colors['primary_green'],
                bg=self.colors['background']).pack(pady=20)
        
        # Liste des employés
        listbox = tk.Listbox(selection_window,
                            font=('Segoe UI', 11),
                            height=10)
        listbox.pack(fill='both', expand=True, padx=20, pady=(0, 20))
        
        for emp in employees:
            emp_id, first_name, last_name, matricule = emp
            listbox.insert(tk.END, f"{first_name} {last_name} (Matricule: {matricule})")
            
        # Boutons
        buttons_frame = tk.Frame(selection_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=20, pady=(0, 20))
        
        def generate_selected():
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("Attention", "Veuillez sélectionner un employé")
                return
                
            selected_emp = employees[selection[0]]
            selection_window.destroy()
            self.create_employee_sheet_report(selected_emp[0], format_type)
            
        tk.Button(buttons_frame,
                 text=f"Générer {format_type.upper()}",
                 font=('Segoe UI', 11, 'bold'),
                 bg=self.colors['primary_green'],
                 fg='white',
                 relief='flat',
                 bd=0,
                 padx=20,
                 pady=8,
                 cursor='hand2',
                 command=generate_selected).pack(side='right', padx=(10, 0))
        
        tk.Button(buttons_frame,
                 text="Annuler",
                 font=('Segoe UI', 11),
                 bg=self.colors['text_light'],
                 fg='white',
                 relief='flat',
                 bd=0,
                 padx=20,
                 pady=8,
                 cursor='hand2',
                 command=selection_window.destroy).pack(side='right')
                 
    def create_employee_sheet_report(self, employee_id, format_type):
        """Créer la fiche détaillée d'un employé"""
        try:
            # Récupérer les données complètes de l'employé
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT * FROM employees WHERE id = ?', (employee_id,))
            employee = cursor.fetchone()
            
            if not employee:
                messagebox.showerror("Erreur", "Employé non trouvé")
                return
                
            # Récupérer l'historique de carrière
            cursor.execute('''
                SELECT act_number, nature, subject, act_date, effective_date
                FROM career_history 
                WHERE employee_id = ?
                ORDER BY act_date DESC
            ''', (employee_id,))
            career_history = cursor.fetchall()
            
            # Récupérer les congés
            cursor.execute('''
                SELECT lt.name, l.start_date, l.end_date, l.days_count, l.status
                FROM leaves l
                JOIN leave_types lt ON l.leave_type_id = lt.id
                WHERE l.employee_id = ?
                ORDER BY l.start_date DESC
                LIMIT 10
            ''', (employee_id,))
            recent_leaves = cursor.fetchall()
            
            conn.close()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            employee_name = f"{employee[2]}_{employee[3]}"
            
            if format_type == 'pdf':
                filename = f"fiche_{employee_name}_{timestamp}.pdf"
                self.create_employee_sheet_pdf(employee, career_history, recent_leaves, filename)
            else:  # excel
                filename = f"fiche_{employee_name}_{timestamp}.xlsx"
                self.create_employee_sheet_excel(employee, career_history, recent_leaves, filename)
                
            messagebox.showinfo("Succès", f"Fiche employé générée: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_employee_sheet_pdf(self, employee, career_history, recent_leaves, filename):
        """Créer le PDF de la fiche employé"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph(f"Fiche Employé - {employee[2]} {employee[3]}", title_style))
        story.append(Spacer(1, 20))
        
        # Informations personnelles
        story.append(Paragraph("Informations Personnelles", styles['Heading2']))
        
        personal_data = [
            ['Matricule:', employee[1]],
            ['Nom Complet:', f"{employee[2]} {employee[3]}"],
            ['Genre:', employee[4] or ''],
            ['Date de Naissance:', employee[5] or ''],
            ['Lieu de Naissance:', employee[6] or ''],
            ['Adresse:', employee[7] or ''],
            ['Téléphone:', employee[8] or ''],
            ['Email:', employee[9] or ''],
            ['Situation Matrimoniale:', employee[10] or ''],
            ['Personnes à Charge:', str(employee[11]) if employee[11] else '0']
        ]
        
        personal_table = Table(personal_data, colWidths=[150, 300])
        personal_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(personal_table)
        story.append(Spacer(1, 20))
        
        # Informations contractuelles
        story.append(Paragraph("Informations Contractuelles", styles['Heading2']))
        
        contract_data = [
            ['Date d\'Embauche:', employee[14] or ''],
            ['Type de Contrat:', employee[15] or ''],
            ['Début de Contrat:', employee[16] or ''],
            ['Fin de Contrat:', employee[17] or ''],
            ['Département:', employee[18] or ''],
            ['Poste/Fonction:', employee[19] or ''],
            ['Statut:', employee[20] or '']
        ]
        
        contract_table = Table(contract_data, colWidths=[150, 300])
        contract_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(contract_table)
        story.append(Spacer(1, 20))
        
        # Historique de carrière (si disponible)
        if career_history:
            story.append(Paragraph("Historique de Carrière", styles['Heading2']))
            
            career_data = [['N° Acte', 'Nature', 'Date Acte', 'Date Effet']]
            for act in career_history[:5]:  # Limiter à 5 derniers actes
                career_data.append([
                    act[0] or '',
                    act[1] or '',
                    act[3] or '',
                    act[4] or ''
                ])
                
            career_table = Table(career_data)
            career_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(career_table)
            
        doc.build(story)
        
    def create_employee_sheet_excel(self, employee, career_history, recent_leaves, filename):
        """Créer le fichier Excel de la fiche employé"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fiche Employé"
        
        # Titre
        ws['A1'] = f"Fiche Employé - {employee[2]} {employee[3]}"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:D1')
        
        row = 3
        
        # Informations personnelles
        ws[f'A{row}'] = "INFORMATIONS PERSONNELLES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        personal_fields = [
            ('Matricule:', employee[1]),
            ('Nom Complet:', f"{employee[2]} {employee[3]}"),
            ('Genre:', employee[4]),
            ('Date de Naissance:', employee[5]),
            ('Lieu de Naissance:', employee[6]),
            ('Adresse:', employee[7]),
            ('Téléphone:', employee[8]),
            ('Email:', employee[9]),
            ('Situation Matrimoniale:', employee[10]),
            ('Personnes à Charge:', employee[11])
        ]
        
        for label, value in personal_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value or ''
            row += 1
            
        row += 2
        
        # Informations contractuelles
        ws[f'A{row}'] = "INFORMATIONS CONTRACTUELLES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        contract_fields = [
            ('Date d\'Embauche:', employee[14]),
            ('Type de Contrat:', employee[15]),
            ('Début de Contrat:', employee[16]),
            ('Fin de Contrat:', employee[17]),
            ('Département:', employee[18]),
            ('Poste/Fonction:', employee[19]),
            ('Statut:', employee[20])
        ]
        
        for label, value in contract_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value or ''
            row += 1
            
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        
        wb.save(filename)
        
    def generate_annual_leave_report(self, format_type):
        """Générer le rapport annuel des congés"""
        try:
            current_year = datetime.now().year
            
            # Récupérer les données des congés
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT e.matricule, e.first_name, e.last_name, 
                       COUNT(l.id) as total_leaves,
                       SUM(l.days_count) as total_days,
                       GROUP_CONCAT(lt.name || ': ' || l.days_count || ' jours', '; ') as leave_details
                FROM employees e
                LEFT JOIN leaves l ON e.id = l.employee_id 
                    AND strftime('%Y', date(substr(l.start_date, 7, 4) || '-' || substr(l.start_date, 4, 2) || '-' || substr(l.start_date, 1, 2))) = ?
                LEFT JOIN leave_types lt ON l.leave_type_id = lt.id
                WHERE e.status = 'Active'
                GROUP BY e.id, e.matricule, e.first_name, e.last_name
                ORDER BY e.last_name, e.first_name
            ''', (str(current_year),))
            
            leave_data = cursor.fetchall()
            conn.close()
            
            if not leave_data:
                messagebox.showwarning("Attention", "Aucune donnée de congé trouvée")
                return
                
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"rapport_conges_{current_year}_{timestamp}.pdf"
                self.create_annual_leave_pdf(leave_data, current_year, filename)
            else:  # excel
                filename = f"rapport_conges_{current_year}_{timestamp}.xlsx"
                self.create_annual_leave_excel(leave_data, current_year, filename)
                
            messagebox.showinfo("Succès", f"Rapport des congés généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_annual_leave_pdf(self, leave_data, year, filename):
        """Créer le PDF du rapport annuel des congés"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph(f"Rapport Annuel des Congés - {year}", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tableau des congés
        data = [['Matricule', 'Employé', 'Nb Congés', 'Total Jours', 'Solde Restant']]
        
        for emp_data in leave_data:
            matricule, first_name, last_name, total_leaves, total_days, leave_details = emp_data
            
            # Calcul du solde (30 jours par défaut - jours pris)
            annual_allowance = 30
            days_taken = total_days or 0
            remaining_balance = annual_allowance - days_taken
            
            data.append([
                matricule,
                f"{first_name} {last_name}",
                str(total_leaves or 0),
                str(days_taken),
                str(remaining_balance)
            ])
            
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(table)
        doc.build(story)
        
    def create_annual_leave_excel(self, leave_data, year, filename):
        """Créer le fichier Excel du rapport annuel des congés"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Congés {year}"
        
        # Titre
        ws['A1'] = f"Rapport Annuel des Congés - {year}"
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:E1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        # En-têtes
        headers = ['Matricule', 'Employé', 'Nb Congés Pris', 'Total Jours Pris', 'Solde Restant']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Données
        for row, emp_data in enumerate(leave_data, 5):
            matricule, first_name, last_name, total_leaves, total_days, leave_details = emp_data
            
            # Calcul du solde
            annual_allowance = 30
            days_taken = total_days or 0
            remaining_balance = annual_allowance - days_taken
            
            ws.cell(row=row, column=1, value=matricule)
            ws.cell(row=row, column=2, value=f"{first_name} {last_name}")
            ws.cell(row=row, column=3, value=total_leaves or 0)
            ws.cell(row=row, column=4, value=days_taken)
            ws.cell(row=row, column=5, value=remaining_balance)
            
            # Colorer en rouge si solde négatif
            if remaining_balance < 0:
                ws.cell(row=row, column=5).font = Font(color='FF0000', bold=True)
                
        # Ajuster les largeurs de colonnes
        # Ajuster les largeurs de colonnes (version corrigée)
        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[column_letter]:
                # Ignorer les cellules fusionnées
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                try:
                    if cell.value:
                        # Ajouter 2 pour un peu d'espace
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            # Définir une largeur minimale et maximale
            adjusted_width = max(max_length + 2, 15)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 40)
        wb.save(filename)
        
    def generate_hr_statistics_report(self, format_type):
        """Générer le rapport de statistiques RH"""
        try:
            # Récupérer toutes les statistiques
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Statistiques générales
            cursor.execute('SELECT COUNT(*) FROM employees WHERE status = "Active"')
            total_active = cursor.fetchone()[0]
            
            cursor.execute('SELECT COUNT(*) FROM employees')
            total_employees = cursor.fetchone()[0]
            
            # Répartition par département
            cursor.execute('''
                SELECT department, COUNT(*) 
                FROM employees 
                WHERE status = "Active" AND department IS NOT NULL AND department != ""
                GROUP BY department 
                ORDER BY COUNT(*) DESC
            ''')
            dept_stats = cursor.fetchall()
            
            # Répartition par type de contrat
            cursor.execute('''
                SELECT contract_type, COUNT(*) 
                FROM employees 
                WHERE status = "Active" AND contract_type IS NOT NULL AND contract_type != ""
                GROUP BY contract_type
            ''')
            contract_stats = cursor.fetchall()
            
            # Statistiques des congés (année courante)
            current_year = datetime.now().year
            cursor.execute('''
                SELECT COUNT(*), SUM(days_count)
                FROM leaves 
                WHERE strftime('%Y', date(substr(start_date, 7, 4) || '-' || substr(start_date, 4, 2) || '-' || substr(start_date, 1, 2))) = ?
            ''', (str(current_year),))
            leave_stats = cursor.fetchone()
            
            conn.close()
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            if format_type == 'pdf':
                filename = f"statistiques_rh_{timestamp}.pdf"
                self.create_hr_statistics_pdf(
                    total_active, total_employees, dept_stats, 
                    contract_stats, leave_stats, filename
                )
            else:  # excel
                filename = f"statistiques_rh_{timestamp}.xlsx"
                self.create_hr_statistics_excel(
                    total_active, total_employees, dept_stats, 
                    contract_stats, leave_stats, filename
                )
                
            messagebox.showinfo("Succès", f"Rapport statistiques généré: {filename}")
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la génération: {str(e)}")
            
    def create_hr_statistics_pdf(self, total_active, total_employees, dept_stats, contract_stats, leave_stats, filename):
        """Créer le PDF des statistiques RH"""
        doc = SimpleDocTemplate(filename, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Titre
        title_style = styles['Title']
        title_style.textColor = colors.HexColor(self.colors['primary_green'])
        story.append(Paragraph("Statistiques RH ", title_style))
        story.append(Spacer(1, 20))
        
        # Date de génération
        story.append(Paragraph(f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
        story.append(Spacer(1, 30))
        
        # Statistiques générales
        story.append(Paragraph("Statistiques Générales", styles['Heading2']))
        
        general_data = [
            ['Total Employés:', str(total_employees)],
            ['Employés Actifs:', str(total_active)],
            ['Taux d\'Activité:', f"{(total_active/total_employees*100):.1f}%" if total_employees > 0 else "0%"]
        ]
        
        general_table = Table(general_data, colWidths=[200, 100])
        general_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(general_table)
        story.append(Spacer(1, 20))
        
        # Répartition par département
        if dept_stats:
            story.append(Paragraph("Répartition par Département", styles['Heading2']))
            
            dept_data = [['Département', 'Nombre d\'Employés', 'Pourcentage']]
            for dept, count in dept_stats:
                percentage = (count / total_active * 100) if total_active > 0 else 0
                dept_data.append([dept, str(count), f"{percentage:.1f}%"])
                
            dept_table = Table(dept_data)
            dept_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(self.colors['primary_green'])),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            story.append(dept_table)
            story.append(Spacer(1, 20))
            
        # Statistiques des congés
        story.append(Paragraph(f"Statistiques des Congés {datetime.now().year}", styles['Heading2']))
        
        total_leave_requests = leave_stats[0] or 0
        total_leave_days = leave_stats[1] or 0
        
        leave_data = [
            ['Total Demandes de Congés:', str(total_leave_requests)],
            ['Total Jours de Congés:', str(total_leave_days)],
            ['Moyenne par Demande:', f"{(total_leave_days/total_leave_requests):.1f} jours" if total_leave_requests > 0 else "0 jours"]
        ]
        
        leave_table = Table(leave_data, colWidths=[200, 100])
        leave_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        story.append(leave_table)
        
        doc.build(story)
        
    def create_hr_statistics_excel(self, total_active, total_employees, dept_stats, contract_stats, leave_stats, filename):
        """Créer le fichier Excel des statistiques RH"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Statistiques RH"
        
        # Titre
        ws['A1'] = "Statistiques RH "
        ws['A1'].font = Font(size=16, bold=True, color='2E7D32')
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A2'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = Font(size=10)
        
        row = 4
        
        # Statistiques générales
        ws[f'A{row}'] = "STATISTIQUES GÉNÉRALES"
        ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
        row += 2
        
        ws[f'A{row}'] = "Total Employés:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_employees
        row += 1
        
        ws[f'A{row}'] = "Employés Actifs:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total_active
        row += 1
        
        ws[f'A{row}'] = "Taux d'Activité:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{(total_active/total_employees*100):.1f}%" if total_employees > 0 else "0%"
        row += 3
        
        # Répartition par département
        if dept_stats:
            ws[f'A{row}'] = "RÉPARTITION PAR DÉPARTEMENT"
            ws[f'A{row}'].font = Font(size=12, bold=True, color='2E7D32')
            row += 2
            
            # En-têtes
            headers = ['Département', 'Nombre', 'Pourcentage']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
            row += 1
            
            # Données
            for dept, count in dept_stats:
                percentage = (count / total_active * 100) if total_active > 0 else 0
                ws.cell(row=row, column=1, value=dept)
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                row += 1
                
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        wb.save(filename)
        
    def show_settings_module(self):
        """Module de configuration"""
        self.clear_main_content()
        self.set_active_nav_button("⚙️ Configuration")
        
        # Titre
        title = tk.Label(self.main_content,
                        text="⚙️ Configuration du Système",
                        font=('Segoe UI', 18, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 30))
        
        # Frame pour les options de configuration
        config_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        config_frame.pack(fill='both', expand=True, padx=50, pady=20)
        
        # Section Utilisateurs
        users_frame = tk.LabelFrame(config_frame,
                                   text="👥 Gestion des Utilisateurs",
                                   font=('Segoe UI', 14, 'bold'),
                                   fg=self.colors['primary_green'],
                                   bg=self.colors['background'],
                                   padx=20,
                                   pady=20)
        users_frame.pack(fill='x', pady=(0, 20))
        
        # Boutons de gestion des utilisateurs
        users_buttons_frame = tk.Frame(users_frame, bg=self.colors['background'])
        users_buttons_frame.pack(fill='x')
        
        add_user_btn = tk.Button(users_buttons_frame,
                                text="➕ Ajouter Utilisateur",
                                font=('Segoe UI', 11, 'bold'),
                                bg=self.colors['primary_green'],
                                fg='white',
                                relief='flat',
                                bd=0,
                                padx=20,
                                pady=10,
                                cursor='hand2',
                                command=self.add_user)
        add_user_btn.pack(side='left', padx=(0, 10))
        
        change_password_btn = tk.Button(users_buttons_frame,
                                       text="🔑 Changer Mot de Passe",
                                       font=('Segoe UI', 11, 'bold'),
                                       bg=self.colors['accent_green'],
                                       fg=self.colors['text_dark'],
                                       relief='flat',
                                       bd=0,
                                       padx=20,
                                       pady=10,
                                       cursor='hand2',
                                       command=self.change_password)
        change_password_btn.pack(side='left')
        
        # Section Base de Données
        db_frame = tk.LabelFrame(config_frame,
                                text="💾 Base de Données",
                                font=('Segoe UI', 14, 'bold'),
                                fg=self.colors['primary_green'],
                                bg=self.colors['background'],
                                padx=20,
                                pady=20)
        db_frame.pack(fill='x', pady=(0, 20))
        
        # Boutons de gestion de la base de données
        db_buttons_frame = tk.Frame(db_frame, bg=self.colors['background'])
        db_buttons_frame.pack(fill='x')
        
        backup_btn = tk.Button(db_buttons_frame,
                              text="💾 Sauvegarder",
                              font=('Segoe UI', 11, 'bold'),
                              bg=self.colors['success'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=self.backup_database)
        backup_btn.pack(side='left', padx=(0, 10))
        
        restore_btn = tk.Button(db_buttons_frame,
                               text="📁 Restaurer",
                               font=('Segoe UI', 11, 'bold'),
                               bg=self.colors['warning'],
                               fg='white',
                               relief='flat',
                               bd=0,
                               padx=20,
                               pady=10,
                               cursor='hand2',
                               command=self.restore_database)
        restore_btn.pack(side='left')
        
        # Section Informations Système
        info_frame = tk.LabelFrame(config_frame,
                                  text="ℹ️ Informations Système",
                                  font=('Segoe UI', 14, 'bold'),
                                  fg=self.colors['primary_green'],
                                  bg=self.colors['background'],
                                  padx=20,
                                  pady=20)
        info_frame.pack(fill='x')
        
        # Informations système
        info_text = tk.Text(info_frame,
                           font=('Segoe UI', 10),
                           bg=self.colors['light_gray'],
                           fg=self.colors['text_dark'],
                           relief='flat',
                           height=8,
                           wrap='word',
                           state='disabled')
        info_text.pack(fill='both', expand=True)
        
        # Remplir les informations système
        system_info = f"""
Version de l'Application: 1.0.0
Base de Données: SQLite ({self.db_path})
Dossier Documents: {os.path.abspath(self.documents_folder)}
Dossier Photos: {os.path.abspath(self.photos_folder)}
Dossier Courriers: {os.path.abspath(self.courriers_folder)}

Statistiques:
- Nombre total d'employés: {self.get_total_employees()}
- Nombre d'utilisateurs: {self.get_total_users()}
- Taille de la base de données: {self.get_db_size()}

Développé pour Sen pro
© 2025 - Système de Gestion RH
        """
        
        info_text.config(state='normal')
        info_text.insert('1.0', system_info.strip())
        info_text.config(state='disabled')
        
    def add_user(self):
        """Ajouter un nouvel utilisateur"""
        # Fenêtre de saisie
        user_window = tk.Toplevel(self.root)
        user_window.title("Ajouter Utilisateur")
        user_window.geometry("400x300")
        user_window.configure(bg=self.colors['background'])
        user_window.transient(self.root)
        user_window.grab_set()
        
        # Variables
        username_var = tk.StringVar()
        password_var = tk.StringVar()
        role_var = tk.StringVar(value='user')
        
        # Titre
        title = tk.Label(user_window,
                        text="👤 Nouvel Utilisateur",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(user_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=30)
        
        # Nom d'utilisateur
        tk.Label(form_frame,
                text="Nom d'utilisateur:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        username_entry = tk.Entry(form_frame,
                                 textvariable=username_var,
                                 font=('Segoe UI', 11),
                                 width=30,
                                 relief='solid',
                                 bd=1)
        username_entry.pack(fill='x', pady=(0, 15))
        
        # Mot de passe
        tk.Label(form_frame,
                text="Mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        password_entry = tk.Entry(form_frame,
                                 textvariable=password_var,
                                 font=('Segoe UI', 11),
                                 width=30,
                                 show='*',
                                 relief='solid',
                                 bd=1)
        password_entry.pack(fill='x', pady=(0, 15))
        
        # Rôle
        tk.Label(form_frame,
                text="Rôle:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        role_combo = ttk.Combobox(form_frame,
                                 textvariable=role_var,
                                 values=['user', 'admin'],
                                 font=('Segoe UI', 11),
                                 width=27,
                                 state='readonly')
        role_combo.pack(fill='x', pady=(0, 20))
        
        # Boutons
        buttons_frame = tk.Frame(user_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=30, pady=(0, 20))
        
        def save_user():
            username = username_var.get().strip()
            password = password_var.get().strip()
            role = role_var.get()
            
            if not username or not password:
                messagebox.showerror("Erreur", "Tous les champs sont obligatoires")
                return
                
            # Vérifier l'unicité
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
            
            if cursor.fetchone():
                messagebox.showerror("Erreur", "Ce nom d'utilisateur existe déjà")
                conn.close()
                return
                
            # Créer l'utilisateur
            try:
                password_hash = hashlib.sha256(password.encode()).hexdigest()
                cursor.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                              (username, password_hash, role))
                conn.commit()
                messagebox.showinfo("Succès", "Utilisateur créé avec succès")
                user_window.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur lors de la création: {str(e)}")
            finally:
                conn.close()
                
        save_btn = tk.Button(buttons_frame,
                            text="💾 Créer",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=save_user)
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=user_window.destroy)
        cancel_btn.pack(side='right')
        
        username_entry.focus()
        
    def change_password(self):
        """Changer le mot de passe de l'utilisateur actuel"""
        # Fenêtre de changement de mot de passe
        pwd_window = tk.Toplevel(self.root)
        pwd_window.title("Changer Mot de Passe")
        pwd_window.geometry("400x250")
        pwd_window.configure(bg=self.colors['background'])
        pwd_window.transient(self.root)
        pwd_window.grab_set()
        
        # Variables
        current_pwd_var = tk.StringVar()
        new_pwd_var = tk.StringVar()
        confirm_pwd_var = tk.StringVar()
        
        # Titre
        title = tk.Label(pwd_window,
                        text="🔑 Changer Mot de Passe",
                        font=('Segoe UI', 16, 'bold'),
                        fg=self.colors['primary_green'],
                        bg=self.colors['background'])
        title.pack(pady=(20, 20))
        
        # Formulaire
        form_frame = tk.Frame(pwd_window, bg=self.colors['background'])
        form_frame.pack(fill='both', expand=True, padx=30)
        
        # Mot de passe actuel
        tk.Label(form_frame,
                text="Mot de passe actuel:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        current_pwd_entry = tk.Entry(form_frame,
                                    textvariable=current_pwd_var,
                                    font=('Segoe UI', 11),
                                    width=30,
                                    show='*',
                                    relief='solid',
                                    bd=1)
        current_pwd_entry.pack(fill='x', pady=(0, 10))
        
        # Nouveau mot de passe
        tk.Label(form_frame,
                text="Nouveau mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        new_pwd_entry = tk.Entry(form_frame,
                                textvariable=new_pwd_var,
                                font=('Segoe UI', 11),
                                width=30,
                                show='*',
                                relief='solid',
                                bd=1)
        new_pwd_entry.pack(fill='x', pady=(0, 10))
        
        # Confirmer nouveau mot de passe
        tk.Label(form_frame,
                text="Confirmer nouveau mot de passe:",
                font=('Segoe UI', 11),
                fg=self.colors['text_dark'],
                bg=self.colors['background']).pack(anchor='w', pady=(0, 5))
        
        confirm_pwd_entry = tk.Entry(form_frame,
                                    textvariable=confirm_pwd_var,
                                    font=('Segoe UI', 11),
                                    width=30,
                                    show='*',
                                    relief='solid',
                                    bd=1)
        confirm_pwd_entry.pack(fill='x', pady=(0, 15))
        
        # Boutons
        buttons_frame = tk.Frame(pwd_window, bg=self.colors['background'])
        buttons_frame.pack(fill='x', padx=30, pady=(0, 20))
        
        def save_password():
            current_pwd = current_pwd_var.get()
            new_pwd = new_pwd_var.get()
            confirm_pwd = confirm_pwd_var.get()
            
            if not all([current_pwd, new_pwd, confirm_pwd]):
                messagebox.showerror("Erreur", "Tous les champs sont obligatoires")
                return
                
            if new_pwd != confirm_pwd:
                messagebox.showerror("Erreur", "Les nouveaux mots de passe ne correspondent pas")
                return
                
            if len(new_pwd) < 4:
                messagebox.showerror("Erreur", "Le mot de passe doit contenir au moins 4 caractères")
                return
                
            # Vérifier le mot de passe actuel
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            current_pwd_hash = hashlib.sha256(current_pwd.encode()).hexdigest()
            cursor.execute('SELECT id FROM users WHERE id = ? AND password_hash = ?',
                          (self.current_user['id'], current_pwd_hash))
            
            if not cursor.fetchone():
                messagebox.showerror("Erreur", "Mot de passe actuel incorrect")
                conn.close()
                return
                
            # Mettre à jour le mot de passe
            try:
                new_pwd_hash = hashlib.sha256(new_pwd.encode()).hexdigest()
                cursor.execute('UPDATE users SET password_hash = ? WHERE id = ?',
                              (new_pwd_hash, self.current_user['id']))
                conn.commit()
                messagebox.showinfo("Succès", "Mot de passe modifié avec succès")
                pwd_window.destroy()
            except sqlite3.Error as e:
                messagebox.showerror("Erreur", f"Erreur lors de la modification: {str(e)}")
            finally:
                conn.close()
                
        save_btn = tk.Button(buttons_frame,
                            text="💾 Modifier",
                            font=('Segoe UI', 12, 'bold'),
                            bg=self.colors['primary_green'],
                            fg='white',
                            relief='flat',
                            bd=0,
                            padx=20,
                            pady=10,
                            cursor='hand2',
                            command=save_password)
        save_btn.pack(side='right', padx=(10, 0))
        
        cancel_btn = tk.Button(buttons_frame,
                              text="❌ Annuler",
                              font=('Segoe UI', 12),
                              bg=self.colors['text_light'],
                              fg='white',
                              relief='flat',
                              bd=0,
                              padx=20,
                              pady=10,
                              cursor='hand2',
                              command=pwd_window.destroy)
        cancel_btn.pack(side='right')
        
        current_pwd_entry.focus()
        
    def backup_database(self):
        """Sauvegarder la base de données"""
        try:
            # Sélectionner le dossier de destination
            backup_path = filedialog.asksaveasfilename(
                title="Sauvegarder la base de données",
                defaultextension=".db",
                filetypes=[("Base de données SQLite", "*.db"), ("Tous les fichiers", "*.*")],
                initialname=f"hr_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
            )
            
            if backup_path:
                # Copier la base de données
                shutil.copy2(self.db_path, backup_path)
                messagebox.showinfo("Succès", f"Base de données sauvegardée avec succès:\n{backup_path}")
                
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la sauvegarde: {str(e)}")
            
    def restore_database(self):
        """Restaurer la base de données"""
        if messagebox.askyesno("Confirmation", 
                              "Êtes-vous sûr de vouloir restaurer la base de données ?\n\nCette action remplacera toutes les données actuelles."):
            
            try:
                # Sélectionner le fichier de sauvegarde
                backup_file = filedialog.askopenfilename(
                    title="Sélectionner la sauvegarde à restaurer",
                    filetypes=[("Base de données SQLite", "*.db"), ("Tous les fichiers", "*.*")]
                )
                
                if backup_file:
                    # Remplacer la base de données actuelle
                    shutil.copy2(backup_file, self.db_path)
                    messagebox.showinfo("Succès", "Base de données restaurée avec succès.\n\nL'application va redémarrer.")
                    
                    # Redémarrer l'application
                    self.root.quit()
                    
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la restauration: {str(e)}")
                
    def get_total_employees(self):
        """Obtenir le nombre total d'employés"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM employees')
            count = cursor.fetchone()[0]
            conn.close()
            return count
        except:
            return 0
            
    def get_total_users(self):
        """Obtenir le nombre total d'utilisateurs"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('SELECT COUNT(*) FROM users')
            count = cursor.fetchone()[0]
            conn.close()
            return count
        except:
            return 0
            
    def get_db_size(self):
        """Obtenir la taille de la base de données"""
        try:
            size = os.path.getsize(self.db_path)
            if size < 1024:
                return f"{size} bytes"
            elif size < 1024 * 1024:
                return f"{size / 1024:.1f} KB"
            else:
                return f"{size / (1024 * 1024):.1f} MB"
        except:
            return "Inconnue"
            
    def clear_main_content(self):
        """Nettoyer le contenu principal"""
        for widget in self.main_content.winfo_children():
            widget.destroy()
            
    def logout(self):
        """Déconnexion"""
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir vous déconnecter ?"):
            self.current_user = None
            self.current_employee_id = None
            self.show_login_screen()

    def start_move(self, event):
        """Enregistrer la position de départ du clic pour le déplacement"""
        self.x = event.x
        self.y = event.y

    def do_move(self, event):
        """Déplacer la fenêtre en fonction du mouvement de la souris"""
        deltax = event.x - self.x
        deltay = event.y - self.y
        x = self.root.winfo_x() + deltax
        y = self.root.winfo_y() + deltay
        self.root.geometry(f"+{x}+{y}")

    def close_app(self):
        """Fermer l'application proprement"""
        if messagebox.askokcancel("Quitter", "Êtes-vous sûr de vouloir quitter l'application ?"):
            self.root.destroy()

    def run(self):
        """Lancer l'application et la centrer correctement"""
        # Forcer la mise à jour de la fenêtre pour avoir les dimensions
        self.root.update_idletasks()
        
        # Obtenir la taille de l'écran
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Taille de la fenêtre (définie dans __init__)
        window_width = 1400
        window_height = 900
        
        # Calculer la position pour centrer
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        
        # Appliquer la géométrie
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # Démarrer la boucle principale
        self.root.mainloop()
    
    def show_ocr_module(self):
        """Affiche l'interface du module OCR pour l'extraction de texte."""
        self.clear_main_content()
        self.set_active_nav_button("✍️ OCR - Extraire Texte")

        # Titre
        title = tk.Label(self.main_content, text="✍️ OCR - Extraction de Texte d'Images et PDF",
                        font=('Segoe UI', 18, 'bold'), fg=self.colors['primary_green'], bg=self.colors['background'])
        title.pack(pady=(20, 20))

        # Conteneur principal
        main_frame = tk.Frame(self.main_content, bg=self.colors['background'])
        main_frame.pack(fill='both', expand=True, padx=20, pady=10)

        # Panneau de gauche pour la sélection et l'aperçu
        left_panel = tk.Frame(main_frame, bg=self.colors['white'], bd=1, relief='solid')
        left_panel.pack(side='left', fill='y', padx=(0, 10))

        tk.Label(left_panel, text="Source", font=('Segoe UI', 14, 'bold'), bg=self.colors['white']).pack(pady=10)

        load_btn = tk.Button(left_panel, text="📂 Charger Image ou PDF", font=('Segoe UI', 11, 'bold'),
                             bg=self.colors['primary_green'], fg='white', relief='flat',
                             command=self._select_and_process_file)
        load_btn.pack(pady=10, padx=10, fill='x')

        self.ocr_file_label = tk.Label(left_panel, text="Aucun fichier sélectionné", font=('Segoe UI', 10),
                                       bg=self.colors['white'], wraplength=280)
        self.ocr_file_label.pack(pady=5, padx=10)
        
        self.ocr_image_preview = tk.Label(left_panel, bg=self.colors['light_gray'])
        self.ocr_image_preview.pack(pady=10, padx=10, fill='both', expand=True)

        # Panneau de droite pour le résultat
        right_panel = tk.Frame(main_frame, bg=self.colors['white'], bd=1, relief='solid')
        right_panel.pack(side='right', fill='both', expand=True)

        result_toolbar = tk.Frame(right_panel, bg=self.colors['white'])
        result_toolbar.pack(fill='x', pady=5, padx=10)
        
        tk.Label(result_toolbar, text="Texte Extrait", font=('Segoe UI', 14, 'bold'), bg=self.colors['white']).pack(side='left')
        
        copy_btn = tk.Button(result_toolbar, text="📋 Copier", font=('Segoe UI', 10),
                             bg=self.colors['accent_green'], command=self._copy_text_to_clipboard)
        copy_btn.pack(side='right', padx=5)

        save_btn = tk.Button(result_toolbar, text="💾 Sauvegarder (.txt)", font=('Segoe UI', 10),
                             bg=self.colors['accent_green'], command=self._save_text_as_file)
        save_btn.pack(side='right')

        text_frame = tk.Frame(right_panel)
        text_frame.pack(fill='both', expand=True, padx=10, pady=(0, 10))

        self.ocr_result_text = tk.Text(text_frame, wrap='word', font=('Segoe UI', 11), relief='flat',
                                       fg=self.colors['text_dark'], bg=self.colors['light_gray'])
        
        text_scrollbar = ttk.Scrollbar(text_frame, orient='vertical', command=self.ocr_result_text.yview)
        self.ocr_result_text.configure(yscrollcommand=text_scrollbar.set)
        
        self.ocr_result_text.pack(side='left', fill='both', expand=True)
        text_scrollbar.pack(side='right', fill='y')

    def _select_and_process_file(self):
        """Ouvre une boîte de dialogue pour sélectionner un fichier et lance le traitement OCR."""
        file_path = filedialog.askopenfilename(
            title="Sélectionner une image ou un PDF",
            filetypes=[
                ("Fichiers Image", "*.png *.jpg *.jpeg *.bmp *.tiff"),
                ("Fichiers PDF", "*.pdf"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if not file_path:
            return

        self.ocr_file_label.config(text=os.path.basename(file_path))
        self.ocr_result_text.delete('1.0', tk.END)
        self.ocr_result_text.insert('1.0', "Traitement en cours, veuillez patienter...")
        self.root.update_idletasks()

        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext in ['.pdf']:
            self._process_pdf_ocr(file_path)
        else:
            self._process_image_ocr(file_path)

    def _process_image_ocr(self, file_path):
        """Traite une seule image avec Tesseract et affiche le résultat."""
        try:
            # Afficher un aperçu de l'image
            img = Image.open(file_path)
            img.thumbnail((280, 400)) # Redimensionne pour l'aperçu
            photo = ImageTk.PhotoImage(img)
            self.ocr_image_preview.config(image=photo)
            self.ocr_image_preview.image = photo

            # Extraire le texte
            extracted_text = pytesseract.image_to_string(Image.open(file_path), lang='fra')
            
            self.ocr_result_text.delete('1.0', tk.END)
            self.ocr_result_text.insert('1.0', extracted_text or "Aucun texte n'a pu être détecté.")

        except pytesseract.TesseractNotFoundError:
            messagebox.showerror("Erreur Tesseract",
                                 "Le programme Tesseract n'est pas installé ou n'est pas dans le PATH.\n"
                                 "Veuillez l'installer et vérifier le chemin dans le code.")
        except Exception as e:
            messagebox.showerror("Erreur de Traitement", f"Une erreur est survenue: {e}")
            self.ocr_result_text.delete('1.0', tk.END)

    def _process_pdf_ocr(self, file_path):
        """Convertit un PDF en images, puis traite chaque image avec Tesseract."""
        self.ocr_image_preview.config(image=None, text="Aperçu non\ndisponible\npour les PDF")
        self.ocr_image_preview.image = None
        
        try:
            # --- CHEMIN MIS À JOUR AVEC VOTRE VERSION ---
            poppler_path = r"C:\poppler-24.08.0\Library\bin"
            
            images = convert_from_path(file_path, poppler_path=poppler_path)
            full_text = ""
            
            for i, image in enumerate(images):
                self.ocr_result_text.delete('1.0', tk.END)
                self.ocr_result_text.insert('1.0', f"Traitement de la page {i+1}/{len(images)}...")
                self.root.update_idletasks()
                
                text = pytesseract.image_to_string(image, lang='fra')
                full_text += f"--- PAGE {i+1} ---\n{text}\n\n"

            self.ocr_result_text.delete('1.0', tk.END)
            self.ocr_result_text.insert('1.0', full_text or "Aucun texte n'a pu être détecté dans le PDF.")

        except Exception as e:
            messagebox.showerror("Erreur de Traitement PDF",
                                 "Impossible de traiter le PDF. Assurez-vous que Poppler est bien installé et que le chemin est correct.\n\n"
                                 f"Détail de l'erreur: {e}")
            self.ocr_result_text.delete('1.0', tk.END)

    def _copy_text_to_clipboard(self):
        """Copie le texte extrait dans le presse-papiers."""
        text = self.ocr_result_text.get('1.0', tk.END).strip()
        if text:
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("Copié", "Le texte a été copié dans le presse-papiers.")
        else:
            messagebox.showwarning("Attention", "Aucun texte à copier.")

    def _save_text_as_file(self):
        """Sauvegarde le texte extrait dans un fichier .txt."""
        text = self.ocr_result_text.get('1.0', tk.END).strip()
        if not text:
            messagebox.showwarning("Attention", "Aucun texte à sauvegarder.")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Fichiers Texte", "*.txt"), ("Tous les fichiers", "*.*")],
            title="Sauvegarder le texte extrait"
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(text)
                messagebox.showinfo("Succès", f"Fichier sauvegardé avec succès:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Erreur de Sauvegarde", f"Impossible de sauvegarder le fichier: {e}")
    
    def logout(self):
        """Déconnexion"""
        if messagebox.askyesno("Confirmation", "Êtes-vous sûr de vouloir vous déconnecter ?"):
            self.current_user = None
            self.current_employee_id = None
            self.show_login_screen()

    def run(self):
        """Lancer l'application et la centrer correctement"""
        # Forcer la mise à jour de la fenêtre pour avoir les dimensions
        self.root.update_idletasks()
        
        # Obtenir la taille de l'écran
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Taille de la fenêtre (définie dans __init__)
        window_width = 1400
        window_height = 900
        
        # Calculer la position pour centrer
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        
        # Appliquer la géométrie
        self.root.geometry(f'{window_width}x{window_height}+{x}+{y}')
        
        # Démarrer la boucle principale
        self.root.mainloop()

# Point d'entrée de l'application
if __name__ == "__main__":
    try:
        app = HRManagementApp()
        app.run()
    except Exception as e:
        print(f"Erreur lors du démarrage de l'application: {str(e)}")
        input("Appuyez sur Entrée pour fermer...")